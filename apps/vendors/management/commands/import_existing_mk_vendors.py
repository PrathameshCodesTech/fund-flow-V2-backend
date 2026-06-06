from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.core.models import Organization, ScopeNode
from apps.users.services import send_password_reset_for_user
from apps.vendors.models import (
    MarketingStatus,
    OperationalStatus,
    UserVendorAssignment,
    Vendor,
)


STATE_BY_REGION_CODE = {
    "1": "Jammu and Kashmir",
    "2": "Himachal Pradesh",
    "3": "Punjab",
    "4": "Chandigarh",
    "5": "Uttarakhand",
    "6": "Haryana",
    "7": "Delhi",
    "8": "Rajasthan",
    "9": "Uttar Pradesh",
    "10": "Bihar",
    "11": "Sikkim",
    "12": "Arunachal Pradesh",
    "13": "Nagaland",
    "14": "Manipur",
    "15": "Mizoram",
    "16": "Tripura",
    "17": "Meghalaya",
    "18": "Assam",
    "19": "West Bengal",
    "20": "Jharkhand",
    "21": "Odisha",
    "22": "Chhattisgarh",
    "23": "Madhya Pradesh",
    "24": "Gujarat",
    "25": "Daman and Diu",
    "26": "Dadra and Nagar Haveli",
    "27": "Maharashtra",
    "28": "Andhra Pradesh",
    "29": "Karnataka",
    "30": "Goa",
    "31": "Lakshadweep",
    "32": "Kerala",
    "33": "Tamil Nadu",
    "34": "Puducherry",
    "35": "Andaman and Nicobar Islands",
    "36": "Telangana",
    "37": "Andhra Pradesh",
}


EMAIL_COLUMNS = [
    "Email1-Communication",
    "Email1-Add. Independent",
    "Email2-Communication",
    "Email2-Add. Independent",
    "Email3-Communication",
    "Email3-Add. Independent",
    "Email4-Communication",
    "Email4-Add. Independent",
]


@dataclass
class ImportStats:
    rows_seen: int = 0
    vendors_created: int = 0
    vendors_updated: int = 0
    users_created: int = 0
    users_updated: int = 0
    assignments_created: int = 0
    assignments_updated: int = 0
    reset_emails_sent: int = 0
    skipped_no_email: int = 0
    skipped_missing_name_or_sap: int = 0


class Command(BaseCommand):
    help = (
        "Import existing MK vendors from the SAP-style vendor workbook. "
        "By default this is a dry run and sends no email."
    )

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Path to MK Vendors.xlsx")
        parser.add_argument("--org-code", default="horizon", help="Organization code. Default: horizon")
        parser.add_argument("--scope", default="Marketing", help="Scope node name. Default: Marketing")
        parser.add_argument(
            "--commit",
            action="store_true",
            help="Actually write changes. Without this flag, the command only reports what would happen.",
        )
        parser.add_argument(
            "--create-users",
            action="store_true",
            help="Create/update vendor portal users for vendors that have an email.",
        )
        parser.add_argument(
            "--send-password-reset",
            action="store_true",
            help="Send password reset email to vendor users after creating/updating the user assignment.",
        )
        parser.add_argument(
            "--requested-by-email",
            default="",
            help="Optional admin email to use as the password reset requester/audit actor.",
        )
        parser.add_argument(
            "--default-payment-mode",
            default="",
            help="Optional value to seed preferred payment mode, for example NEFT. Blank by default.",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file"]).expanduser()
        if not file_path.exists():
            raise CommandError(f"File does not exist: {file_path}")

        commit = bool(options["commit"])
        create_users = bool(options["create_users"])
        send_resets = bool(options["send_password_reset"])
        if send_resets and not create_users:
            raise CommandError("--send-password-reset requires --create-users")

        org = Organization.objects.get(code=options["org_code"])
        scope_node = ScopeNode.objects.get(org=org, name__iexact=options["scope"], is_active=True)
        requested_by = self._get_requested_by(options["requested_by_email"])

        rows = self._load_rows(file_path)
        stats = ImportStats(rows_seen=len(rows))

        if not commit:
            self.stdout.write(self.style.WARNING("DRY RUN: no database changes and no emails will be sent."))

        def run_import():
            for row_number, row in rows:
                self._process_row(
                    row_number=row_number,
                    row=row,
                    org=org,
                    scope_node=scope_node,
                    create_users=create_users,
                    send_resets=send_resets,
                    requested_by=requested_by,
                    default_payment_mode=options["default_payment_mode"].strip(),
                    commit=commit,
                    stats=stats,
                )

        if commit:
            with transaction.atomic():
                run_import()
        else:
            run_import()

        self._print_summary(stats, commit=commit, create_users=create_users, send_resets=send_resets)

    def _load_rows(self, file_path: Path) -> list[tuple[int, dict[str, object]]]:
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError("openpyxl is required. Install project requirements first.") from exc

        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        raw_headers = [self._clean_cell(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        headers = self._dedupe_headers(raw_headers)

        rows = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(self._clean_cell(value) for value in values):
                continue
            rows.append((row_number, dict(zip(headers, values))))
        return rows

    def _process_row(
        self,
        *,
        row_number: int,
        row: dict[str, object],
        org: Organization,
        scope_node: ScopeNode,
        create_users: bool,
        send_resets: bool,
        requested_by,
        default_payment_mode: str,
        commit: bool,
        stats: ImportStats,
    ) -> None:
        payload = self._build_vendor_payload(row, org=org, scope_node=scope_node, default_payment_mode=default_payment_mode)
        vendor_name = payload["vendor_name"]
        sap_vendor_id = payload["sap_vendor_id"]
        if not vendor_name or not sap_vendor_id:
            stats.skipped_missing_name_or_sap += 1
            self.stdout.write(f"SKIP row {row_number}: missing vendor name or SAP vendor id")
            return

        email = payload["email"]
        vendor = None
        vendor_created = False
        if commit:
            vendor, vendor_created = Vendor.objects.update_or_create(
                org=org,
                sap_vendor_id=sap_vendor_id,
                defaults=payload,
            )
        else:
            vendor_created = not Vendor.objects.filter(org=org, sap_vendor_id=sap_vendor_id).exists()

        if vendor_created:
            stats.vendors_created += 1
        else:
            stats.vendors_updated += 1

        if not email:
            stats.skipped_no_email += 1
            self.stdout.write(f"NO EMAIL row {row_number}: {vendor_name} ({sap_vendor_id}) - vendor only")
            return

        if not create_users:
            return

        user_created = False
        assignment_created = False
        if commit:
            user, user_created = self._ensure_user(email=email, vendor_name=vendor_name)
            assignment, assignment_created = UserVendorAssignment.objects.update_or_create(
                user=user,
                vendor=vendor,
                defaults={"is_active": True},
            )
            vendor.portal_email = email
            vendor.portal_user_id = str(user.pk)
            vendor.save(update_fields=["portal_email", "portal_user_id", "updated_at"])

            if send_resets:
                send_password_reset_for_user(target_user=user, requested_by=requested_by)
                stats.reset_emails_sent += 1
        else:
            User = get_user_model()
            user_created = not User.objects.filter(email__iexact=email).exists()
            assignment_created = False

        if user_created:
            stats.users_created += 1
        else:
            stats.users_updated += 1

        if assignment_created:
            stats.assignments_created += 1
        else:
            stats.assignments_updated += 1

    def _build_vendor_payload(
        self,
        row: dict[str, object],
        *,
        org: Organization,
        scope_node: ScopeNode,
        default_payment_mode: str,
    ) -> dict[str, object]:
        vendor_name = self._first_value(row, "Name", "Name 2")
        sap_vendor_id = self._first_value(row, "Business Partner")
        email = self._first_email(row)
        phone = self._first_value(row, "Telephone")
        gstin = self._first_value(row, "Tax Number 3").upper()
        pan = self._first_value(row, "Permanent account number").upper()
        if not pan and len(gstin) >= 12:
            pan = gstin[2:12]

        street_1 = self._first_value(row, "Street")
        street_2 = self._first_value(row, "Street 2")
        street_3 = self._first_value(row, "Street 3")
        address_line1 = street_1 or street_2 or street_3
        address_line2 = street_2 if address_line1 != street_2 else ""
        address_line3 = street_3 if address_line1 != street_3 else ""
        region_code = self._first_value(row, "Region")
        state = STATE_BY_REGION_CODE.get(region_code.lstrip("0"), "")
        country = self._country_name(self._first_value(row, "Country/Region Key"))

        msme_number = self._first_value(row, "MSME No.")
        account_number = self._first_value(row, "Account No:Bank1")
        beneficiary_name = self._first_value(row, "Acc. Holder:Bank1") or vendor_name
        ifsc = self._first_value(row, "IFSC:Bank1").upper()

        return {
            "org": org,
            "scope_node": scope_node,
            "vendor_name": vendor_name,
            "email": email,
            "phone": phone,
            "title": "",
            "vendor_type": "Organisation",
            "fax": "",
            "region": "",
            "head_office_no": "",
            "gst_registered": bool(gstin),
            "gstin": gstin,
            "pan": pan,
            "address_line1": address_line1,
            "address_line2": address_line2,
            "address_line3": address_line3,
            "city": self._first_value(row, "City"),
            "state": state,
            "country": country,
            "pincode": self._first_value(row, "Postal Code"),
            "preferred_payment_mode": default_payment_mode,
            "beneficiary_name": beneficiary_name,
            "beneficiary_account_number": account_number,
            "bank_name": "",
            "bank_address": "",
            "bank_email": "",
            "account_number": account_number,
            "bank_account_number": account_number,
            "bank_account_type": "",
            "ifsc": ifsc,
            "micr_code": "",
            "neft_code": ifsc,
            "bank_branch_address_line1": "",
            "bank_branch_address_line2": "",
            "bank_branch_city": "",
            "bank_branch_state": "",
            "bank_branch_country": "",
            "bank_branch_pincode": "",
            "bank_phone": "",
            "bank_fax": "",
            "authorized_signatory_name": "",
            "msme_registered": bool(msme_number),
            "msme_registration_number": msme_number,
            "msme_enterprise_type": "",
            "declaration_accepted": None,
            "contact_persons_json": [],
            "head_office_address_json": {},
            "tax_registration_details_json": {
                "tin_no": "",
                "cst_no": "",
                "lst_no": "",
                "esic_reg_no": "",
                "pan_ref_no": "",
                "ppf_no": "",
            },
            "sap_vendor_id": sap_vendor_id,
            "po_mandate_enabled": False,
            "marketing_status": MarketingStatus.APPROVED,
            "operational_status": OperationalStatus.ACTIVE,
            "portal_email": email,
        }

    def _ensure_user(self, *, email: str, vendor_name: str):
        User = get_user_model()
        first_name, last_name = self._split_name(vendor_name)
        user, created = User.objects.get_or_create(
            email=email,
            defaults={
                "first_name": first_name,
                "last_name": last_name,
                "is_active": True,
                "is_staff": False,
                "is_superuser": False,
            },
        )
        changed_fields = []
        if not user.first_name:
            user.first_name = first_name
            changed_fields.append("first_name")
        if not user.last_name:
            user.last_name = last_name
            changed_fields.append("last_name")
        if not user.is_active:
            user.is_active = True
            changed_fields.append("is_active")
        if user.is_staff:
            user.is_staff = False
            changed_fields.append("is_staff")
        if user.is_superuser:
            user.is_superuser = False
            changed_fields.append("is_superuser")
        if changed_fields:
            user.save(update_fields=changed_fields + ["updated_at"])
        return user, created

    def _get_requested_by(self, email: str):
        if not email:
            return None
        User = get_user_model()
        try:
            return User.objects.get(email__iexact=email)
        except User.DoesNotExist as exc:
            raise CommandError(f"--requested-by-email user not found: {email}") from exc

    def _first_email(self, row: dict[str, object]) -> str:
        for column in EMAIL_COLUMNS:
            value = self._first_value(row, column).lower()
            if value and self._looks_like_email(value):
                return value
        return ""

    def _first_value(self, row: dict[str, object], *columns: str) -> str:
        for column in columns:
            value = self._clean_cell(row.get(column))
            if value:
                return value
        return ""

    def _clean_cell(self, value) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _dedupe_headers(self, headers: list[str]) -> list[str]:
        seen = {}
        result = []
        for header in headers:
            name = header or "Unnamed"
            count = seen.get(name, 0)
            seen[name] = count + 1
            result.append(name if count == 0 else f"{name}__{count + 1}")
        return result

    def _looks_like_email(self, value: str) -> bool:
        return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", value))

    def _country_name(self, value: str) -> str:
        if not value or value.upper() == "IN":
            return "India"
        return value

    def _split_name(self, vendor_name: str) -> tuple[str, str]:
        parts = vendor_name.strip().split()
        if not parts:
            return "Vendor", ""
        if len(parts) == 1:
            return parts[0][:150], ""
        return parts[0][:150], " ".join(parts[1:])[:150]

    def _print_summary(self, stats: ImportStats, *, commit: bool, create_users: bool, send_resets: bool) -> None:
        self.stdout.write("")
        self.stdout.write("MK vendor import summary")
        self.stdout.write(f"Mode: {'COMMIT' if commit else 'DRY RUN'}")
        self.stdout.write(f"Create users: {create_users}")
        self.stdout.write(f"Send password reset: {send_resets}")
        self.stdout.write(f"Rows seen: {stats.rows_seen}")
        self.stdout.write(f"Vendors created: {stats.vendors_created}")
        self.stdout.write(f"Vendors updated: {stats.vendors_updated}")
        self.stdout.write(f"Users created: {stats.users_created}")
        self.stdout.write(f"Users updated/reused: {stats.users_updated}")
        self.stdout.write(f"Assignments created: {stats.assignments_created}")
        self.stdout.write(f"Assignments updated/reused: {stats.assignments_updated}")
        self.stdout.write(f"Password reset emails sent: {stats.reset_emails_sent}")
        self.stdout.write(f"Skipped no email: {stats.skipped_no_email}")
        self.stdout.write(f"Skipped missing name/SAP id: {stats.skipped_missing_name_or_sap}")

        if not commit:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("Run again with --commit to write these changes."))
