"""
Vendor domain services.

All business logic for the vendor onboarding and activation lifecycle lives here.
No view-layer imports.  Django mail is called indirectly via apps.vendors.email
so tests can mock at a single point.
"""
import io
import logging
import os
import re
import secrets
import zipfile
from datetime import timedelta
from pathlib import Path
from xml.etree import ElementTree

from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import transaction
from django.utils import timezone

from apps.vendors.models import (
    FinanceActionType,
    FinanceDecisionChoice,
    InvitationStatus,
    MarketingStatus,
    OperationalStatus,
    SubmissionMode,
    SubmissionStatus,
    UserVendorAssignment,
    Vendor,
    VendorAttachment,
    VendorActivationToken,
    VendorFinanceActionToken,
    VendorFinanceDecision,
    VendorInvitation,
    VendorOnboardingSubmission,
    VendorProfileRevision,
    VendorProfileRevisionStatus,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Custom exceptions
# ---------------------------------------------------------------------------

class InvitationExpiredError(ValueError):
    """Invitation token has passed its expiry."""


class InvitationNotFoundError(ValueError):
    """No invitation found for this token (or it was cancelled)."""


class SubmissionStateError(ValueError):
    """Submission is in the wrong state for the requested operation."""


class VendorStateError(ValueError):
    """Vendor is in the wrong state for the requested operation."""


class FinanceTokenError(ValueError):
    """Finance action token is invalid, expired, or already used."""


class POMandate(ValueError):
    """Legacy exception kept for compatibility; PO numbers are optional."""


_EDITABLE_SUBMISSION_STATUSES = {
    SubmissionStatus.DRAFT,
    SubmissionStatus.REOPENED,
    SubmissionStatus.FINANCE_REJECTED,
}


# ---------------------------------------------------------------------------
# VRF field mappings (label → normalized field name)
# ---------------------------------------------------------------------------

_VRF_LABEL_MAP = {
    # Core identity
    "vendor name": "vendor_name",
    "vendor_name": "vendor_name",
    "title": "title",
    "vendor type": "vendor_type",
    "type of vendor": "vendor_type",
    "vendor_type": "vendor_type",
    "gst registered": "gst_registered",
    "gst_registered": "gst_registered",
    "gst registration": "gst_registered",
    "gstin": "gstin",
    "gstin number": "gstin",
    "pan": "pan",
    "pan no": "pan",
    "pan no.": "pan",
    "email": "email",
    "email id": "email",
    "e-mail": "email",
    "phone": "phone",
    "phone no": "phone",
    "phone no.": "phone",
    "mobile": "phone",
    "fax": "fax",
    "fax no": "fax",
    "fax no.": "fax",
    "region": "region",
    "head office no": "head_office_no",
    "head office / site no.": "head_office_no",
    "head office / site no": "head_office_no",
    # Address
    "address line 1": "address_line1",
    "address line1": "address_line1",
    "address line 2": "address_line2",
    "address line2": "address_line2",
    "address line 3": "address_line3",
    "address line3": "address_line3",
    "city": "city",
    "state": "state",
    "country": "country",
    "pincode": "pincode",
    "pin code": "pincode",
    "zip code": "pincode",
    # Bank — core
    "bank name": "bank_name",
    "bank": "bank_name",
    "bank address": "bank_address",
    "bank email": "bank_email",
    "bank email address": "bank_email",
    "bank e-mail": "bank_email",
    "bank e-mail address": "bank_email",
    "beneficiary name": "beneficiary_name",
    "beneficiary account no": "beneficiary_account_number",
    "beneficiary account no.": "beneficiary_account_number",
    "beneficiary account number": "beneficiary_account_number",
    "account number": "account_number",
    "account no": "account_number",
    "bank account no": "bank_account_number",
    "bank account no.": "bank_account_number",
    "bank account number": "bank_account_number",
    "account type": "bank_account_type",
    "bank account type": "bank_account_type",
    "ifsc": "ifsc",
    "ifsc code": "ifsc",
    "preferred payment mode": "preferred_payment_mode",
    "preffered payment mode": "preferred_payment_mode",
    "payment mode": "preferred_payment_mode",
    "bank micr code": "micr_code",
    "micr code": "micr_code",
    "micr": "micr_code",
    "neft code": "neft_code",
    # Bank — branch contact
    "bank branch address line 1": "bank_branch_address_line1",
    "bank branch address line1": "bank_branch_address_line1",
    "branch address line 1": "bank_branch_address_line1",
    "branch address line1": "bank_branch_address_line1",
    "bank branch address line 2": "bank_branch_address_line2",
    "bank branch address line2": "bank_branch_address_line2",
    "branch address line 2": "bank_branch_address_line2",
    "branch address line2": "bank_branch_address_line2",
    "bank branch city": "bank_branch_city",
    "branch city": "bank_branch_city",
    "bank branch state": "bank_branch_state",
    "branch state": "bank_branch_state",
    "bank branch country": "bank_branch_country",
    "branch country": "bank_branch_country",
    "bank branch pincode": "bank_branch_pincode",
    "branch pincode": "bank_branch_pincode",
    "bank phone": "bank_phone",
    "branch phone": "bank_phone",
    "bank fax": "bank_fax",
    "branch fax": "bank_fax",
    # MSME / compliance
    "msme registered": "msme_registered",
    "msme registration number": "msme_registration_number",
    "msme_registration_number": "msme_registration_number",
    "udyam registration no": "msme_registration_number",
    "udyam_reg_no": "msme_registration_number",
    "enterprise type": "msme_enterprise_type",
    "enterprise_type": "msme_enterprise_type",
    "msme_enterprise_type": "msme_enterprise_type",
    "authorized signatory name": "authorized_signatory_name",
    "authorized_signatory_name": "authorized_signatory_name",
    "declaration accepted": "declaration_accepted",
    "declaration_accepted": "declaration_accepted",
}

_HEAD_OFFICE_LABEL_MAP = {
    "head office address line 1": "address_line1",
    "head office address line1": "address_line1",
    "address line 1": "address_line1",
    "address line1": "address_line1",
    "head office address line 2": "address_line2",
    "head office address line2": "address_line2",
    "address line 2": "address_line2",
    "address line2": "address_line2",
    "head office city": "city",
    "city": "city",
    "head office state": "state",
    "state": "state",
    "head office country": "country",
    "country": "country",
    "head office pincode": "pincode",
    "pincode": "pincode",
    "pin code": "pincode",
    "head office phone": "phone",
    "phone no": "phone",
    "phone no.": "phone",
    "head office fax": "fax",
    "fax no": "fax",
    "fax no.": "fax",
}

_TAX_REGISTRATION_LABEL_MAP = {
    "tax registration nos.": "tax_registration_nos",
    "tax registration nos": "tax_registration_nos",
    "tin no.": "tin_no",
    "tin no": "tin_no",
    "cst no.": "cst_no",
    "cst no": "cst_no",
    "lst no.": "lst_no",
    "lst no": "lst_no",
    "esic reg. no.": "esic_reg_no",
    "esic reg. no": "esic_reg_no",
    "esic reg no": "esic_reg_no",
    "pan ref. no.": "pan_ref_no",
    "pan ref. no": "pan_ref_no",
    "pan ref no": "pan_ref_no",
    "ppf no.": "ppf_no",
    "ppf no": "ppf_no",
}

_CONTACT_PERSON_LABEL_MAP = {
    "contact 1 name": (0, "name"),
    "1) name": (0, "name"),
    "1 name": (0, "name"),
    "contact 1 designation": (0, "designation"),
    "1) designation": (0, "designation"),
    "1 designation": (0, "designation"),
    "contact 1 email": (0, "email"),
    "1) email address": (0, "email"),
    "1 email address": (0, "email"),
    "contact 1 telephone": (0, "telephone"),
    "1) telephone": (0, "telephone"),
    "1 telephone": (0, "telephone"),
    "contact 2 name": (1, "name"),
    "2) name": (1, "name"),
    "2 name": (1, "name"),
    "contact 2 designation": (1, "designation"),
    "2) designation": (1, "designation"),
    "2 designation": (1, "designation"),
    "contact 2 email": (1, "email"),
    "2) email address": (1, "email"),
    "2 email address": (1, "email"),
    "contact 2 telephone": (1, "telephone"),
    "2) telephone": (1, "telephone"),
    "2 telephone": (1, "telephone"),
}

_BANK_DETAILS_LABEL_MAP = {
    "address line": "bank_branch_address_line1",
    "address line 1": "bank_branch_address_line1",
    "address line1": "bank_branch_address_line1",
    "addess line2": "bank_branch_address_line2",
    "address line 2": "bank_branch_address_line2",
    "address line2": "bank_branch_address_line2",
    "city": "bank_branch_city",
    "pin code": "bank_branch_pincode",
    "pincode": "bank_branch_pincode",
    "state": "bank_branch_state",
    "country": "bank_branch_country",
    "phone no": "bank_phone",
    "phone no.": "bank_phone",
    "fax no": "bank_fax",
    "fax no.": "bank_fax",
    "email address": "bank_email",
}

_KNOWN_KEYS = set(_VRF_LABEL_MAP.values())

# Keys that map directly from normalized name to model field
_NORM_FIELD_MAP = {
    # Core identity
    "title": "normalized_title",
    "vendor_name": "normalized_vendor_name",
    "vendor_type": "normalized_vendor_type",
    "email": "normalized_email",
    "phone": "normalized_phone",
    "fax": "normalized_fax",
    "region": "normalized_region",
    "head_office_no": "normalized_head_office_no",
    "gst_registered": "normalized_gst_registered",
    "gstin": "normalized_gstin",
    "pan": "normalized_pan",
    # Address
    "address_line1": "normalized_address_line1",
    "address_line2": "normalized_address_line2",
    "address_line3": "normalized_address_line3",
    "city": "normalized_city",
    "state": "normalized_state",
    "country": "normalized_country",
    "pincode": "normalized_pincode",
    # Bank core
    "preferred_payment_mode": "normalized_preferred_payment_mode",
    "beneficiary_name": "normalized_beneficiary_name",
    "beneficiary_account_number": "normalized_beneficiary_account_number",
    "bank_name": "normalized_bank_name",
    "bank_address": "normalized_bank_address",
    "bank_email": "normalized_bank_email",
    "account_number": "normalized_account_number",
    "bank_account_number": "normalized_bank_account_number",
    "bank_account_type": "normalized_bank_account_type",
    "ifsc": "normalized_ifsc",
    "micr_code": "normalized_micr_code",
    "neft_code": "normalized_neft_code",
    # Bank branch contact
    "bank_branch_address_line1": "normalized_bank_branch_address_line1",
    "bank_branch_address_line2": "normalized_bank_branch_address_line2",
    "bank_branch_city": "normalized_bank_branch_city",
    "bank_branch_state": "normalized_bank_branch_state",
    "bank_branch_country": "normalized_bank_branch_country",
    "bank_branch_pincode": "normalized_bank_branch_pincode",
    "bank_phone": "normalized_bank_phone",
    "bank_fax": "normalized_bank_fax",
    # MSME / compliance
    "authorized_signatory_name": "normalized_authorized_signatory_name",
    "msme_registered": "normalized_msme_registered",
    "msme_registration_number": "normalized_msme_registration_number",
    "msme_enterprise_type": "normalized_msme_enterprise_type",
    "declaration_accepted": "declaration_accepted",
}

# Snapshot field names — Vendor field names that participate in profile revisions.
# Key = snapshot dict key (also the field label). Value = Vendor model field name.
_VENDOR_PROFILE_SNAPSHOT_FIELDS = {
    "vendor_name": "vendor_name",
    "email": "email",
    "phone": "phone",
    "title": "title",
    "vendor_type": "vendor_type",
    "fax": "fax",
    "region": "region",
    "head_office_no": "head_office_no",
    "gst_registered": "gst_registered",
    "gstin": "gstin",
    "pan": "pan",
    "address_line1": "address_line1",
    "address_line2": "address_line2",
    "address_line3": "address_line3",
    "city": "city",
    "state": "state",
    "country": "country",
    "pincode": "pincode",
    "preferred_payment_mode": "preferred_payment_mode",
    "beneficiary_name": "beneficiary_name",
    "beneficiary_account_number": "beneficiary_account_number",
    "bank_name": "bank_name",
    "bank_address": "bank_address",
    "bank_email": "bank_email",
    "account_number": "account_number",
    "bank_account_number": "bank_account_number",
    "bank_account_type": "bank_account_type",
    "ifsc": "ifsc",
    "micr_code": "micr_code",
    "neft_code": "neft_code",
    "bank_branch_address_line1": "bank_branch_address_line1",
    "bank_branch_address_line2": "bank_branch_address_line2",
    "bank_branch_city": "bank_branch_city",
    "bank_branch_state": "bank_branch_state",
    "bank_branch_country": "bank_branch_country",
    "bank_branch_pincode": "bank_branch_pincode",
    "bank_phone": "bank_phone",
    "bank_fax": "bank_fax",
    "authorized_signatory_name": "authorized_signatory_name",
    "msme_registered": "msme_registered",
    "msme_registration_number": "msme_registration_number",
    "msme_enterprise_type": "msme_enterprise_type",
    "declaration_accepted": "declaration_accepted",
    "contact_persons_json": "contact_persons_json",
    "head_office_address_json": "head_office_address_json",
    "tax_registration_details_json": "tax_registration_details_json",
}

_SUBMISSION_TO_VENDOR_FIELD_MAP = {
    "normalized_title": "title",
    "normalized_vendor_name": "vendor_name",
    "normalized_vendor_type": "vendor_type",
    "normalized_email": "email",
    "normalized_phone": "phone",
    "normalized_fax": "fax",
    "normalized_region": "region",
    "normalized_head_office_no": "head_office_no",
    "normalized_gst_registered": "gst_registered",
    "normalized_gstin": "gstin",
    "normalized_pan": "pan",
    "normalized_address_line1": "address_line1",
    "normalized_address_line2": "address_line2",
    "normalized_address_line3": "address_line3",
    "normalized_city": "city",
    "normalized_state": "state",
    "normalized_country": "country",
    "normalized_pincode": "pincode",
    "normalized_preferred_payment_mode": "preferred_payment_mode",
    "normalized_beneficiary_name": "beneficiary_name",
    "normalized_beneficiary_account_number": "beneficiary_account_number",
    "normalized_bank_name": "bank_name",
    "normalized_bank_address": "bank_address",
    "normalized_bank_email": "bank_email",
    "normalized_account_number": "account_number",
    "normalized_bank_account_number": "bank_account_number",
    "normalized_bank_account_type": "bank_account_type",
    "normalized_ifsc": "ifsc",
    "normalized_micr_code": "micr_code",
    "normalized_neft_code": "neft_code",
    "normalized_bank_branch_address_line1": "bank_branch_address_line1",
    "normalized_bank_branch_address_line2": "bank_branch_address_line2",
    "normalized_bank_branch_city": "bank_branch_city",
    "normalized_bank_branch_state": "bank_branch_state",
    "normalized_bank_branch_country": "bank_branch_country",
    "normalized_bank_branch_pincode": "bank_branch_pincode",
    "normalized_bank_phone": "bank_phone",
    "normalized_bank_fax": "bank_fax",
    "normalized_authorized_signatory_name": "authorized_signatory_name",
    "normalized_msme_registered": "msme_registered",
    "normalized_msme_registration_number": "msme_registration_number",
    "normalized_msme_enterprise_type": "msme_enterprise_type",
    "declaration_accepted": "declaration_accepted",
}


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

def _start_finance_review(submission: VendorOnboardingSubmission) -> None:
    """
    Authoritative business transition: move a submission into active finance review.

    Does (in order):
      1. Generate canonical VRF Excel workbook — REQUIRED; raises if fails
      2. Mark submission as SENT_TO_FINANCE with tokens — atomic
      3. Send finance handoff email with workbook — REQUIRED; raises if fails
         (whole transaction rolls back if email cannot be sent with workbook)
      4. Internal audit log notification — non-fatal

    Call within @transaction.atomic.
    """
    from apps.vendors.notifications import (
        send_finance_handoff_notification,
        notify_internal_submission_received,
    )

    # ── 1. Generate canonical VRF Excel (REQUIRED — must succeed before state change) ─
    try:
        generate_vendor_export_excel(submission)
    except Exception as exc:
        raise SubmissionStateError(
            f"Cannot complete finance handoff: VRF workbook generation failed — {exc}"
        ) from exc

    now = timezone.now()

    # ── 2. Authoritative state transition + tokens ────────────────────────────
    submission.status = SubmissionStatus.SENT_TO_FINANCE
    submission.submitted_at = now
    submission.finance_sent_at = now
    submission.save(update_fields=["status", "submitted_at", "finance_sent_at", "updated_at"])

    expiry_hours = getattr(settings, "VENDOR_FINANCE_TOKEN_EXPIRY_HOURS", 72)
    expires_at = now + timedelta(hours=expiry_hours)

    VendorFinanceActionToken.objects.create(
        submission=submission,
        action_type=FinanceActionType.APPROVE,
        token=_generate_token(),
        expires_at=expires_at,
    )
    VendorFinanceActionToken.objects.create(
        submission=submission,
        action_type=FinanceActionType.REJECT,
        token=_generate_token(),
        expires_at=expires_at,
    )

    # ── 3. Finance handoff email (REQUIRED — raises on failure, rolls back transaction) ─
    send_finance_handoff_notification(submission)

    # ── 4. Internal audit log notification (non-fatal) ────────────────────────
    try:
        notify_internal_submission_received(submission)
    except Exception as exc:
        _logger.warning(
            "Internal submission-received notification failed for submission_id=%s: %s",
            submission.pk, exc,
        )


def _generate_token() -> str:
    return secrets.token_urlsafe(48)


def _apply_normalized_fields(submission: VendorOnboardingSubmission, normalized: dict) -> None:
    """Write normalized dict values onto submission model fields."""
    for key, field in _NORM_FIELD_MAP.items():
        if key in normalized:
            value = normalized[key]
            if key == "gst_registered":
                if isinstance(value, bool):
                    setattr(submission, field, value)
                elif isinstance(value, str):
                    setattr(submission, field, value.strip().lower() in ("yes", "true", "1", "y"))
                else:
                    setattr(submission, field, bool(value))
            else:
                setattr(submission, field, str(value).strip() if value is not None else "")


def _extract_normalized_from_payload(payload: dict) -> tuple[dict, dict, dict, dict, dict]:
    """
    Split payload into normalized core fields, remaining raw data,
    and JSON blocks (contact_persons, head_office_address, tax_registration_details).

    Returns (normalized_dict, raw_dict, contact_persons, head_office_address, tax_registration_details).
    """
    normalized = {}
    raw = {}
    contact_persons = None
    head_office_address = None
    tax_registration_details = None

    JSON_BLOCK_KEYS = {"contact_persons", "head_office_address", "tax_registration_details"}

    for key, value in payload.items():
        lower_key = key.strip().lower()
        mapped = _VRF_LABEL_MAP.get(lower_key, lower_key.replace(" ", "_"))
        if mapped in JSON_BLOCK_KEYS and isinstance(value, (dict, list)):
            if mapped == "contact_persons":
                contact_persons = value if isinstance(value, list) else [value]
            elif mapped == "head_office_address":
                head_office_address = value if isinstance(value, dict) else {}
            elif mapped == "tax_registration_details":
                tax_registration_details = value if isinstance(value, dict) else {}
        elif mapped in _KNOWN_KEYS:
            normalized[mapped] = value
        else:
            raw[key] = value

    return normalized, raw, contact_persons, head_office_address, tax_registration_details


def _apply_normalized_fields(submission: VendorOnboardingSubmission, normalized: dict) -> None:
    """Write normalized dict values onto submission model fields."""
    from apps.vendors.models import ALLOWED_MSME_ENTERPRISE_TYPES
    for key, field in _NORM_FIELD_MAP.items():
        if key in normalized:
            value = normalized[key]
            if key in ("gst_registered", "msme_registered", "declaration_accepted"):
                if isinstance(value, bool):
                    pass  # use as-is
                elif isinstance(value, str):
                    value = value.strip().lower() in ("yes", "true", "1", "y")
                else:
                    value = bool(value)
            elif key == "msme_enterprise_type":
                normalized_value = str(value).strip().lower()
                if normalized_value and normalized_value not in ALLOWED_MSME_ENTERPRISE_TYPES:
                    raise ValueError(
                        f"msme_enterprise_type must be one of {sorted(ALLOWED_MSME_ENTERPRISE_TYPES)}; "
                        f"got '{value}'."
                    )
                value = normalized_value
            else:
                value = str(value).strip() if value is not None else ""
            setattr(submission, field, value)


def _enforce_invitation_email(submission: VendorOnboardingSubmission) -> None:
    """
    The invitation email is authoritative for vendor onboarding.

    Manual form and Excel upload payloads may include an email field, but vendors
    must not be able to change the identity/email that received the invite.
    """
    invitation_email = (submission.invitation.vendor_email or "").strip()
    if not invitation_email:
        return
    submission.normalized_email = invitation_email
    raw = dict(submission.raw_form_data or {})
    raw["email"] = invitation_email
    raw["Email Id"] = invitation_email
    submission.raw_form_data = raw


def _get_export_dir() -> Path:
    media_root = getattr(settings, "MEDIA_ROOT", settings.BASE_DIR / "media")
    export_dir = Path(media_root) / "vendor_exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    return export_dir


def _save_source_excel(submission: "VendorOnboardingSubmission", file_bytes: bytes, original_name: str) -> None:
    """Save the original uploaded Excel to disk and record the path on the submission."""
    try:
        source_dir = Path(getattr(settings, "MEDIA_ROOT", settings.BASE_DIR / "media")) / "vendor_source_uploads"
        source_dir.mkdir(parents=True, exist_ok=True)
        safe_name = Path(original_name).name or "upload.xlsx"
        dest_path = str(source_dir / f"sub_{submission.pk}_source_{safe_name}")
        with open(dest_path, "wb") as fh:
            fh.write(file_bytes)
        submission.source_excel_file = dest_path
        submission.save(update_fields=["source_excel_file", "updated_at"])
    except Exception as exc:
        _logger.warning("Failed to save source Excel for submission_id=%s: %s", submission.pk, exc)


def _build_audit_log(user, action: str, resource_type: str, resource_id: int, metadata: dict = None) -> None:
    """Write to AuditLog if the audit app is available."""
    try:
        from apps.audit.models import AuditLog
        AuditLog.objects.create(
            user=user,
            action=action,
            resource_type=resource_type,
            resource_id=resource_id,
            metadata=metadata or {},
        )
    except Exception:
        pass  # Audit failure must never block business logic


def _build_vendor_defaults_from_submission(submission: VendorOnboardingSubmission, invitation) -> dict:
    defaults = {
        "org": invitation.org,
        "scope_node": invitation.scope_node,
        "sap_vendor_id": submission.finance_vendor_code or "",
        "marketing_status": MarketingStatus.PENDING,
        "operational_status": OperationalStatus.WAITING_MARKETING_APPROVAL,
        "contact_persons_json": submission.contact_persons_json or [],
        "head_office_address_json": submission.head_office_address_json or {},
        "tax_registration_details_json": submission.tax_registration_details_json or {},
    }
    for submission_field, vendor_field in _SUBMISSION_TO_VENDOR_FIELD_MAP.items():
        defaults[vendor_field] = getattr(submission, submission_field, None)
    defaults["vendor_name"] = defaults.get("vendor_name") or invitation.vendor_name_hint or "Unknown Vendor"
    defaults["email"] = defaults.get("email") or ""
    defaults["phone"] = defaults.get("phone") or ""
    return defaults


def _get_editable_submission_for_invitation(
    invitation: VendorInvitation,
) -> VendorOnboardingSubmission | None:
    """
    Return the latest editable submission for an invitation.

    If the latest submission is already beyond the editable phase, block any
    attempt to create a fresh submission until finance explicitly reopens it.
    """
    latest = invitation.submissions.order_by("-created_at", "-id").first()
    if latest is None:
        return None

    if latest.status in _EDITABLE_SUBMISSION_STATUSES:
        return latest

    raise SubmissionStateError(
        "This submission has already been sent for review. "
        "You can edit it again only after it is reopened."
    )


# ---------------------------------------------------------------------------
# 1. create_vendor_invitation
# ---------------------------------------------------------------------------

@transaction.atomic
def create_vendor_invitation(
    org,
    scope_node,
    vendor_email: str,
    invited_by=None,
    vendor_name_hint: str = "",
    expires_at=None,
) -> VendorInvitation:
    """
    Create a new vendor invitation with a secure token and send the
    onboarding email to the vendor.

    Returns:
        VendorInvitation with status=pending
    """
    invitation = VendorInvitation.objects.create(
        org=org,
        scope_node=scope_node,
        vendor_email=vendor_email,
        invited_by=invited_by,
        vendor_name_hint=vendor_name_hint,
        token=_generate_token(),
        expires_at=expires_at,
        status=InvitationStatus.PENDING,
    )
    _build_audit_log(
        user=invited_by,
        action="vendor_invitation_created",
        resource_type="VendorInvitation",
        resource_id=invitation.pk,
        metadata={"vendor_email": vendor_email},
    )

    # Send invitation email (mockable in tests)
    _send_invitation_email(invitation, invited_by)

    return invitation


import logging

_logger = logging.getLogger(__name__)


def _send_invitation_email(invitation: VendorInvitation, invited_by) -> None:
    """Send the vendor onboarding invitation email. Mocked in tests."""
    from apps.vendors.email import send_vendor_invitation_email

    portal_base = getattr(settings, "VENDOR_PORTAL_BASE_URL", "http://localhost:5173")
    onboarding_url = f"{portal_base}/vendor/onboarding/{invitation.token}"
    if invited_by:
        invited_by_name = invited_by.get_full_name().strip() or invited_by.email
    else:
        invited_by_name = "Horizon Industrial Parks"

    try:
        send_vendor_invitation_email(
            vendor_email=invitation.vendor_email,
            vendor_name_hint=invitation.vendor_name_hint,
            onboarding_url=onboarding_url,
            invited_by_name=invited_by_name,
        )
    except Exception:
        _logger.exception(
            "Failed to send vendor invitation email for %s (invitation_id=%s)",
            invitation.vendor_email,
            invitation.pk,
        )
        raise


# ---------------------------------------------------------------------------
# 2. get_invitation_by_token
# ---------------------------------------------------------------------------

def get_invitation_by_token(token: str) -> VendorInvitation:
    """
    Look up and validate an invitation by its public token.

    Raises:
        InvitationNotFoundError — if token doesn't exist or invitation is cancelled
        InvitationExpiredError  — if invitation is past its expiry; also marks it expired
    """
    try:
        invitation = VendorInvitation.objects.select_related("org", "scope_node").get(token=token)
    except VendorInvitation.DoesNotExist:
        raise InvitationNotFoundError(f"No active invitation for token.")

    if invitation.status == InvitationStatus.CANCELLED:
        raise InvitationNotFoundError("This invitation has been cancelled.")

    if invitation.is_expired() and invitation.status not in (InvitationStatus.EXPIRED,):
        invitation.status = InvitationStatus.EXPIRED
        invitation.save(update_fields=["status", "updated_at"])

    if invitation.status == InvitationStatus.EXPIRED:
        raise InvitationExpiredError("This invitation has expired.")

    # Mark as opened if still pending
    if invitation.status == InvitationStatus.PENDING:
        invitation.status = InvitationStatus.OPENED
        invitation.save(update_fields=["status", "updated_at"])

    return invitation


# ---------------------------------------------------------------------------
# 3. create_or_update_submission_from_manual
# ---------------------------------------------------------------------------

@transaction.atomic
def create_or_update_submission_from_manual(
    invitation: VendorInvitation,
    payload: dict,
    submitted_by=None,
    finalize: bool = False,
) -> VendorOnboardingSubmission:
    """
    Save or update a manual-entry submission for the given invitation.

    - Stores the full payload in raw_form_data.
    - Extracts normalized core fields.
    - If finalize=True, sets status=submitted and submitted_at=now.

    Raises:
        SubmissionStateError if invitation's active submission is past submitted state.
    """
    # Get existing draft/reopened submission or create new
    submission = _get_editable_submission_for_invitation(invitation)
    if submission is None:
        submission = VendorOnboardingSubmission(invitation=invitation)

    if False and submission.pk and submission.status not in (
        SubmissionStatus.DRAFT, SubmissionStatus.REOPENED
    ):
        raise SubmissionStateError(
            f"Submission {submission.pk} is in status '{submission.status}' — cannot edit."
        )

    normalized, remaining_raw, contact_persons, head_office_address, tax_registration_details = _extract_normalized_from_payload(payload)

    # Preserve any previously stored raw data and merge
    merged_raw = {**submission.raw_form_data, **payload}
    submission.raw_form_data = merged_raw
    submission.submission_mode = SubmissionMode.MANUAL

    try:
        _apply_normalized_fields(submission, normalized)
        _enforce_invitation_email(submission)
    except ValueError as exc:
        raise SubmissionStateError(str(exc)) from exc

    # Store structured JSON blocks
    if contact_persons is not None:
        submission.contact_persons_json = contact_persons
    if head_office_address is not None:
        submission.head_office_address_json = head_office_address
    if tax_registration_details is not None:
        submission.tax_registration_details_json = tax_registration_details

    # Persist payload state before finance transition.
    # For new submissions pk is None — save() inserts; for existing drafts it updates.
    submission.save()

    if finalize:
        # Now that submission is persisted, move it into active finance review.
        # _start_finance_review expects a persisted record with a PK.
        _start_finance_review(submission)
        _build_audit_log(
            user=submitted_by,
            action="vendor_submission_finalized",
            resource_type="VendorOnboardingSubmission",
            resource_id=submission.pk,
            metadata={"mode": "manual"},
        )

    return submission


# ---------------------------------------------------------------------------
# 4. create_or_update_submission_from_excel
# ---------------------------------------------------------------------------

@transaction.atomic
def create_or_update_submission_from_excel(
    invitation: VendorInvitation,
    file_obj,
    submitted_by=None,
    finalize: bool = False,
) -> VendorOnboardingSubmission:
    """
    Parse a VRF-style Excel upload and create/update the submission.

    - Reads each row as label → value pairs.
    - Maps known labels to normalized fields.
    - Unknown labels are preserved in raw_form_data.
    - If finalize=True, sets status=submitted.

    Raises:
        SubmissionStateError if the existing submission is past the editable stage.
    """
    import io as _io
    import openpyxl

    file_obj.seek(0)
    _file_bytes = file_obj.read()
    _file_name = getattr(file_obj, "name", "upload.xlsx")
    wb = openpyxl.load_workbook(_io.BytesIO(_file_bytes), data_only=True)
    ws = wb.active

    extracted: dict = {}
    contact_persons = [
        {"type": "general_queries", "name": "", "designation": "", "email": "", "telephone": ""},
        {"type": "secondary", "name": "", "designation": "", "email": "", "telephone": ""},
    ]
    contact_has_values = [False, False]
    head_office_address: dict = {}
    tax_registration_details: dict = {}
    current_section = ""

    for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
        label, value = row[0], row[1]
        if label is None:
            continue
        label_str = str(label).strip()
        if not label_str:
            continue
        lower = label_str.lower().replace("*", "").strip().rstrip(":").strip()
        if value is None:
            if any(
                marker in lower
                for marker in (
                    "vendor information",
                    "supplying / billing address",
                    "tax registration",
                    "head office address",
                    "contact persons",
                    "payment details",
                    "bank details",
                    "msme declaration",
                )
            ):
                current_section = lower
            continue

        if "bank details" in current_section and lower in _BANK_DETAILS_LABEL_MAP:
            key = _BANK_DETAILS_LABEL_MAP[lower]
            extracted[key] = value
            continue

        if lower in _CONTACT_PERSON_LABEL_MAP:
            idx, field_name = _CONTACT_PERSON_LABEL_MAP[lower]
            contact_persons[idx][field_name] = str(value).strip()
            contact_has_values[idx] = True
            continue

        if "head office address" in current_section and lower in _HEAD_OFFICE_LABEL_MAP:
            head_office_address[_HEAD_OFFICE_LABEL_MAP[lower]] = str(value).strip()
            continue

        if lower in _TAX_REGISTRATION_LABEL_MAP:
            tax_registration_details[_TAX_REGISTRATION_LABEL_MAP[lower]] = str(value).strip()
            continue

        key = _VRF_LABEL_MAP.get(lower, label_str)
        extracted[key] = value

    if any(contact_has_values):
        extracted["contact_persons"] = [
            cp
            for cp in contact_persons
            if any(
                str(cp.get(field, "")).strip()
                for field in ("name", "designation", "email", "telephone")
            )
        ]
    if head_office_address:
        extracted["head_office_address"] = head_office_address
    if tax_registration_details:
        extracted["tax_registration_details"] = tax_registration_details

    normalized, _, contact_persons_json, head_office_json, tax_registration_json = _extract_normalized_from_payload(extracted)

    submission = _get_editable_submission_for_invitation(invitation)
    if submission is None:
        submission = VendorOnboardingSubmission(invitation=invitation)

    if False and submission.pk and submission.status not in (
        SubmissionStatus.DRAFT, SubmissionStatus.REOPENED
    ):
        raise SubmissionStateError(
            f"Submission {submission.pk} is in status '{submission.status}' — cannot edit."
        )

    merged_raw = {**submission.raw_form_data, **extracted}
    submission.raw_form_data = merged_raw
    submission.submission_mode = SubmissionMode.EXCEL_UPLOAD

    _apply_normalized_fields(submission, normalized)
    _enforce_invitation_email(submission)

    if contact_persons_json is not None:
        submission.contact_persons_json = contact_persons_json
    if head_office_json is not None:
        submission.head_office_address_json = head_office_json
    if tax_registration_json is not None:
        submission.tax_registration_details_json = tax_registration_json

    # Persist payload state before finance transition.
    submission.save()

    # Save the original uploaded file now that we have a pk.
    _save_source_excel(submission, _file_bytes, _file_name)

    if finalize:
        _start_finance_review(submission)
        _build_audit_log(
            user=submitted_by,
            action="vendor_submission_finalized",
            resource_type="VendorOnboardingSubmission",
            resource_id=submission.pk,
            metadata={"mode": "excel_upload"},
        )

    return submission


# ---------------------------------------------------------------------------
# 5. attach_document
# ---------------------------------------------------------------------------

_MSME_DECLARATION_ALLOWED_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".docx"}
_MSME_DECLARATION_ERROR = (
    "Please upload the completed MSME Declaration Form downloaded from this page. "
    "The uploaded file does not contain the expected MSME declaration text."
)


def _read_uploaded_file_bytes(file_obj) -> bytes:
    try:
        position = file_obj.tell()
    except (AttributeError, OSError):
        position = None

    try:
        if hasattr(file_obj, "seek"):
            file_obj.seek(0)
        data = file_obj.read()
    finally:
        if position is not None and hasattr(file_obj, "seek"):
            file_obj.seek(position)

    return data or b""


def _extract_docx_text(file_bytes: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as docx_zip:
            xml_bytes = docx_zip.read("word/document.xml")
    except (KeyError, zipfile.BadZipFile) as exc:
        raise ValueError("Unable to read MSME Declaration Form Word document.") from exc

    try:
        root = ElementTree.fromstring(xml_bytes)
    except ElementTree.ParseError as exc:
        raise ValueError("Unable to read MSME Declaration Form Word document.") from exc

    return " ".join(text for text in root.itertext() if text)


def _extract_pdf_text(file_bytes: bytes) -> str:
    text_parts: list[str] = []

    try:
        from PyPDF2 import PdfReader

        reader = PdfReader(io.BytesIO(file_bytes))
        for page in reader.pages[:3]:
            text_parts.append(page.extract_text() or "")
    except Exception:
        # Scanned PDFs often have no embedded text. OCR fallback below handles them.
        pass

    extracted_text = "\n".join(text_parts)
    if _looks_like_msme_declaration(extracted_text):
        return extracted_text

    try:
        import fitz
        from PIL import Image

        document = fitz.open(stream=file_bytes, filetype="pdf")
        ocr_parts: list[str] = []
        for page_index in range(min(3, document.page_count)):
            page = document.load_page(page_index)
            pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            image = Image.open(io.BytesIO(pixmap.tobytes("png")))
            ocr_parts.append(_ocr_image(image))
        return "\n".join([extracted_text, *ocr_parts])
    except ImportError as exc:
        raise ValueError(
            "MSME declaration verification is not configured. Install PyMuPDF, Pillow, and pytesseract."
        ) from exc
    except Exception as exc:
        raise ValueError("Unable to verify the MSME Declaration Form PDF.") from exc


def _configure_tesseract() -> None:
    try:
        import pytesseract
    except ImportError as exc:
        raise ValueError(
            "MSME declaration OCR is not configured. Install pytesseract on the backend."
        ) from exc

    configured_cmd = getattr(settings, "VENDOR_MSME_TESSERACT_CMD", "")
    if configured_cmd:
        pytesseract.pytesseract.tesseract_cmd = configured_cmd
        return

    windows_default = Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe")
    if os.name == "nt" and windows_default.exists():
        pytesseract.pytesseract.tesseract_cmd = str(windows_default)


def _ocr_image(image) -> str:
    _configure_tesseract()
    try:
        import pytesseract

        return pytesseract.image_to_string(image)
    except Exception as exc:
        raise ValueError("Unable to OCR the MSME Declaration Form upload.") from exc


def _extract_image_text(file_bytes: bytes) -> str:
    try:
        from PIL import Image

        image = Image.open(io.BytesIO(file_bytes))
        return _ocr_image(image)
    except ImportError as exc:
        raise ValueError("MSME declaration image verification requires Pillow.") from exc
    except Exception as exc:
        raise ValueError("Unable to verify the MSME Declaration Form image.") from exc


def _normalize_ocr_text(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").lower()).strip()


def _looks_like_msme_declaration(text: str) -> bool:
    normalized = _normalize_ocr_text(text)
    if not normalized:
        return False

    has_msme_marker = (
        "msme" in normalized
        or "micro small and medium" in normalized
        or "micro, small and medium" in normalized
    )
    has_declaration_marker = "declaration" in normalized or "declare" in normalized
    if not (has_msme_marker and has_declaration_marker):
        return False

    supporting_markers = (
        "udyam",
        "uan",
        "registration number",
        "enterprise",
        "micro",
        "small",
        "medium",
        "authorized signatory",
        "authorised signatory",
        "signatory",
    )
    score = sum(1 for marker in supporting_markers if marker in normalized)
    return score >= 3


def _validate_msme_declaration_upload(file_obj, file_name: str) -> None:
    ext = Path(file_name).suffix.lower()
    if ext not in _MSME_DECLARATION_ALLOWED_EXTENSIONS:
        allowed = ", ".join(sorted(_MSME_DECLARATION_ALLOWED_EXTENSIONS))
        raise ValueError(f"MSME Declaration Form must be uploaded as one of: {allowed}.")

    file_bytes = _read_uploaded_file_bytes(file_obj)
    if not file_bytes:
        raise ValueError("MSME Declaration Form upload is empty.")

    if ext == ".docx":
        extracted_text = _extract_docx_text(file_bytes)
    elif ext == ".pdf":
        extracted_text = _extract_pdf_text(file_bytes)
    else:
        extracted_text = _extract_image_text(file_bytes)

    if not _looks_like_msme_declaration(extracted_text):
        raise ValueError(_MSME_DECLARATION_ERROR)

    if hasattr(file_obj, "seek"):
        file_obj.seek(0)


def _ensure_submission_attachments_editable(submission: VendorOnboardingSubmission) -> None:
    if submission.status not in _EDITABLE_SUBMISSION_STATUSES:
        raise SubmissionStateError(
            f"Submission {submission.pk} is in status '{submission.status}' — attachments cannot be changed."
        )


def _delete_attachment_file(attachment: VendorAttachment) -> None:
    if attachment.file:
        try:
            attachment.file.delete(save=False)
        except Exception:
            # File deletion failure should not leave stale DB rows blocking replacement.
            pass


def _delete_attachment_record(attachment: VendorAttachment) -> None:
    _delete_attachment_file(attachment)
    attachment.delete()


def remove_submission_attachment(
    submission: VendorOnboardingSubmission,
    attachment_id: int,
) -> None:
    """
    Remove one attachment from an editable vendor onboarding submission.

    The attachment must belong to the supplied submission.  This is used by the
    public invitation flow, so never accept arbitrary attachment deletion.
    """
    _ensure_submission_attachments_editable(submission)
    try:
        attachment = submission.attachments.get(pk=attachment_id)
    except VendorAttachment.DoesNotExist as exc:
        raise ValueError("Attachment not found for this submission.") from exc

    from apps.vendors.models import ACTIVE_VENDOR_ATTACHMENT_DOCUMENT_TYPES

    if attachment.document_type in ACTIVE_VENDOR_ATTACHMENT_DOCUMENT_TYPES:
        attachments_to_delete = list(
            submission.attachments.filter(document_type=attachment.document_type)
        )
    else:
        attachments_to_delete = [attachment]

    for existing_attachment in attachments_to_delete:
        _delete_attachment_record(existing_attachment)


def attach_document(
    submission: VendorOnboardingSubmission,
    title: str,
    file_obj=None,
    file_name: str = "",
    file_url: str = "",
    document_type: str = "",
    uploaded_by=None,
) -> VendorAttachment:
    """
    Create a VendorAttachment record.

    If file_obj is provided (Django uploaded file or file-like), it is stored
    via the FileField on the model (controlled media storage).  file_name is
    derived from file_obj.name if not explicitly supplied.

    For backward compatibility, file_name + file_url without a file_obj still
    work (metadata-only record).

    Raises:
        ValueError — document_type is not in ALLOWED_ATTACHMENT_DOCUMENT_TYPES
    """
    from apps.vendors.models import ACTIVE_VENDOR_ATTACHMENT_DOCUMENT_TYPES

    _ensure_submission_attachments_editable(submission)

    if document_type and document_type not in ACTIVE_VENDOR_ATTACHMENT_DOCUMENT_TYPES:
        raise ValueError(
            f"document_type '{document_type}' is not allowed. "
            f"Accepted types: {', '.join(sorted(ACTIVE_VENDOR_ATTACHMENT_DOCUMENT_TYPES))}"
        )

    resolved_name = file_name
    if file_obj and not resolved_name:
        from pathlib import Path
        resolved_name = Path(getattr(file_obj, "name", "attachment")).name

    if file_obj and document_type == "msme_declaration_form":
        _validate_msme_declaration_upload(file_obj, resolved_name)

    if document_type in ACTIVE_VENDOR_ATTACHMENT_DOCUMENT_TYPES:
        existing_attachments = list(
            submission.attachments.filter(document_type=document_type)
        )
        for existing_attachment in existing_attachments:
            _delete_attachment_record(existing_attachment)

    attachment = VendorAttachment(
        submission=submission,
        title=title,
        file_name=resolved_name,
        file_url=file_url,
        document_type=document_type,
        uploaded_by=uploaded_by,
    )
    if file_obj:
        attachment.file.save(resolved_name, file_obj, save=False)
    attachment.save()
    return attachment


def get_required_attachment_types(submission: VendorOnboardingSubmission) -> tuple[str, ...]:
    from apps.vendors.models import (
        REQUIRED_MSME_ATTACHMENT_DOCUMENT_TYPES,
        REQUIRED_VENDOR_ATTACHMENT_DOCUMENT_TYPES,
    )

    required = list(REQUIRED_VENDOR_ATTACHMENT_DOCUMENT_TYPES)
    if submission.normalized_msme_registered:
        required.extend(REQUIRED_MSME_ATTACHMENT_DOCUMENT_TYPES)
    return tuple(required)


def get_missing_required_attachment_labels(submission: VendorOnboardingSubmission) -> list[str]:
    from apps.vendors.models import VENDOR_ATTACHMENT_DOCUMENT_TYPE_LABELS

    existing_types = set(
        submission.attachments.exclude(document_type="")
        .values_list("document_type", flat=True)
    )
    return [
        VENDOR_ATTACHMENT_DOCUMENT_TYPE_LABELS.get(document_type, document_type)
        for document_type in get_required_attachment_types(submission)
        if document_type not in existing_types
    ]


# ---------------------------------------------------------------------------
# 6. generate_vendor_export_excel
# ---------------------------------------------------------------------------

def generate_vendor_export_excel(submission: VendorOnboardingSubmission) -> str:
    """
    Generate the canonical VRF export workbook from the submission's data.

    Writes the file to MEDIA_ROOT/vendor_exports/sub_{id}_vrf.xlsx.
    Updates submission.exported_excel_file and saves the field.

    Returns the filesystem path.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendor Registration Form"

    # Styles
    header_font = Font(bold=True, size=13)
    section_font = Font(bold=True, size=11)
    section_fill = PatternFill("solid", fgColor="D9E1F2")
    label_font = Font(bold=True)

    def _section(row_num: int, title: str):
        cell = ws.cell(row=row_num, column=1, value=title)
        cell.font = section_font
        cell.fill = section_fill
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
        return row_num + 1

    def _row(row_num: int, label: str, value):
        ws.cell(row=row_num, column=1, value=label).font = label_font
        ws.cell(row=row_num, column=2, value=str(value) if value is not None else "")
        return row_num + 1

    # Title
    title_cell = ws.cell(row=1, column=1, value="Horizon - Vendor Registration Form (Export)")
    title_cell.font = header_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    ws.cell(row=2, column=1, value=f"Submission ID: {submission.pk}")
    ws.cell(row=2, column=2, value=f"Status: {submission.status}")

    row = 4

    # Section 1: Vendor Information
    row = _section(row, "SECTION 1: VENDOR INFORMATION")
    row = _row(row, "Title", submission.normalized_title)
    row = _row(row, "Vendor Name", submission.normalized_vendor_name)
    row = _row(row, "Vendor Type", submission.normalized_vendor_type)
    row = _row(row, "GST Registration", "Registered" if submission.normalized_gst_registered else "Un-Registered" if submission.normalized_gst_registered is False else "")
    row = _row(row, "GSTIN Number", submission.normalized_gstin)
    row += 1

    # Section 2: Supplying / Billing Address
    row = _section(row, "SECTION 2: SUPPLYING / BILLING ADDRESS")
    row = _row(row, "Address Line 1", submission.normalized_address_line1)
    row = _row(row, "Address Line 2", submission.normalized_address_line2)
    row = _row(row, "Address Line 3", submission.normalized_address_line3)
    row = _row(row, "City", submission.normalized_city)
    row = _row(row, "Pin Code", submission.normalized_pincode)
    row = _row(row, "State", submission.normalized_state)
    row = _row(row, "Country", submission.normalized_country)
    row = _row(row, "Phone no", submission.normalized_phone)
    row = _row(row, "Fax no", submission.normalized_fax)
    row = _row(row, "Email Id", submission.normalized_email)
    row = _row(row, "Region", submission.normalized_region)
    row = _row(row, "Head Office no", submission.normalized_head_office_no)
    row += 1

    # Section 3: Tax Registration Nos
    row = _section(row, "SECTION 3: TAX REGISTRATION NOS")
    taxd = submission.tax_registration_details_json or {}
    row = _row(row, "TIN NO", taxd.get("tin_no", ""))
    row = _row(row, "CST No.", taxd.get("cst_no", ""))
    row = _row(row, "LST No.", taxd.get("lst_no", ""))
    row = _row(row, "PAN No.", submission.normalized_pan)
    row = _row(row, "ESIC Reg NO", taxd.get("esic_reg_no", ""))
    row = _row(row, "PAN Ref. No.", taxd.get("pan_ref_no", ""))
    row = _row(row, "PPF No.", taxd.get("ppf_no", ""))
    row += 1

    # Section 4: Head Office Address
    row = _section(row, "SECTION 4: HEAD OFFICE ADDRESS")
    hoa = submission.head_office_address_json or {}
    row = _row(row, "Address line 1", hoa.get("address_line1", ""))
    row = _row(row, "Address line 2", hoa.get("address_line2", ""))
    row = _row(row, "City", hoa.get("city", ""))
    row = _row(row, "Pincode", hoa.get("pincode", ""))
    row = _row(row, "State", hoa.get("state", ""))
    row = _row(row, "Country", hoa.get("country", ""))
    row = _row(row, "Phone no", hoa.get("phone", ""))
    row = _row(row, "Fax no", hoa.get("fax", ""))
    row += 1

    # Section 5: Contact Persons
    row = _section(row, "SECTION 5: CONTACT PERSONS")
    row = _row(row, "General Queries", "")
    contact_persons = submission.contact_persons_json or []
    contact_1 = contact_persons[0] if len(contact_persons) > 0 else {}
    contact_2 = contact_persons[1] if len(contact_persons) > 1 else {}
    row = _row(row, "1) Name", contact_1.get("name", ""))
    row = _row(row, "1) Designation", contact_1.get("designation", ""))
    row = _row(row, "1) Email Address", contact_1.get("email", ""))
    row = _row(row, "1) Telephone", contact_1.get("telephone", ""))
    row = _row(row, "2) Name", contact_2.get("name", ""))
    row = _row(row, "2) Designation", contact_2.get("designation", ""))
    row = _row(row, "2) Email Address", contact_2.get("email", ""))
    row = _row(row, "2) Telephone", contact_2.get("telephone", ""))
    row += 1

    # Section 6: Payment Details
    row = _section(row, "SECTION 6: PAYMENT DETAILS")
    row = _row(row, "Preffered Payment Mode", submission.normalized_preferred_payment_mode)
    row += 1

    # Section 7: Bank Details
    row = _section(row, "SECTION 7: BANK DETAILS")
    row = _row(row, "Bank Name", submission.normalized_bank_name)
    row = _row(row, "Bank Address", submission.normalized_bank_address)
    row = _row(row, "Address line", submission.normalized_bank_branch_address_line1)
    row = _row(row, "Addess line2", submission.normalized_bank_branch_address_line2)
    row = _row(row, "City", submission.normalized_bank_branch_city)
    row = _row(row, "Pin code", submission.normalized_bank_branch_pincode)
    row = _row(row, "State", submission.normalized_bank_branch_state)
    row = _row(row, "Country", submission.normalized_bank_branch_country)
    row = _row(row, "Phone No", submission.normalized_bank_phone)
    row = _row(row, "Fax No", submission.normalized_bank_fax)
    row = _row(row, "Beneficiary Name", submission.normalized_beneficiary_name)
    row = _row(row, "Beneficiary Account No", submission.normalized_beneficiary_account_number)
    row = _row(row, "Bank Account No", submission.normalized_bank_account_number)
    row = _row(row, "Bank account type", submission.normalized_bank_account_type)
    row = _row(row, "Bank MICR code", submission.normalized_micr_code)
    row = _row(row, "NEFT code", submission.normalized_neft_code)
    row = _row(row, "IFSC code", submission.normalized_ifsc)
    row = _row(row, "Email Address", submission.normalized_bank_email)
    row += 1

    # Section 8: MSME Declaration
    row = _section(row, "SECTION 8: MSME DECLARATION")
    row = _row(row, "MSME Registered", "Yes" if submission.normalized_msme_registered else "No" if submission.normalized_msme_registered is False else "")
    row = _row(row, "MSME Registration Number", submission.normalized_msme_registration_number)
    row = _row(row, "Enterprise Type", submission.normalized_msme_enterprise_type)
    row = _row(row, "Authorized Signatory Name", submission.normalized_authorized_signatory_name)
    row += 1

    # Additional raw data
    raw = submission.raw_form_data or {}
    extra = {k: v for k, v in raw.items()
             if k not in _NORM_FIELD_MAP and k not in _KNOWN_KEYS
             and k not in ("contact_persons", "head_office_address", "tax_registration_details")}
    if extra:
        row = _section(row, "ADDITIONAL INFORMATION")
        for k, v in extra.items():
            row = _row(row, k, v)

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40

    # Save
    export_dir = _get_export_dir()
    file_path = str(export_dir / f"sub_{submission.pk}_vrf.xlsx")
    wb.save(file_path)

    submission.exported_excel_file = file_path
    submission.save(update_fields=["exported_excel_file", "updated_at"])

    return file_path


# ---------------------------------------------------------------------------
# 7. send_submission_to_finance
# ---------------------------------------------------------------------------

@transaction.atomic
def send_submission_to_finance(
    submission: VendorOnboardingSubmission,
    triggered_by=None,
) -> VendorOnboardingSubmission:
    """
    Generate export workbook, create finance action tokens, send email to
    configured finance recipients, and transition submission to sent_to_finance.

    Submission must be in status=submitted or reopened.

    Raises:
        SubmissionStateError — if submission is not in a sendable state.
    """
    if submission.status not in (SubmissionStatus.SUBMITTED, SubmissionStatus.REOPENED):
        raise SubmissionStateError(
            f"Submission {submission.pk} is in '{submission.status}' — "
            "must be 'submitted' or 'reopened' to send to finance."
        )

    # Generate / refresh export Excel
    export_path = generate_vendor_export_excel(submission)

    # Create (or refresh) finance action tokens
    expiry_hours = getattr(settings, "VENDOR_FINANCE_TOKEN_EXPIRY_HOURS", 72)
    expires_at = timezone.now() + timedelta(hours=expiry_hours)

    approve_token = VendorFinanceActionToken.objects.create(
        submission=submission,
        action_type=FinanceActionType.APPROVE,
        token=_generate_token(),
        expires_at=expires_at,
    )
    reject_token = VendorFinanceActionToken.objects.create(
        submission=submission,
        action_type=FinanceActionType.REJECT,
        token=_generate_token(),
        expires_at=expires_at,
    )

    base_url = getattr(
        settings,
        "VENDOR_FINANCE_PORTAL_BASE_URL",
        getattr(settings, "VENDOR_PORTAL_BASE_URL", "http://localhost:3000"),
    )
    # Finance emails expose one review entry point; approve/reject happens inside
    # the review page using the paired action tokens.
    approve_url = f"{base_url}/vendor/finance/{approve_token.token}"
    reject_url = approve_url

    # Gather attachment URLs (legacy file_url only — FileField uploads shown on review page)
    attachment_urls = list(
        submission.attachments.exclude(file_url="").values_list("file_url", flat=True)
    )

    # Resolve inviting user and scope for email context
    invitation = submission.invitation
    inviting_user = None
    scope_name = None
    if invitation:
        inviter = getattr(invitation, "invited_by", None)
        if inviter:
            inviting_user = inviter.get_full_name().strip() or inviter.email
        if invitation.scope_node:
            scope_name = getattr(invitation.scope_node, "name", None)

    # Send email (mockable in tests)
    from apps.vendors.email import send_finance_email
    send_finance_email(
        submission_id=submission.pk,
        vendor_name=submission.normalized_vendor_name or "Unknown Vendor",
        approve_url=approve_url,
        reject_url=reject_url,
        inviting_user=inviting_user,
        scope_name=scope_name,
        exported_excel_path=export_path,
        attachment_urls=attachment_urls,
    )

    submission.status = SubmissionStatus.SENT_TO_FINANCE
    submission.finance_sent_at = timezone.now()
    submission.save(update_fields=["status", "finance_sent_at", "updated_at"])

    _build_audit_log(
        user=triggered_by,
        action="vendor_submission_sent_to_finance",
        resource_type="VendorOnboardingSubmission",
        resource_id=submission.pk,
        metadata={"approve_token": approve_token.pk, "reject_token": reject_token.pk},
    )

    return submission


# ---------------------------------------------------------------------------
# 8. finance_approve_submission
# ---------------------------------------------------------------------------

@transaction.atomic
def finance_approve_submission(
    token_str: str,
    sap_vendor_id: str,
    note: str = "",
) -> tuple[VendorOnboardingSubmission, Vendor]:
    """
    Finance approves the submission via the token link.

    - Token must be action_type=approve, not expired, not used.
    - Submission must be sent_to_finance or reopened.
    - Creates VendorFinanceDecision (approved).
    - Creates or updates the Vendor master record.
    - Transitions submission to marketing_pending.

    Returns (submission, vendor).

    Raises:
        FinanceTokenError      — invalid/expired/used token
        SubmissionStateError   — submission in wrong state
        ValueError             — sap_vendor_id missing
    """
    token = _get_valid_finance_token(token_str, expected_action=FinanceActionType.APPROVE)
    return approve_vendor_submission_finance(
        submission=token.submission,
        sap_vendor_id=sap_vendor_id,
        note=note,
        token=token,
        actor=None,
    )


@transaction.atomic
def approve_vendor_submission_finance(
    submission: VendorOnboardingSubmission,
    sap_vendor_id: str,
    note: str = "",
    token: VendorFinanceActionToken | None = None,
    actor=None,
) -> tuple[VendorOnboardingSubmission, Vendor]:
    """
    Finance approves a vendor onboarding submission.

    Shared by public email-token links and authenticated finance portal actions.
    """
    if not sap_vendor_id or not sap_vendor_id.strip():
        raise ValueError("sap_vendor_id is required for finance approval.")

    if submission.status not in (SubmissionStatus.SENT_TO_FINANCE, SubmissionStatus.REOPENED):
        raise SubmissionStateError(
            f"Submission {submission.pk} is in '{submission.status}' — cannot approve."
        )

    now = timezone.now()

    _consume_submission_finance_tokens(submission, now)

    # Create finance decision
    VendorFinanceDecision.objects.create(
        submission=submission,
        decision=FinanceDecisionChoice.APPROVED,
        sap_vendor_id=sap_vendor_id.strip(),
        note=note,
        acted_via_token=token,
        acted_at=now,
    )

    # Update submission
    submission.finance_vendor_code = sap_vendor_id.strip()
    submission.status = SubmissionStatus.MARKETING_PENDING
    submission.save(update_fields=["finance_vendor_code", "status", "updated_at"])

    # Create or upsert Vendor
    invitation = submission.invitation
    vendor_defaults = _build_vendor_defaults_from_submission(submission, invitation)
    vendor_defaults["sap_vendor_id"] = sap_vendor_id.strip()
    vendor, _ = Vendor.objects.update_or_create(
        onboarding_submission=submission,
        defaults=vendor_defaults,
    )

    _build_audit_log(
        user=actor,
        action="vendor_finance_approved",
        resource_type="VendorOnboardingSubmission",
        resource_id=submission.pk,
        metadata={"sap_vendor_id": sap_vendor_id, "vendor_id": vendor.pk},
    )

    # Notify vendor and inviter of the approval + marketing next step
    from apps.vendors.notifications import notify_vendor_approved
    try:
        notify_vendor_approved(submission, vendor)
    except Exception:
        pass  # Notification failure does not roll back the state transition

    return submission, vendor


# ---------------------------------------------------------------------------
# 9. finance_reject_submission
# ---------------------------------------------------------------------------

@transaction.atomic
def finance_reject_submission(
    token_str: str,
    note: str = "",
) -> VendorOnboardingSubmission:
    """
    Finance rejects the submission via the token link.

    Returns updated submission.

    Raises:
        FinanceTokenError    — invalid/expired/used token
        SubmissionStateError — submission in wrong state
    """
    token = _get_valid_finance_token(token_str, expected_action=FinanceActionType.REJECT)
    return reject_vendor_submission_finance(
        submission=token.submission,
        note=note,
        token=token,
        actor=None,
    )


@transaction.atomic
def reject_vendor_submission_finance(
    submission: VendorOnboardingSubmission,
    note: str = "",
    token: VendorFinanceActionToken | None = None,
    actor=None,
) -> VendorOnboardingSubmission:
    """
    Finance rejects a vendor onboarding submission.

    Shared by public email-token links and authenticated finance portal actions.
    """

    if submission.status not in (SubmissionStatus.SENT_TO_FINANCE, SubmissionStatus.REOPENED):
        raise SubmissionStateError(
            f"Submission {submission.pk} is in '{submission.status}' — cannot reject."
        )

    now = timezone.now()
    _consume_submission_finance_tokens(submission, now)

    VendorFinanceDecision.objects.create(
        submission=submission,
        decision=FinanceDecisionChoice.REJECTED,
        note=note,
        acted_via_token=token,
        acted_at=now,
    )

    submission.status = SubmissionStatus.FINANCE_REJECTED
    submission.save(update_fields=["status", "updated_at"])

    _build_audit_log(
        user=actor,
        action="vendor_finance_rejected",
        resource_type="VendorOnboardingSubmission",
        resource_id=submission.pk,
        metadata={"note": note},
    )

    # Reopen immediately so the original onboarding link becomes editable for
    # vendor correction. The finance rejection remains captured in
    # VendorFinanceDecision and audit history.
    reopen_submission(submission, reopened_by=actor, note=note)

    # Notify vendor and inviter of the rejection
    from apps.vendors.notifications import notify_vendor_rejected
    try:
        notify_vendor_rejected(submission, note=note)
    except Exception:
        pass  # Notification failure does not roll back the state transition

    return submission


# ---------------------------------------------------------------------------
# 10. reopen_submission
# ---------------------------------------------------------------------------

@transaction.atomic
def reopen_submission(
    submission: VendorOnboardingSubmission,
    reopened_by=None,
    note: str = "",
) -> VendorOnboardingSubmission:
    """
    Reopen a finance-rejected submission so the vendor can correct and re-send.

    Only allowed from finance_rejected status.

    Raises:
        SubmissionStateError — if submission is not finance_rejected
    """
    if submission.status != SubmissionStatus.FINANCE_REJECTED:
        raise SubmissionStateError(
            f"Submission {submission.pk} is in '{submission.status}' — "
            "can only reopen from 'finance_rejected'."
        )

    submission.status = SubmissionStatus.REOPENED
    submission.save(update_fields=["status", "updated_at"])
    VendorInvitation.objects.filter(pk=submission.invitation_id).update(
        status=InvitationStatus.OPENED,
    )

    _build_audit_log(
        user=reopened_by,
        action="vendor_submission_reopened",
        resource_type="VendorOnboardingSubmission",
        resource_id=submission.pk,
        metadata={"note": note},
    )

    from apps.vendors.notifications import notify_vendor_reopened
    try:
        notify_vendor_reopened(submission, note=note)
    except Exception:
        pass  # Notification failure does not roll back the state transition

    return submission


# ---------------------------------------------------------------------------
# 10. Shared portal access helpers
# ---------------------------------------------------------------------------

def get_vendor_email(vendor: Vendor) -> str:
    """Return vendor email or fallback from onboarding submission."""
    if vendor.email:
        return vendor.email
    if vendor.onboarding_submission_id:
        sub = vendor.onboarding_submission
        if sub and getattr(sub, "normalized_email", None):
            return sub.normalized_email
    raise VendorStateError(f"Vendor {vendor.pk} has no email — cannot create portal user.")


def ensure_vendor_portal_user(vendor: Vendor):
    """
    Create or reuse a portal user for vendor.email.
    User is always active (usable after password is set).
    Returns (user, created).
    """
    from django.contrib.auth import get_user_model
    User = get_user_model()
    vendor_email = get_vendor_email(vendor)
    user, created = User.objects.get_or_create(
        email=vendor_email,
        defaults={"is_active": True},
    )
    if not user.is_active:
        user.is_active = True
        user.save(update_fields=["is_active"])
    return user, created, vendor_email


def ensure_vendor_portal_assignment(vendor: Vendor, user) -> tuple:
    """Create or reactivate UserVendorAssignment. Returns (assignment, created)."""
    assignment, created = UserVendorAssignment.objects.update_or_create(
        user=user,
        vendor=vendor,
        defaults={"is_active": True},
    )
    return assignment, created


def create_vendor_activation_token(user, vendor, vendor_email: str, actor=None):
    """
    Invalidate all existing un-used tokens for this user,
    then create a fresh activation token.
    Returns (token, created).
    """
    now = timezone.now()
    expiry_days = getattr(settings, "VENDOR_ACTIVATION_TOKEN_EXPIRY_DAYS", 7)
    expires_at = now + timedelta(days=expiry_days)

    # Invalidate any existing un-used tokens for this user
    VendorActivationToken.objects.filter(
        uid=str(user.pk), used_at__isnull=True
    ).update(used_at=now)

    token = VendorActivationToken.objects.create(
        uid=str(user.pk),
        token=secrets.token_urlsafe(48),
        expires_at=expires_at,
        sent_by=actor,
        vendor=vendor,
        sent_at=timezone.now(),
    )
    return token, True


def send_vendor_activation_for_vendor(vendor: Vendor, actor=None):
    """
    Full portal activation flow: ensure user, assignment, token, send email.

    Idempotent:
    - If user/assignment already exist, they are reused.
    - Old unused tokens are invalidated before new one is created.
    - Email send is mandatory — raises on failure.

    Args:
        vendor:    Vendor instance (must be active)
        actor:     User who triggered this (for audit log, can be None)

    Returns dict with result metadata:
        {
            "email": str,
            "user_created": bool,
            "assignment_created": bool,
            "token_created": bool,
        }

    Raises:
        VendorStateError — if vendor is not active or has no email
    """
    if vendor.operational_status != OperationalStatus.ACTIVE:
        raise VendorStateError(
            f"Vendor {vendor.pk} is in '{vendor.operational_status}' — "
            "must be 'active' to send portal activation."
        )

    user, user_created, vendor_email = ensure_vendor_portal_user(vendor)
    assignment, assignment_created = ensure_vendor_portal_assignment(vendor, user)
    token, token_created = create_vendor_activation_token(user, vendor, vendor_email, actor=actor)

    base_url = getattr(settings, "VENDOR_PORTAL_BASE_URL", "http://localhost:3000")
    activation_url = f"{base_url}/vendor/activate/{user.pk}/{token.token}"

    from apps.vendors.email import send_vendor_activation_email
    send_vendor_activation_email(
        vendor_email=vendor_email,
        vendor_name=vendor.vendor_name,
        activation_url=activation_url,
    )

    # Update cached portal tracking fields on Vendor
    vendor.portal_activation_sent_at = timezone.now()
    vendor.portal_user_id = str(user.pk)
    vendor.portal_email = vendor_email
    vendor.save(update_fields=["portal_activation_sent_at", "portal_user_id", "portal_email", "updated_at"])

    _build_audit_log(
        user=actor,
        action="vendor_activation_resent",
        resource_type="Vendor",
        resource_id=vendor.pk,
        metadata={
            "portal_user_id": user.pk,
            "activation_token_id": token.pk,
            "assignment_id": assignment.pk,
            "user_created": user_created,
            "assignment_created": assignment_created,
            "token_created": token_created,
            "email": vendor_email,
        },
    )

    return {
        "email": vendor_email,
        "user_created": user_created,
        "assignment_created": assignment_created,
        "token_created": token_created,
    }


def send_vendor_contact_activation_notices(vendor: Vendor, primary_vendor_email: str) -> dict:
    """
    Notify listed vendor business contacts after final approval.

    Informational contact notices are best-effort. A delivery failure must not
    roll back vendor activation or create access for the contact recipient.
    """
    from apps.vendors.email import send_vendor_contact_activation_notice_email

    excluded_emails = {
        email.casefold()
        for email in (primary_vendor_email, vendor.email, vendor.portal_email)
        if email
    }
    sent_emails = []
    skipped_emails = []
    failed_emails = []

    for contact in vendor.contact_persons_json or []:
        if not isinstance(contact, dict):
            continue
        contact_email = str(contact.get("email") or "").strip()
        if not contact_email:
            continue
        normalized_email = contact_email.casefold()
        if normalized_email in excluded_emails:
            skipped_emails.append(contact_email)
            continue
        try:
            validate_email(contact_email)
        except ValidationError:
            skipped_emails.append(contact_email)
            logger.warning(
                "Skipping invalid vendor contact notification email vendor_id=%s email=%s",
                vendor.pk,
                contact_email,
            )
            continue

        excluded_emails.add(normalized_email)
        try:
            send_vendor_contact_activation_notice_email(
                contact_email=contact_email,
                contact_name=str(contact.get("name") or "").strip(),
                vendor_name=vendor.vendor_name,
                primary_vendor_email=primary_vendor_email,
            )
            sent_emails.append(contact_email)
        except Exception:
            failed_emails.append(contact_email)
            logger.exception(
                "Vendor contact activation notice delivery failed vendor_id=%s email=%s",
                vendor.pk,
                contact_email,
            )

    return {
        "contact_notice_sent_emails": sent_emails,
        "contact_notice_skipped_emails": skipped_emails,
        "contact_notice_failed_emails": failed_emails,
    }


# ---------------------------------------------------------------------------
# 11. approve_vendor_marketing
# ---------------------------------------------------------------------------

@transaction.atomic
def approve_vendor_marketing(
    vendor: Vendor,
    approved_by,
) -> Vendor:
    """
    Marketing approves the vendor, making it operational and initiating portal activation.

    - vendor must be operational_status=waiting_marketing_approval.
    - Sets marketing_status=approved, operational_status=active.
    - Clears po_mandate_enabled; PO numbers are optional invoice metadata.
    - Sets linked submission.status=activated.
    - Sends portal activation email (mandatory — rolls back on failure).

    Raises:
        VendorStateError — if vendor is not in waiting_marketing_approval
    """
    if vendor.operational_status != OperationalStatus.WAITING_MARKETING_APPROVAL:
        raise VendorStateError(
            f"Vendor {vendor.pk} is in '{vendor.operational_status}' — "
            "must be 'waiting_marketing_approval' to approve."
        )

    now = timezone.now()
    vendor.marketing_status = MarketingStatus.APPROVED
    vendor.operational_status = OperationalStatus.ACTIVE
    vendor.approved_by_marketing = approved_by
    vendor.approved_at = now
    vendor.po_mandate_enabled = False
    vendor.save(update_fields=[
        "marketing_status", "operational_status",
        "approved_by_marketing", "approved_at", "po_mandate_enabled",
        "updated_at",
    ])

    # Transition linked submission
    if vendor.onboarding_submission_id:
        VendorOnboardingSubmission.objects.filter(pk=vendor.onboarding_submission_id).update(
            status=SubmissionStatus.ACTIVATED,
        )

    # Send portal activation (mandatory — raises on failure, rolls back transaction)
    activation_result = send_vendor_activation_for_vendor(vendor, actor=approved_by)
    contact_notice_result = send_vendor_contact_activation_notices(
        vendor,
        primary_vendor_email=activation_result["email"],
    )

    _build_audit_log(
        user=approved_by,
        action="vendor_marketing_approved",
        resource_type="Vendor",
        resource_id=vendor.pk,
        metadata={
            **activation_result,
            **contact_notice_result,
        },
    )

    return vendor


# ---------------------------------------------------------------------------
# 12. reject_vendor_marketing
# ---------------------------------------------------------------------------

@transaction.atomic
def reject_vendor_marketing(
    vendor: Vendor,
    rejected_by,
    note: str = "",
) -> Vendor:
    """
    Marketing rejects the vendor.

    - Sets marketing_status=rejected, operational_status=inactive.
    - Sets linked submission.status=rejected.

    Raises:
        VendorStateError — if vendor is not in waiting_marketing_approval
    """
    if vendor.operational_status != OperationalStatus.WAITING_MARKETING_APPROVAL:
        raise VendorStateError(
            f"Vendor {vendor.pk} is in '{vendor.operational_status}' — "
            "must be 'waiting_marketing_approval' to reject."
        )

    vendor.marketing_status = MarketingStatus.REJECTED
    vendor.operational_status = OperationalStatus.INACTIVE
    vendor.save(update_fields=["marketing_status", "operational_status", "updated_at"])

    if vendor.onboarding_submission_id:
        VendorOnboardingSubmission.objects.filter(pk=vendor.onboarding_submission_id).update(
            status=SubmissionStatus.REJECTED,
        )

    _build_audit_log(
        user=rejected_by,
        action="vendor_marketing_rejected",
        resource_type="Vendor",
        resource_id=vendor.pk,
        metadata={"note": note},
    )

    return vendor


# ---------------------------------------------------------------------------
# 13. assert_vendor_can_submit_invoice
# ---------------------------------------------------------------------------

def assert_vendor_can_submit_invoice(vendor: Vendor, po_number: str = None) -> None:
    """
    Gate check before a vendor submits an invoice.

    Raises:
        VendorStateError        — vendor is not active
        VendorProfileHoldError  — vendor has an active profile revision hold
    """
    if vendor.operational_status != OperationalStatus.ACTIVE:
        raise VendorStateError(
            f"Vendor {vendor.pk} is not active (status: '{vendor.operational_status}'). "
            "Cannot submit invoice."
        )
    assert_vendor_profile_not_on_hold(vendor)


# ---------------------------------------------------------------------------
# Private: finance token validation
# ---------------------------------------------------------------------------

def _get_valid_finance_token(
    token_str: str,
    expected_action: str,
) -> VendorFinanceActionToken:
    """
    Fetch and validate a finance action token.

    Raises:
        FinanceTokenError — if not found, wrong action, expired, or used
    """
    try:
        token = VendorFinanceActionToken.objects.select_related("submission").get(token=token_str)
    except VendorFinanceActionToken.DoesNotExist:
        raise FinanceTokenError("Invalid finance action token.")

    if token.action_type != expected_action:
        raise FinanceTokenError(
            f"Token is for action '{token.action_type}', expected '{expected_action}'."
        )
    if token.is_used():
        raise FinanceTokenError("This token has already been used.")
    if token.is_expired():
        raise FinanceTokenError("This token has expired.")

    return token


def _consume_submission_finance_tokens(
    submission: VendorOnboardingSubmission,
    used_at,
) -> None:
    """Mark all still-open finance action tokens for this submission as used."""
    submission.finance_tokens.filter(used_at__isnull=True).update(used_at=used_at)


# ---------------------------------------------------------------------------
# Finalize submission (public endpoint helper)
# ---------------------------------------------------------------------------

@transaction.atomic
def finalize_submission(
    submission: VendorOnboardingSubmission,
    submitted_by=None,
) -> VendorOnboardingSubmission:
    """
    Finalize an editable submission, triggering automatic finance review.

    Under Option B (auto-send-to-finance), this is the canonical finalization
    entry point for draft/reopened/correction submissions. It transitions the submission
    directly to sent_to_finance state.

    Raises:
        SubmissionStateError — if submission is not editable
        ValueError           — if minimum required fields are missing
    """
    if submission.status not in _EDITABLE_SUBMISSION_STATUSES:
        raise SubmissionStateError(
            f"Submission {submission.pk} is in '{submission.status}' — cannot finalize."
        )

    if not submission.normalized_vendor_name:
        raise ValueError("vendor_name is required before finalizing.")
    if not submission.normalized_email:
        raise ValueError("email is required before finalizing.")
    missing_attachments = get_missing_required_attachment_labels(submission)
    if missing_attachments:
        raise ValueError(
            "Required attachments missing: " + ", ".join(missing_attachments)
        )

    # Trigger auto-finance transition (sets status, creates tokens, sends emails)
    _start_finance_review(submission)

    # Mark invitation as submitted
    VendorInvitation.objects.filter(pk=submission.invitation_id).update(
        status=InvitationStatus.SUBMITTED,
    )

    _build_audit_log(
        user=submitted_by,
        action="vendor_submission_finalized",
        resource_type="VendorOnboardingSubmission",
        resource_id=submission.pk,
        metadata={},
    )

    return submission


# ---------------------------------------------------------------------------
# Vendor Profile Revision
# ---------------------------------------------------------------------------

class VendorProfileHoldError(ValueError):
    """Vendor is on profile hold — operation blocked until revision resolves."""


def assert_vendor_profile_not_on_hold(vendor: Vendor) -> None:
    """
    Raise VendorProfileHoldError if vendor has an active profile revision hold.
    Call at invoice submission, workflow approval, and finance action gate points.
    """
    if vendor.profile_change_pending:
        reason = vendor.profile_hold_reason or "A profile revision is pending review."
        raise VendorProfileHoldError(
            f"Vendor {vendor.pk} is on profile hold: {reason}"
        )


def build_vendor_live_snapshot(vendor: Vendor) -> dict:
    """
    Extract the approved live profile from the Vendor model directly.
    This is the authoritative source of truth for the vendor's approved profile.
    Snapshot keys match _VENDOR_PROFILE_SNAPSHOT_FIELDS keys.
    """
    snapshot = {key: getattr(vendor, field, None) for key, field in _VENDOR_PROFILE_SNAPSHOT_FIELDS.items()}
    submission = vendor.onboarding_submission
    if not submission:
        return snapshot

    fallback_pairs = {
        "title": submission.normalized_title,
        "vendor_name": submission.normalized_vendor_name,
        "vendor_type": submission.normalized_vendor_type,
        "email": submission.normalized_email,
        "phone": submission.normalized_phone,
        "fax": submission.normalized_fax,
        "region": submission.normalized_region,
        "head_office_no": submission.normalized_head_office_no,
        "gst_registered": submission.normalized_gst_registered,
        "gstin": submission.normalized_gstin,
        "pan": submission.normalized_pan,
        "address_line1": submission.normalized_address_line1,
        "address_line2": submission.normalized_address_line2,
        "address_line3": submission.normalized_address_line3,
        "city": submission.normalized_city,
        "state": submission.normalized_state,
        "country": submission.normalized_country,
        "pincode": submission.normalized_pincode,
        "preferred_payment_mode": submission.normalized_preferred_payment_mode,
        "beneficiary_name": submission.normalized_beneficiary_name,
        "beneficiary_account_number": submission.normalized_beneficiary_account_number,
        "bank_name": submission.normalized_bank_name,
        "bank_address": submission.normalized_bank_address,
        "bank_email": submission.normalized_bank_email,
        "account_number": submission.normalized_account_number,
        "bank_account_number": submission.normalized_bank_account_number,
        "bank_account_type": submission.normalized_bank_account_type,
        "ifsc": submission.normalized_ifsc,
        "micr_code": submission.normalized_micr_code,
        "neft_code": submission.normalized_neft_code,
        "bank_branch_address_line1": submission.normalized_bank_branch_address_line1,
        "bank_branch_address_line2": submission.normalized_bank_branch_address_line2,
        "bank_branch_city": submission.normalized_bank_branch_city,
        "bank_branch_state": submission.normalized_bank_branch_state,
        "bank_branch_country": submission.normalized_bank_branch_country,
        "bank_branch_pincode": submission.normalized_bank_branch_pincode,
        "bank_phone": submission.normalized_bank_phone,
        "bank_fax": submission.normalized_bank_fax,
        "authorized_signatory_name": submission.normalized_authorized_signatory_name,
        "msme_registered": submission.normalized_msme_registered,
        "msme_registration_number": submission.normalized_msme_registration_number,
        "msme_enterprise_type": submission.normalized_msme_enterprise_type,
        "declaration_accepted": submission.declaration_accepted,
        "contact_persons_json": submission.contact_persons_json,
        "head_office_address_json": submission.head_office_address_json,
        "tax_registration_details_json": submission.tax_registration_details_json,
    }

    for key, fallback in fallback_pairs.items():
        current = snapshot.get(key)
        if current in ("", None, [], {}) and fallback not in ("", None, [], {}):
            snapshot[key] = fallback
    return snapshot


def compute_changed_fields(proposed: dict, source: dict) -> list:
    """Compare two profile snapshots. Return sorted list of differing field keys."""
    return sorted(
        key for key in set(proposed) | set(source)
        if proposed.get(key) != source.get(key)
    )


def merge_profile_snapshot_with_source(proposed: dict, source: dict) -> dict:
    """
    Build a complete proposed snapshot using the source snapshot as baseline.

    The portal edit form submits only editable flat fields. Without merging,
    untouched nested/source fields appear as removed and inflate changed_fields_json.
    """
    merged = (source or {}).copy()
    merged.update(proposed or {})
    return merged


def rebase_profile_revision_source(proposed: dict, source: dict, live: dict) -> dict:
    """
    Refresh a stale revision source snapshot using the current live profile only
    for fields where the proposed draft already matches the live value.

    This keeps actual draft edits intact while collapsing false positives caused
    by old/empty source snapshots created before the vendor profile was fully set.
    """
    rebased = (source or {}).copy()
    all_keys = set(rebased) | set(proposed or {}) | set(live or {})
    for key in all_keys:
      if (proposed or {}).get(key) == (live or {}).get(key):
          rebased[key] = (live or {}).get(key)
    return rebased


@transaction.atomic
def get_or_create_editable_profile_revision(vendor: Vendor, actor=None) -> VendorProfileRevision:
    """
    Return the existing editable (draft/reopened) profile revision for vendor,
    or create a new one seeded with the current live snapshot.
    Does NOT place the vendor on hold — that happens at submit time.
    """
    existing = vendor.profile_revisions.filter(
        status__in=[VendorProfileRevisionStatus.DRAFT, VendorProfileRevisionStatus.REOPENED]
    ).order_by("-created_at").first()
    if existing:
        current_live_snapshot = build_vendor_live_snapshot(vendor)
        rebased_source_snapshot = rebase_profile_revision_source(
            existing.proposed_snapshot_json or {},
            existing.source_revision_snapshot_json or {},
            current_live_snapshot,
        )
        normalized_snapshot = merge_profile_snapshot_with_source(
            existing.proposed_snapshot_json or {},
            rebased_source_snapshot,
        )
        normalized_changed_fields = compute_changed_fields(
            normalized_snapshot,
            rebased_source_snapshot,
        )
        if (
            rebased_source_snapshot != (existing.source_revision_snapshot_json or {})
            or
            normalized_snapshot != (existing.proposed_snapshot_json or {})
            or normalized_changed_fields != (existing.changed_fields_json or [])
        ):
            existing.source_revision_snapshot_json = rebased_source_snapshot
            existing.proposed_snapshot_json = normalized_snapshot
            existing.changed_fields_json = normalized_changed_fields
            existing.save(update_fields=["source_revision_snapshot_json", "proposed_snapshot_json", "changed_fields_json", "updated_at"])
        return existing

    last = vendor.profile_revisions.order_by("-revision_number").first()
    revision_number = (last.revision_number + 1) if last else 1

    source_snapshot = build_vendor_live_snapshot(vendor)
    return VendorProfileRevision.objects.create(
        vendor=vendor,
        revision_number=revision_number,
        status=VendorProfileRevisionStatus.DRAFT,
        proposed_snapshot_json=source_snapshot.copy(),
        source_revision_snapshot_json=source_snapshot,
        changed_fields_json=[],
        created_by=actor,
        updated_by=actor,
    )


@transaction.atomic
def save_draft_profile_revision(
    revision: VendorProfileRevision,
    proposed_snapshot: dict,
    actor=None,
) -> VendorProfileRevision:
    """
    Update proposed_snapshot_json and recompute changed_fields_json.
    Revision must be draft or reopened.
    """
    if revision.status not in (VendorProfileRevisionStatus.DRAFT, VendorProfileRevisionStatus.REOPENED):
        raise SubmissionStateError(
            f"Revision {revision.pk} is in '{revision.status}' — cannot edit."
        )
    normalized_snapshot = merge_profile_snapshot_with_source(
        proposed_snapshot,
        revision.source_revision_snapshot_json or {},
    )
    revision.proposed_snapshot_json = normalized_snapshot
    revision.changed_fields_json = compute_changed_fields(
        normalized_snapshot, revision.source_revision_snapshot_json or {}
    )
    revision.updated_by = actor
    revision.save(update_fields=["proposed_snapshot_json", "changed_fields_json", "updated_by", "updated_at"])
    return revision


@transaction.atomic
def submit_profile_revision(revision: VendorProfileRevision, actor=None) -> VendorProfileRevision:
    """
    Submit a profile revision for internal review. Places the vendor on hold.
    Transitions: draft/reopened → submitted.
    """
    if revision.status not in (VendorProfileRevisionStatus.DRAFT, VendorProfileRevisionStatus.REOPENED):
        raise SubmissionStateError(
            f"Revision {revision.pk} is in '{revision.status}' — cannot submit."
        )
    if not revision.changed_fields_json:
        raise ValueError("No fields have changed — nothing to submit.")

    now = timezone.now()
    revision.status = VendorProfileRevisionStatus.SUBMITTED
    revision.submitted_at = now
    revision.updated_by = actor
    revision.save(update_fields=["status", "submitted_at", "updated_by", "updated_at"])

    vendor = revision.vendor
    vendor.profile_change_pending = True
    vendor.profile_hold_reason = f"Profile revision #{revision.revision_number} submitted for review."
    vendor.active_profile_revision = revision
    vendor.profile_hold_started_at = now
    vendor.save(update_fields=[
        "profile_change_pending", "profile_hold_reason",
        "active_profile_revision", "profile_hold_started_at", "updated_at",
    ])

    _build_audit_log(
        user=actor,
        action="vendor_profile_revision_submitted",
        resource_type="VendorProfileRevision",
        resource_id=revision.pk,
        metadata={"vendor_id": vendor.pk, "revision_number": revision.revision_number},
    )
    return revision


@transaction.atomic
def finance_approve_profile_revision(revision: VendorProfileRevision, actor=None) -> VendorProfileRevision:
    """
    Finance approves a submitted profile revision.
    Transitions to finance_approved; apply is the only valid final path after that.
    """
    if revision.status != VendorProfileRevisionStatus.SUBMITTED:
        raise SubmissionStateError(
            f"Revision {revision.pk} is in '{revision.status}' — finance cannot approve."
        )
    now = timezone.now()
    revision.status = VendorProfileRevisionStatus.FINANCE_APPROVED
    revision.approved_at = now
    revision.updated_by = actor
    revision.save(update_fields=["status", "approved_at", "updated_by", "updated_at"])

    _build_audit_log(
        user=actor,
        action="vendor_profile_revision_finance_approved",
        resource_type="VendorProfileRevision",
        resource_id=revision.pk,
        metadata={"vendor_id": revision.vendor_id},
    )
    return revision


@transaction.atomic
def finance_reject_profile_revision(
    revision: VendorProfileRevision,
    actor=None,
    note: str = "",
) -> VendorProfileRevision:
    """
    Finance rejects a submitted profile revision. Lifts vendor hold.
    """
    if revision.status != VendorProfileRevisionStatus.SUBMITTED:
        raise SubmissionStateError(
            f"Revision {revision.pk} is in '{revision.status}' — finance cannot reject."
        )
    revision.status = VendorProfileRevisionStatus.FINANCE_REJECTED
    revision.updated_by = actor
    revision.save(update_fields=["status", "updated_by", "updated_at"])

    _lift_vendor_profile_hold(revision.vendor)

    _build_audit_log(
        user=actor,
        action="vendor_profile_revision_finance_rejected",
        resource_type="VendorProfileRevision",
        resource_id=revision.pk,
        metadata={"vendor_id": revision.vendor_id, "note": note},
    )
    return revision


@transaction.atomic
def reopen_profile_revision(
    revision: VendorProfileRevision,
    actor=None,
    note: str = "",
) -> VendorProfileRevision:
    """
    Reopen a finance-rejected revision for vendor corrections.
    Transitions: finance_rejected → reopened. Re-places vendor on hold.
    """
    if revision.status != VendorProfileRevisionStatus.FINANCE_REJECTED:
        raise SubmissionStateError(
            f"Revision {revision.pk} is in '{revision.status}' — can only reopen from 'finance_rejected'."
        )
    now = timezone.now()
    revision.status = VendorProfileRevisionStatus.REOPENED
    revision.updated_by = actor
    revision.save(update_fields=["status", "updated_by", "updated_at"])

    vendor = revision.vendor
    vendor.profile_change_pending = True
    vendor.profile_hold_reason = f"Profile revision #{revision.revision_number} reopened for corrections."
    vendor.active_profile_revision = revision
    vendor.profile_hold_started_at = now
    vendor.save(update_fields=[
        "profile_change_pending", "profile_hold_reason",
        "active_profile_revision", "profile_hold_started_at", "updated_at",
    ])

    _build_audit_log(
        user=actor,
        action="vendor_profile_revision_reopened",
        resource_type="VendorProfileRevision",
        resource_id=revision.pk,
        metadata={"vendor_id": revision.vendor_id, "note": note},
    )
    return revision




@transaction.atomic
def apply_vendor_profile_revision(revision: VendorProfileRevision, actor=None) -> VendorProfileRevision:
    """
    Write proposed_snapshot_json fields directly onto the Vendor model.
    Vendor is the authoritative approved live profile source of truth.
    Mark the revision as applied and lift the vendor hold.
    """
    if revision.status != VendorProfileRevisionStatus.FINANCE_APPROVED:
        raise SubmissionStateError(
            f"Revision {revision.pk} is in '{revision.status}' — cannot apply."
        )
    snapshot = revision.proposed_snapshot_json or {}
    vendor = revision.vendor

    update_fields = ["updated_at"]
    for snapshot_key, vendor_field in _VENDOR_PROFILE_SNAPSHOT_FIELDS.items():
        if snapshot_key in snapshot:
            val = snapshot[snapshot_key]
            # Normalise booleans for BooleanFields
            if isinstance(val, str) and vendor._meta.get_field(vendor_field).get_internal_type() == "BooleanField":
                val = val.strip().lower() in ("yes", "true", "1", "y")
            elif val is None:
                val = ""
            elif not isinstance(val, bool) and not isinstance(val, (dict, list)):
                val = str(val).strip()
            current = getattr(vendor, vendor_field, None)
            if val != current:
                setattr(vendor, vendor_field, val)
                update_fields.append(vendor_field)

    now = timezone.now()
    revision.status = VendorProfileRevisionStatus.APPLIED
    revision.applied_at = now
    revision.updated_by = actor
    revision.save(update_fields=["status", "applied_at", "updated_by", "updated_at"])

    _lift_vendor_profile_hold(vendor, extra_update_fields=update_fields)

    _build_audit_log(
        user=actor,
        action="vendor_profile_revision_applied",
        resource_type="VendorProfileRevision",
        resource_id=revision.pk,
        metadata={"vendor_id": vendor.pk, "changed_fields": revision.changed_fields_json},
    )
    return revision


@transaction.atomic
def cancel_profile_revision(revision: VendorProfileRevision, actor=None) -> VendorProfileRevision:
    """
    Cancel an in-progress profile revision. Lifts vendor hold if held by this revision.
    """
    terminal = {
        VendorProfileRevisionStatus.APPLIED,
        VendorProfileRevisionStatus.CANCELLED,
    }
    if revision.status in terminal:
        raise SubmissionStateError(
            f"Revision {revision.pk} is already in terminal state '{revision.status}'."
        )
    revision.status = VendorProfileRevisionStatus.CANCELLED
    revision.updated_by = actor
    revision.save(update_fields=["status", "updated_by", "updated_at"])

    vendor = revision.vendor
    if vendor.active_profile_revision_id == revision.pk:
        _lift_vendor_profile_hold(vendor)

    _build_audit_log(
        user=actor,
        action="vendor_profile_revision_cancelled",
        resource_type="VendorProfileRevision",
        resource_id=revision.pk,
        metadata={"vendor_id": vendor.pk},
    )
    return revision


def _lift_vendor_profile_hold(vendor: Vendor, extra_update_fields: list = None) -> None:
    """Clear all profile hold fields on the Vendor record."""
    vendor.profile_change_pending = False
    vendor.profile_hold_reason = ""
    vendor.active_profile_revision = None
    vendor.profile_hold_started_at = None
    fields = [
        "profile_change_pending", "profile_hold_reason",
        "active_profile_revision", "profile_hold_started_at", "updated_at",
    ]
    if extra_update_fields:
        fields = list(set(fields) | set(extra_update_fields))
    vendor.save(update_fields=fields)
