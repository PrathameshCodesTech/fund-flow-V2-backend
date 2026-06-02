from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal
import hashlib
import io
import json
import logging
import re
import time
from typing import Any
from urllib import error as urllib_error
from urllib import request as urllib_request

from django.conf import settings
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone

from apps.invoices.models import Invoice, InvoiceDocument, InvoiceStatus, VendorInvoiceSubmission, VendorInvoiceSubmissionStatus
from apps.access.models import PermissionAction, PermissionResource
from apps.access.services import user_has_permission_including_ancestors


logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Submission validation result
# ---------------------------------------------------------------------------

@dataclass
class SubmissionValidationResult:
    is_valid: bool
    field_errors: dict[str, list[str]] = field(default_factory=dict)
    warnings: list[dict[str, str]] = field(default_factory=list)


class SubmissionValidationError(ValueError):
    """Raised when submission validation fails with field-level errors."""
    def __init__(self, result: SubmissionValidationResult):
        self.result = result
        super().__init__("Submission validation failed.")


# ---------------------------------------------------------------------------
# Workflow route validation helpers
# ---------------------------------------------------------------------------

def _validate_workflow_route_for_submission(submission, send_to_route):
    """
    Returns (published_version, first_step) if the route is viable for the
    submission's scope node, otherwise raises VendorRouteError.

    Checks:
    1. route.is_active
    2. route.workflow_template.is_active
    3. route.workflow_template.module == "invoice"
    4. Published WorkflowTemplateVersion exists
    5. Template scope is at or above submission's scope_node (ancestry check)
    6. First actionable step has at least one eligible actor
       (default_user active first, else users with required_role at resolved node)
    """
    from apps.core.services import get_ancestors
    from apps.access.selectors import get_users_with_role_at_node
    from apps.workflow.services import get_first_actionable_step, resolve_step_target_node
    from apps.workflow.models import VersionStatus

    if not send_to_route.is_active:
        raise VendorRouteError(
            f"The selected 'Send To' option '{send_to_route.label}' is not active."
        )

    template = send_to_route.workflow_template
    if not template.is_active:
        raise VendorRouteError(
            f"Route '{send_to_route.label}' is misconfigured: "
            "the mapped workflow template is not active. Contact your administrator."
        )

    if template.module != "invoice":
        raise VendorRouteError(
            f"Route '{send_to_route.label}' is misconfigured: "
            "the mapped workflow template is not an invoice workflow. "
            "Contact your administrator."
        )

    published_version = (
        template.versions.filter(status=VersionStatus.PUBLISHED)
        .order_by("-version_number")
        .first()
    )
    if published_version is None:
        raise VendorRouteError(
            f"Route '{send_to_route.label}' is misconfigured: "
            "the mapped workflow template has no published version. "
            "Contact your administrator."
        )

    # Scope ancestry: template's scope_node must be at or above submission's scope_node
    template_scope = template.scope_node
    submission_scope = submission.scope_node
    if template_scope.pk != submission_scope.pk:
        ancestor_paths = submission_scope.get_ancestors_from_path()
        if template_scope.path not in ancestor_paths:
            raise VendorRouteError(
                f"Route '{send_to_route.label}' is not available for invoices "
                f"at '{submission_scope.code}'. The workflow is scoped to "
                f"'{template_scope.code}' which is not an ancestor of this node."
            )

    # Actor viability: first actionable step must have an eligible actor
    first_step = get_first_actionable_step(published_version)
    if first_step is None:
        raise VendorRouteError(
            f"Route '{send_to_route.label}' is misconfigured: "
            "the workflow has no actionable steps. Contact your administrator."
        )

    resolved_node = resolve_step_target_node(first_step, submission_scope)
    if first_step.default_user and first_step.default_user.is_active:
        pass  # valid — default user takes precedence
    elif get_users_with_role_at_node(first_step.required_role, resolved_node).exists():
        pass  # valid — at least one active user has the required role
    else:
        raise VendorRouteError(
            f"Route '{send_to_route.label}' has no eligible approvers configured "
            f"for the first step ('{first_step.name}'). "
            "Assign someone with the role to proceed."
        )

    return published_version, first_step


def validate_vendor_submission_for_submit(submission, send_to_route) -> SubmissionValidationResult:
    """
    Run full submission validation before submit:
      - Blocking (A): submission/vendor/normalized_data field checks
      - Blocking (B): duplicate detection (returned as field_errors)
      - Blocking (C): workflow route viability (raises VendorRouteError)
      - Warnings: non-blocking feedback (low confidence, far due date, missing description)
    Returns SubmissionValidationResult with is_valid, field_errors, warnings.
    """
    nd = submission.normalized_data or {}
    vendor = submission.vendor
    field_errors: dict[str, list[str]] = {}
    warnings: list[dict[str, str]] = []
    invoice_date = None

    # ── A: Submission / vendor / data checks ──────────────────────────────────

    if not nd:
        field_errors["_normalized_data"] = [
            "Submission has no extracted data. Please upload a valid invoice document."
        ]

    if vendor:
        if vendor.operational_status != "active":
            field_errors["vendor_status"] = [
                f"Vendor is currently {vendor.operational_status}. "
                "Only active vendors can receive invoices."
            ]
    else:
        field_errors["vendor"] = ["No vendor is associated with this submission."]

    # Invoice number
    if not nd.get("vendor_invoice_number"):
        field_errors.setdefault("vendor_invoice_number", []).append(
            "Invoice number is required."
        )

    # Invoice date
    invoice_date_str = nd.get("invoice_date", "")
    if not invoice_date_str:
        field_errors.setdefault("invoice_date", []).append("Invoice date is required.")
    else:
        from django.utils.dateparse import parse_date
        invoice_date = parse_date(invoice_date_str)
        if invoice_date is None:
            field_errors.setdefault("invoice_date", []).append(
                "Invoice date must be in YYYY-MM-DD format."
            )

    # Currency
    currency = nd.get("currency", "")
    if not currency:
        field_errors.setdefault("currency", []).append("Currency is required.")
    elif not re.match(r"^[A-Z]{3}$", currency):
        field_errors.setdefault("currency", []).append(
            "Currency must be a 3-letter uppercase code (e.g., INR, USD)."
        )

    # Total amount
    total_str = nd.get("total_amount")
    if not total_str:
        field_errors.setdefault("total_amount", []).append("Total amount is required.")
    else:
        try:
            total = Decimal(str(total_str))
            if total <= 0:
                field_errors.setdefault("total_amount", []).append(
                    "Total amount must be greater than zero."
                )
        except Exception:
            field_errors.setdefault("total_amount", []).append(
                "Total amount must be a valid number."
            )

    # Subtotal + tax vs total consistency
    sub_str = nd.get("subtotal_amount")
    tax_str = nd.get("tax_amount")
    tot_str = nd.get("total_amount")
    if sub_str and tax_str and tot_str:
        try:
            sub = Decimal(str(sub_str))
            tax = Decimal(str(tax_str))
            tot = Decimal(str(tot_str))
            if abs(sub + tax - tot) > Decimal("0.02"):
                field_errors.setdefault("total_amount", []).append(
                    "Total amount does not match subtotal + tax. "
                    "Please review and correct."
                )
        except Exception:
            pass  # already validated as number above

    # Due date
    due_date_str = nd.get("due_date")
    if due_date_str:
        from django.utils.dateparse import parse_date
        due_date = parse_date(due_date_str)
        if due_date is None:
            field_errors.setdefault("due_date", []).append(
                "Due date must be in YYYY-MM-DD format."
            )
        elif invoice_date and due_date < invoice_date:
            field_errors.setdefault("due_date", []).append(
                "Due date cannot be before invoice date."
            )

    # ── B: Duplicate checks (returned as field errors) ───────────────────────

    if nd.get("vendor_invoice_number"):
        # Check active submissions (exclude cancelled/rejected)
        dup_sub = VendorInvoiceSubmission.objects.filter(
            vendor=vendor,
            normalized_data__vendor_invoice_number=nd["vendor_invoice_number"],
        ).exclude(
            status__in=[
                VendorInvoiceSubmissionStatus.CANCELLED,
                VendorInvoiceSubmissionStatus.REJECTED,
            ]
        ).exclude(pk=submission.pk).exists()

        if dup_sub:
            field_errors.setdefault("vendor_invoice_number", []).append(
                "An invoice with this reference has already been submitted."
            )

        # Check active invoices (exclude rejected/paid)
        dup_inv = Invoice.objects.filter(
            vendor=vendor,
            vendor_invoice_number=nd["vendor_invoice_number"],
        ).exclude(
            status__in=[InvoiceStatus.REJECTED, InvoiceStatus.PAID]
        ).exists()

        if dup_inv:
            field_errors.setdefault("vendor_invoice_number", []).append(
                "An invoice with this reference already exists in the system."
            )

    # ── C: Workflow route viability ─────────────────────────────────────────

    if not field_errors:
        # Only validate route if no blocking field errors
        # (route validation raises VendorRouteError which bubbles to caller)
        _validate_workflow_route_for_submission(submission, send_to_route)

    # ── Warnings (non-blocking) ─────────────────────────────────────────────

    conf = getattr(submission, "confidence_score", None)
    if conf is not None and conf < 0.7:
        warnings.append({
            "code": "low_confidence",
            "message": "Extraction confidence is low. Please review all fields carefully.",
        })

    if due_date_str and invoice_date:
        from django.utils.dateparse import parse_date
        due_date = parse_date(due_date_str)
        if due_date and invoice_date:
            if (due_date - invoice_date).days > 90:
                warnings.append({
                    "code": "due_date_far_future",
                    "message": "Due date is unusually far in the future.",
                })

    desc = nd.get("description", "").strip()
    if not desc:
        warnings.append({
            "code": "missing_description",
            "message": "Description is missing.",
        })

    return SubmissionValidationResult(
        is_valid=not field_errors,
        field_errors=field_errors,
        warnings=warnings,
    )


# ---------------------------------------------------------------------------
# Core invoice service (preserved from original)
# ---------------------------------------------------------------------------

@transaction.atomic
def create_invoice(
    title,
    amount,
    currency,
    scope_node,
    created_by,
    po_number: str = "",
    vendor=None,
    enforce_permission: bool = True,
):
    """
    Create a new invoice if the creator has CREATE permission on INVOICE
    at scope_node or any ancestor.

    PO number is optional invoice metadata.
    """
    if enforce_permission and not user_has_permission_including_ancestors(
        created_by, PermissionAction.CREATE, PermissionResource.INVOICE, scope_node
    ):
        raise InvoicePermissionError(
            f"User {created_by} does not have create:invoice permission "
            f"at node {scope_node} or any ancestor."
        )
    return Invoice.objects.create(
        title=title,
        amount=amount,
        currency=currency,
        scope_node=scope_node,
        po_number=po_number,
        vendor=vendor,
        created_by=created_by,
        status=InvoiceStatus.DRAFT,
    )


# ---------------------------------------------------------------------------
# Public invoice services (preserved)
# ---------------------------------------------------------------------------

class InvoicePermissionError(PermissionError):
    """Raised when the actor lacks the required permission on an invoice operation."""


class InvoicePOMandateError(ValueError):
    """Legacy exception kept for compatibility; PO numbers are optional."""


def sync_invoice_status(invoice, new_status):
    """Directly update invoice status. Used by workflow sync."""
    invoice.status = new_status
    invoice.save(update_fields=["status", "updated_at"])
    return invoice


# ---------------------------------------------------------------------------
# Extraction result
# ---------------------------------------------------------------------------

@dataclass
class ExtractionResult:
    raw_cells: dict[str, Any]
    normalized: dict[str, Any]
    confidence: float
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Excel Invoice Extractor
# ---------------------------------------------------------------------------

# Known label → normalized field name
EXCEL_FIELD_MAP = {
    "invoice number": "vendor_invoice_number",
    "invoice #": "vendor_invoice_number",
    "inv #": "vendor_invoice_number",
    "inv number": "vendor_invoice_number",
    "invoice no": "vendor_invoice_number",
    "invoice no.": "vendor_invoice_number",
    "invoice date": "invoice_date",
    "inv date": "invoice_date",
    "date": "invoice_date",
    "due date": "due_date",
    "payment due date": "due_date",
    "po number": "po_number",
    "po #": "po_number",
    "po no": "po_number",
    "po no.": "po_number",
    "purchase order number": "po_number",
    "purchase order no": "po_number",
    "currency": "currency",
    "subtotal": "subtotal_amount",
    "sub total": "subtotal_amount",
    "amount": "subtotal_amount",
    "total before tax": "subtotal_amount",
    "subtotal amount": "subtotal_amount",
    "tax": "tax_amount",
    "tax amount": "tax_amount",
    "gst": "tax_amount",
    "gst amount": "tax_amount",
    "total": "total_amount",
    "grand total": "total_amount",
    "total amount": "total_amount",
    "invoice total": "total_amount",
    "total amount (inr)": "total_amount",
    "description": "description",
    "bill to": "bill_to_name",
    "billed to": "bill_to_name",
    "vendor name": "vendor_name",
    "gstin": "gstin",
    "gst in": "gstin",
    "pan": "pan",
    "invoice number *": "vendor_invoice_number",
    "invoice date *": "invoice_date",
    "due date *": "due_date",
    "po number *": "po_number",
    "currency *": "currency",
    "subtotal amount *": "subtotal_amount",
    "tax amount *": "tax_amount",
    "total amount *": "total_amount",
}


def _normalize_excel_label(label: str) -> str:
    return re.sub(r"[^a-z0-9]", "", label.lower())


def _try_parse_date(value: Any) -> str | None:
    """Parse a date value from Excel — returns ISO date string or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%b %d, %Y", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    # Try Excel serial number
    try:
        serial = float(s)
        if 1 <= serial <= 50000:
            from openpyxl.utils.datetime import from_excel
            dt = from_excel(serial)
            if dt:
                return dt.date().isoformat()
    except Exception:
        pass
    return None


def _try_parse_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    s = str(value).strip().replace(",", "").replace(" ", "")
    try:
        return Decimal(s)
    except Exception:
        return None


def _guess_field(key: str) -> str | None:
    norm = _normalize_excel_label(key)
    return EXCEL_FIELD_MAP.get(norm) or EXCEL_FIELD_MAP.get(key.lower().strip())


def _json_safe_dict(d: dict) -> dict:
    """Convert Decimal/other non-JSON-serializable values to JSON-safe types."""
    result = {}
    for k, v in d.items():
        if isinstance(v, Decimal):
            result[k] = float(v)
        elif isinstance(v, dict):
            result[k] = _json_safe_dict(v)
        elif isinstance(v, list):
            result[k] = [
                float(x) if isinstance(x, Decimal) else x for x in v
            ]
        else:
            result[k] = v
    return result


def _azure_docintel_configured() -> bool:
    return bool(
        getattr(settings, "AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT", "").strip()
        and getattr(settings, "AZURE_DOCUMENT_INTELLIGENCE_KEY", "").strip()
    )


def _azure_extract_field_value(field: dict[str, Any]) -> Any:
    if not isinstance(field, dict):
        return None

    # Prefer semantic value types first.
    for key in ("valueString", "valueDate", "valueTime", "valuePhoneNumber", "valueInteger"):
        value = field.get(key)
        if value not in (None, ""):
            return value

    value_number = field.get("valueNumber")
    if value_number not in (None, ""):
        return Decimal(str(value_number))

    value_currency = field.get("valueCurrency")
    if isinstance(value_currency, dict):
        amount = value_currency.get("amount")
        code = value_currency.get("currencyCode")
        if amount not in (None, ""):
            return Decimal(str(amount))
        if code:
            return code

    value_array = field.get("valueArray")
    if isinstance(value_array, list):
        return value_array

    value_object = field.get("valueObject")
    if isinstance(value_object, dict):
        return value_object

    content = field.get("content")
    if content not in (None, ""):
        return content
    return None


def _azure_extract_description(fields: dict[str, Any]) -> str | None:
    items = fields.get("Items")
    item_values = _azure_extract_field_value(items)
    if not isinstance(item_values, list):
        return None

    descriptions: list[str] = []
    for item in item_values:
        value_object = item.get("valueObject") if isinstance(item, dict) else None
        if not isinstance(value_object, dict):
            continue
        desc_field = value_object.get("Description")
        desc = _azure_extract_field_value(desc_field) if isinstance(desc_field, dict) else None
        if desc:
            descriptions.append(str(desc).strip())
    if descriptions:
        return "; ".join(descriptions[:5])
    return None


def _clean_party_name(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = " ".join(str(value).split()).strip()
    if not text:
        return None

    text = re.sub(r"^[\d\W_]+", "", text).strip()
    text = re.sub(
        r"^(?:vendor\s*name|customer\s*name|bill\s*to|billed\s*to|buyer|consignee|recipient)\s*[:\-]?\s*",
        "",
        text,
        flags=re.IGNORECASE,
    ).strip()
    text = re.sub(r"\s{2,}", " ", text)
    return text or None


def _normalize_extracted_invoice_fields(
    normalized: dict[str, Any],
    submission: VendorInvoiceSubmission | None = None,
) -> dict[str, Any]:
    cleaned = dict(normalized)

    for key in ("vendor_name", "bill_to_name"):
        if key in cleaned:
            cleaned_value = _clean_party_name(cleaned.get(key))
            if cleaned_value:
                cleaned[key] = cleaned_value
            else:
                cleaned.pop(key, None)

    if submission is not None:
        known_vendor_name = _clean_party_name(getattr(submission.vendor, "vendor_name", None))
        if known_vendor_name and not _is_meaningful_party_name(cleaned.get("vendor_name")):
            cleaned["vendor_name"] = known_vendor_name

    return cleaned


def _map_azure_invoice_fields(fields: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}

    field_map = {
        "InvoiceId": "vendor_invoice_number",
        "InvoiceDate": "invoice_date",
        "DueDate": "due_date",
        "PurchaseOrder": "po_number",
        "SubTotal": "subtotal_amount",
        "TotalTax": "tax_amount",
        "InvoiceTotal": "total_amount",
        "VendorName": "vendor_name",
        "CustomerName": "bill_to_name",
        "CustomerAddressRecipient": "bill_to_name",
    }

    for azure_name, normalized_name in field_map.items():
        field = fields.get(azure_name)
        if not isinstance(field, dict):
            continue
        value = _azure_extract_field_value(field)
        if value in (None, ""):
            continue
        normalized[normalized_name] = value

    currency_field = fields.get("InvoiceTotal") or fields.get("AmountDue") or fields.get("SubTotal")
    if isinstance(currency_field, dict):
        currency_info = currency_field.get("valueCurrency")
        if isinstance(currency_info, dict) and currency_info.get("currencyCode"):
            normalized["currency"] = currency_info["currencyCode"]

    description = _azure_extract_description(fields)
    if description:
        normalized["description"] = description

    return _normalize_extracted_invoice_fields(normalized)


def _extract_pdf_with_azure_document_intelligence(submission: VendorInvoiceSubmission) -> ExtractionResult | None:
    if not _azure_docintel_configured():
        return None

    endpoint = settings.AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT.strip().rstrip("/")
    key = settings.AZURE_DOCUMENT_INTELLIGENCE_KEY.strip()
    analyze_url = (
        f"{endpoint}/formrecognizer/v2.1/prebuilt/invoice/analyze"
        "?includeTextDetails=true"
    )

    try:
        with submission.source_file.open("rb") as f:
            payload = f.read()

        post_req = urllib_request.Request(
            analyze_url,
            data=payload,
            method="POST",
            headers={
                "Ocp-Apim-Subscription-Key": key,
                "Content-Type": "application/pdf",
            },
        )
        with urllib_request.urlopen(post_req, timeout=60) as resp:
            operation_location = resp.headers.get("operation-location") or resp.headers.get("Operation-Location")
        if not operation_location:
            raise RuntimeError("Azure Document Intelligence did not return an operation URL.")

        poll_headers = {"Ocp-Apim-Subscription-Key": key}
        result_payload = None
        for _ in range(30):
            poll_req = urllib_request.Request(operation_location, headers=poll_headers, method="GET")
            with urllib_request.urlopen(poll_req, timeout=60) as poll_resp:
                result_payload = json.loads(poll_resp.read().decode("utf-8"))
            status = str(result_payload.get("status", "")).lower()
            if status == "succeeded":
                break
            if status == "failed":
                error_detail = result_payload.get("error") or {}
                raise RuntimeError(f"Azure analysis failed: {error_detail}")
            time.sleep(1)
        else:
            raise RuntimeError("Azure analysis timed out while polling for invoice results.")

        analyze_result = result_payload.get("analyzeResult") or {}
        document_results = analyze_result.get("documentResults") or []
        if not document_results:
            raise RuntimeError("Azure returned no invoice document results.")
        first_doc = document_results[0] or {}
        fields = first_doc.get("fields") or {}
        normalized = _map_azure_invoice_fields(fields)
        confidence = _score_confidence(normalized)

        raw_cells = {
            "extraction_method": "azure_document_intelligence",
            "azure_status": result_payload.get("status"),
            "azure_model": "prebuilt-invoice",
            "azure_fields": fields,
        }
        warnings: list[str] = []
        errors: list[str] = []
        if not normalized:
            warnings.append("Azure extraction did not return usable invoice fields.")
            errors.append("Azure extraction returned no usable invoice fields.")
        elif confidence < 0.5:
            warnings.append("Azure extraction returned only partial invoice data. Please review carefully.")

        return ExtractionResult(
            raw_cells=raw_cells,
            normalized={k: v for k, v in normalized.items() if v not in (None, "")},
            confidence=confidence,
            warnings=warnings,
            errors=errors,
        )
    except urllib_error.HTTPError as exc:
        detail = ""
        try:
            detail = exc.read().decode("utf-8", errors="replace")
        except Exception:
            detail = str(exc)
        raise RuntimeError(f"Azure Document Intelligence HTTP {exc.code}: {detail}") from exc
    except urllib_error.URLError as exc:
        raise RuntimeError(f"Azure Document Intelligence connection failed: {exc.reason}") from exc


def _is_meaningful_party_name(value: Any) -> bool:
    if value in (None, ""):
        return False
    text = " ".join(str(value).split()).strip()
    if len(text) < 4:
        return False
    if len(re.findall(r"[A-Za-z]", text)) < 4:
        return False
    lowered = text.lower()
    bad_markers = (
        "government",
        "signature",
        "authorised signatory",
        "authorized signature",
        "ship to",
        "bill to",
    )
    return not any(marker in lowered for marker in bad_markers)


def _decimal_or_none(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except Exception:
        return None


def _amount_set_quality(values: dict[str, Any]) -> tuple[int, Decimal]:
    subtotal = _decimal_or_none(values.get("subtotal_amount"))
    tax = _decimal_or_none(values.get("tax_amount"))
    total = _decimal_or_none(values.get("total_amount"))
    present = sum(v is not None for v in (subtotal, tax, total))
    residual = Decimal("999999999")
    if subtotal is not None and tax is not None and total is not None:
        residual = abs((subtotal + tax) - total)
    return present, residual


def _should_prefer_fallback_amounts(primary_values: dict[str, Any], fallback_values: dict[str, Any]) -> bool:
    primary_present, primary_residual = _amount_set_quality(primary_values)
    fallback_present, fallback_residual = _amount_set_quality(fallback_values)
    primary_total = _decimal_or_none(primary_values.get("total_amount"))
    fallback_total = _decimal_or_none(fallback_values.get("total_amount"))

    if fallback_present < 2:
        return False
    if primary_present < 2:
        return True
    if primary_total is not None and fallback_total is not None:
        delta = abs(primary_total - fallback_total)
        if delta > Decimal("1000") and delta > (primary_total * Decimal("0.05")):
            return False

    if fallback_residual <= Decimal("0.05") and primary_residual > Decimal("0.05"):
        return True
    if fallback_residual < primary_residual and (primary_residual - fallback_residual) > Decimal("0.50"):
        return True
    return False


def _merge_extraction_results(
    primary: ExtractionResult,
    fallback: ExtractionResult,
    submission: VendorInvoiceSubmission | None = None,
) -> ExtractionResult:
    merged = dict(primary.normalized)
    fallback_normalized = fallback.normalized or {}

    for key, value in fallback_normalized.items():
        if merged.get(key) in (None, ""):
            merged[key] = value

    # Prefer cleaner party names from fallback when primary looks weak.
    for key in ("vendor_name", "bill_to_name"):
        primary_value = primary.normalized.get(key)
        fallback_value = fallback_normalized.get(key)
        if not _is_meaningful_party_name(primary_value) and _is_meaningful_party_name(fallback_value):
            merged[key] = fallback_value

    if _should_prefer_fallback_amounts(primary.normalized, fallback_normalized):
        for key in ("subtotal_amount", "tax_amount", "total_amount"):
            fallback_value = fallback_normalized.get(key)
            if fallback_value not in (None, ""):
                merged[key] = fallback_value

    raw_text = fallback.raw_cells.get("raw_text") if isinstance(fallback.raw_cells, dict) else None
    if isinstance(raw_text, str) and raw_text.strip():
        _correct_tax_from_totals_when_split_marked(merged, raw_text)

    merged = _normalize_extracted_invoice_fields(merged, submission=submission)

    confidence = _score_confidence(merged)
    raw_cells = dict(primary.raw_cells)
    raw_cells["fallback_extraction_method"] = fallback.raw_cells.get("extraction_method")
    raw_cells["fallback_used"] = True

    warnings = list(primary.warnings)
    warnings.extend(w for w in fallback.warnings if w not in warnings)
    errors = list(primary.errors)

    if confidence < primary.confidence:
        confidence = primary.confidence

    return ExtractionResult(
        raw_cells=raw_cells,
        normalized={k: v for k, v in merged.items() if v not in (None, "")},
        confidence=confidence,
        warnings=warnings,
        errors=errors,
    )


PDF_FIELD_SEQUENCE = [
    "vendor_invoice_number",
    "invoice_date",
    "due_date",
    "currency",
    "description",
    "subtotal_amount",
    "tax_amount",
    "total_amount",
    "po_number",
]


def _is_pdf_placeholder_line(line: str) -> bool:
    text = line.strip()
    if not text:
        return True
    lowered = text.lower()
    if lowered in {"invoice details", "amounts", "how to submit", "completed sample"}:
        return True
    if text in {"1", "2", "3", "4"}:
        return True
    return (
        lowered.startswith("e.g.")
        or lowered.startswith("yyyy-mm-dd")
        or lowered.startswith("numeric value")
        or lowered.startswith("tax portion")
        or lowered.startswith("grand total")
        or lowered.startswith("required if")
        or lowered.startswith("note:")
        or lowered.startswith("support to confirm")
        or lowered.startswith("fill in all fields")
        or lowered.startswith("log in to the vendor portal")
        or lowered.startswith("attach your invoice file")
        or lowered.startswith("submit ")
        or lowered.startswith("fundflow vendor portal")
        or lowered.startswith("vims vendor portal")
        or lowered.startswith("recommended:")
        or lowered.startswith("brief description")
    )


# ---------------------------------------------------------------------------
# PDF classification
# ---------------------------------------------------------------------------

_DIGITAL_TEXT_MIN_CHARS = 100  # fewer chars → treat as scanned


def _classify_pdf_text(raw_text: str) -> str:
    """
    Returns 'template', 'digital', or 'scanned' based on text density and
    presence of Horizon/VIMS template markers.
    """
    stripped = raw_text.strip()
    if len(stripped) < _DIGITAL_TEXT_MIN_CHARS:
        return "scanned"
    if ("VIMS" in raw_text or "Horizon" in raw_text) and "Vendor Invoice Submission Template" in raw_text:
        return "template"
    return "digital"


# ---------------------------------------------------------------------------
# OCR fallback — PyMuPDF + Tesseract (graceful degradation)
# ---------------------------------------------------------------------------


def _ocr_pdf_pages(file_obj) -> str:
    """
    Rasterize each page with PyMuPDF and run Tesseract OCR.
    Returns concatenated text, or empty string when either library or the
    Tesseract binary is absent.
    """
    try:
        import fitz  # PyMuPDF
    except ImportError:
        logger.debug("PyMuPDF (fitz) not installed — OCR skipped")
        return ""

    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        logger.debug("pytesseract / Pillow not installed — OCR skipped")
        return ""

    try:
        pytesseract.get_tesseract_version()
    except Exception:
        logger.warning("Tesseract binary not found or not reachable — OCR skipped")
        return ""

    try:
        file_obj.seek(0)
        raw_bytes = file_obj.read()
        file_obj.seek(0)
        doc = fitz.open(stream=raw_bytes, filetype="pdf")
        pages_text: list[str] = []
        for page in doc:
            mat = fitz.Matrix(200 / 72, 200 / 72)  # 200 DPI
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
            pages_text.append(pytesseract.image_to_string(img, config="--psm 6"))
        doc.close()
        return "\n".join(pages_text)
    except Exception as exc:
        logger.warning("OCR extraction failed: %s", exc)
        return ""


# ---------------------------------------------------------------------------
# Regex battery for arbitrary invoice PDFs
# ---------------------------------------------------------------------------

_PDF_REGEX_PATTERNS: list[tuple[str, str]] = [
    ("vendor_invoice_number",
     r"(?:invoice\s*(?:no\.?|number|#|num)|inv\w*\s*(?:no\.?|#|number|num))"
     r"[:\s#]*([A-Z0-9][A-Z0-9\-\/\.]{2,})"),
    ("vendor_invoice_number",
     r"(?:tax\s*invoice)\s+([A-Z0-9][A-Z0-9\-\/\.]{2,})"),
    ("invoice_date",
     r"(?:invoice\s*date|inv\.?\s*date|date\s*of\s*invoice|bill\s*date)"
     r"[:\s]*(\d{1,2}[-/\.]\d{1,2}[-/\.]\d{2,4}|\d{4}[-/\.]\d{1,2}[-/\.]\d{1,2}"
     r"|[A-Za-z]{3,9}\s+\d{1,2},?\s+\d{4})"),
    ("invoice_date",
     r"(?:dated|date)\s*[:\s]*(\d{1,2}[-/\.][A-Za-z]{3,9}[-/\.]\d{2,4}|\d{1,2}[-/\.]\d{1,2}[-/\.]\d{2,4}|\d{4}[-/\.]\d{1,2}[-/\.]\d{1,2})"),
    ("due_date",
     r"(?:due\s*date|payment\s*due(?:\s*date)?|pay\s*by)"
     r"[:\s]*(\d{1,2}[-/\.]\d{1,2}[-/\.]\d{2,4}|\d{4}[-/\.]\d{1,2}[-/\.]\d{1,2})"),
    ("currency",
     r"(?:currency)[:\s]*([A-Z]{3})\b"),
    ("po_number",
     r"(?:p\.?o\.?\s*(?:no\.?|number|#)|purchase\s*order\s*(?:no\.?|number|#)?)"
     r"[:\s-]*([A-Z0-9][A-Z0-9\-\/\.]{2,})"),
    ("total_amount",
     r"(?:grand\s*total|invoice\s*total|total\s*amount(?:\s*after\s*tax)?|total\s*due|amount\s*due|balance\s*due|net\s*total|amount\s*chargeable)"
     r"[^\d\n]*([\d,]+\.?\d*)"),
    ("subtotal_amount",
     r"(?:sub\s*total|subtotal|amount\s*before\s*tax|taxable\s*amount|total\s*value\s*of\s*services)"
     r"[^\d\n]*([\d,]+\.?\d*)"),
    ("tax_amount",
     r"(?:total\s*tax\s*amount|tax\s*amount|gst\s*amount|vat\s*amount)"
     r"[^\d\n]*([\d,]+\.?\d*)"),
    ("bill_to_name",
     r"(?:bill\s*to|billed\s*to|invoiced\s*to|client)[:\s]*\n?\s*([A-Z][A-Za-z0-9\s,\.&]{3,60})"),
    ("vendor_name",
     r"(?:from|vendor|supplier|sold\s*by|company\s*name)[:\s]*\n?\s*([A-Z][A-Za-z0-9\s,\.&]{3,60})"),
]

_PDF_DATE_FORMATS = (
    "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y",
    "%d.%m.%Y", "%Y/%m/%d", "%d/%m/%y", "%d-%m-%y", "%d.%m.%y",
    "%d-%b-%Y", "%d-%b-%y", "%d %b %Y", "%d %b %y",
    "%B %d, %Y", "%b %d, %Y", "%d %B %Y",
)

def _parse_pdf_date(s: str) -> str | None:
    s = re.sub(r"(\d)(st|nd|rd|th)\b", r"\1", s.strip(), flags=re.IGNORECASE)
    for fmt in _PDF_DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


_TAX_COMPONENT_LABEL_RE = re.compile(
    r"\b(?:add\s+)?(?:cgst|sgst|igst|utgst|cess|service\s*tax)\b",
    re.IGNORECASE,
)


def _extract_amount_candidates(text: str) -> list[Decimal]:
    candidates: list[Decimal] = []
    for match in re.finditer(r"(?<![A-Z0-9])([\d,]+(?:\.\d{1,2})?)(?!\s*%)", text):
        amount = _try_parse_decimal(match.group(1))
        if amount is not None:
            candidates.append(amount)
    return candidates


def _pick_tax_component_amount(line: str) -> Decimal | None:
    candidates = _extract_amount_candidates(line)
    if not candidates:
        return None

    # Tax-component lines often include a rate and an amount. Prefer the last
    # candidate because layouts commonly end the row with the actual tax value.
    picked = candidates[-1]

    # If the last candidate is implausibly tiny but a larger later candidate
    # exists in the line, prefer the maximum. This protects OCR-ish layouts.
    if picked <= Decimal("100") and any(c > picked for c in candidates):
        picked = max(candidates)

    return picked


def _sum_tax_components(raw_text: str) -> Decimal | None:
    lines = [" ".join(line.split()) for line in raw_text.splitlines()]
    total = Decimal("0")
    matched = 0

    def _looks_like_tax_noise(text: str) -> bool:
        lower = text.lower()
        return "gstin" in lower or "gst number" in lower or lower.startswith("gst:")

    for idx, compact in enumerate(lines):
        if not compact:
            continue
        if _looks_like_tax_noise(compact):
            continue
        if not _TAX_COMPONENT_LABEL_RE.search(compact):
            continue
        amount = _pick_tax_component_amount(compact)
        if amount is None:
            for next_idx in range(idx + 1, min(idx + 3, len(lines))):
                candidate_line = lines[next_idx]
                if not candidate_line or _looks_like_tax_noise(candidate_line):
                    continue
                if _TAX_COMPONENT_LABEL_RE.search(candidate_line):
                    break
                candidate_amount = _pick_tax_component_amount(candidate_line)
                if candidate_amount is not None:
                    amount = candidate_amount
                    break
        if amount is None:
            continue
        total += amount
        matched += 1
    return total if matched >= 1 else None


def _count_tax_component_markers(raw_text: str) -> int:
    seen: set[str] = set()
    for match in _TAX_COMPONENT_LABEL_RE.finditer(raw_text):
        token = re.sub(r"\s+", "", match.group(0).lower())
        token = token.removeprefix("add")
        seen.add(token)
    return len(seen)


def _correct_tax_from_totals_when_split_marked(normalized: dict[str, Any], raw_text: str) -> None:
    subtotal = _decimal_or_none(normalized.get("subtotal_amount"))
    total = _decimal_or_none(normalized.get("total_amount"))
    tax = _decimal_or_none(normalized.get("tax_amount"))
    if subtotal is None or total is None:
        return

    expected_tax = total - subtotal
    if expected_tax <= 0:
        return

    marker_count = _count_tax_component_markers(raw_text)
    if marker_count < 2:
        return

    if tax is None:
        normalized["tax_amount"] = expected_tax
        return

    if abs(expected_tax - tax) <= Decimal("0.05"):
        return

    if expected_tax > tax and abs(expected_tax - (tax * 2)) <= Decimal("0.10"):
        normalized["tax_amount"] = expected_tax


def _amounts_with_positions(raw_text: str) -> list[tuple[int, Decimal]]:
    amounts: list[tuple[int, Decimal]] = []
    for match in re.finditer(r"(?<![A-Z0-9])([\d,]+(?:\.\d{1,2})?)(?![A-Z0-9])", raw_text):
        amount = _try_parse_decimal(match.group(1))
        if amount is not None:
            if amount > Decimal("999999999.99"):
                continue
            amounts.append((match.start(), amount))
    return amounts


def _largest_amount_in_text(raw_text: str) -> Decimal | None:
    amounts = [amount for _, amount in _amounts_with_positions(raw_text)]
    return max(amounts) if amounts else None


def _previous_distinct_amount(raw_text: str, target: Decimal) -> Decimal | None:
    amounts = [amount for _, amount in _amounts_with_positions(raw_text)]
    for amount in sorted(set(amounts), reverse=True):
        if amount < target:
            return amount
    return None


def _currency_marked_amounts(raw_text: str) -> list[Decimal]:
    amounts: list[Decimal] = []
    for pattern in (
        r"\u20b9\s*([\d,]+(?:\.\d{1,2})?)",
        r"Rs\.?\s*([\d,]+(?:\.\d{1,2})?)",
        r"INR\s*([\d,]+(?:\.\d{1,2})?)",
    ):
        for match in re.finditer(pattern, raw_text, re.IGNORECASE):
            amount = _try_parse_decimal(match.group(1))
            if amount is not None:
                amounts.append(amount)
    return amounts


def _extract_from_text_with_regex(raw_text: str) -> dict[str, Any]:
    """Run the regex battery against raw PDF text; first match per field wins."""
    normalized: dict[str, Any] = {}
    for field_name, pattern in _PDF_REGEX_PATTERNS:
        if field_name in normalized:
            continue
        match = re.search(pattern, raw_text, re.IGNORECASE | re.MULTILINE)
        if not match:
            continue
        raw_val = match.group(1).strip()
        if not raw_val:
            continue
        if field_name in ("invoice_date", "due_date"):
            parsed = _parse_pdf_date(raw_val)
            if parsed:
                normalized[field_name] = parsed
        elif field_name in ("total_amount", "subtotal_amount", "tax_amount"):
            parsed = _try_parse_decimal(raw_val)
            if parsed is not None:
                normalized[field_name] = parsed
        else:
            normalized[field_name] = raw_val

    if "currency" not in normalized:
        m = re.search(
            r"\b(INR|USD|EUR|GBP|AED|SGD|AUD|CAD|JPY|CNY|CHF|MYR)\b",
            raw_text, re.IGNORECASE,
        )
        if m:
            normalized["currency"] = m.group(1).upper()
        elif chr(8377) in raw_text or re.search(r"\bindian\s+rupees?\b", raw_text, re.IGNORECASE):
            normalized["currency"] = "INR"

    tax_components = _sum_tax_components(raw_text)
    if tax_components is not None:
        normalized["tax_amount"] = tax_components

    total_amount = normalized.get("total_amount")
    subtotal_amount = normalized.get("subtotal_amount")
    tax_amount = normalized.get("tax_amount")

    if total_amount is None:
        inferred_total = _largest_amount_in_text(raw_text)
        if inferred_total is not None:
            normalized["total_amount"] = inferred_total
            total_amount = inferred_total

    if subtotal_amount is None and total_amount is not None and tax_amount is not None:
        normalized["subtotal_amount"] = total_amount - tax_amount
        subtotal_amount = normalized["subtotal_amount"]

    if tax_amount is None and total_amount is not None and subtotal_amount is not None:
        normalized["tax_amount"] = total_amount - subtotal_amount
        tax_amount = normalized["tax_amount"]

    if subtotal_amount is None and total_amount is not None and tax_amount is not None:
        previous_amount = _previous_distinct_amount(raw_text, total_amount)
        if previous_amount is not None and previous_amount + tax_amount == total_amount:
            normalized["subtotal_amount"] = previous_amount
            subtotal_amount = previous_amount

    if total_amount is not None and subtotal_amount is None and tax_amount is None:
        previous_amount = _previous_distinct_amount(raw_text, total_amount)
        if previous_amount is not None and previous_amount < total_amount:
            inferred_tax = total_amount - previous_amount
            if inferred_tax > 0:
                normalized["subtotal_amount"] = previous_amount
                normalized["tax_amount"] = inferred_tax
                subtotal_amount = previous_amount
                tax_amount = inferred_tax

    if total_amount is None and subtotal_amount is not None and tax_amount is not None:
        normalized["total_amount"] = subtotal_amount + tax_amount

    currency_amounts = _currency_marked_amounts(raw_text)
    if currency_amounts:
        distinct_currency_amounts = sorted(set(currency_amounts), reverse=True)
        currency_total = distinct_currency_amounts[0]

        if total_amount is None or total_amount not in distinct_currency_amounts:
            normalized["total_amount"] = currency_total
            total_amount = currency_total

        subtotal_candidate = next(
            (amount for amount in distinct_currency_amounts if amount < total_amount),
            None,
        )
        if subtotal_candidate is not None:
            inferred_tax = total_amount - subtotal_candidate
            if inferred_tax > 0:
                if subtotal_amount is None or subtotal_amount not in distinct_currency_amounts or subtotal_amount >= total_amount:
                    normalized["subtotal_amount"] = subtotal_candidate
                    subtotal_amount = subtotal_candidate
                if (
                    tax_amount is None
                    or tax_amount > inferred_tax
                    or inferred_tax == tax_amount * 2
                ):
                    normalized["tax_amount"] = inferred_tax
                    tax_amount = inferred_tax

    _correct_tax_from_totals_when_split_marked(normalized, raw_text)

    invoice_ref = normalized.get("vendor_invoice_number")
    if invoice_ref and not re.search(r"\d", str(invoice_ref)):
        normalized.pop("vendor_invoice_number", None)

    po_ref = normalized.get("po_number")
    if po_ref and not re.search(r"\d", str(po_ref)):
        normalized.pop("po_number", None)

    bill_to = normalized.get("bill_to_name")
    if bill_to:
        lines = [ln.strip() for ln in str(bill_to).splitlines() if ln.strip()]
        lines = [ln for ln in lines if ln.lower() not in {"bill to", "ship to", "client"}]
        if lines:
            normalized["bill_to_name"] = lines[0]

    return normalized


# ---------------------------------------------------------------------------
# Vendor layout hook registry (extensibility point, no built-in hooks yet)
# ---------------------------------------------------------------------------
# Maps a lowercase vendor-name substring → hook(normalized, raw_text) -> dict.
# Register vendor-specific rules here or in a separate vendor_hooks.py module.

_VENDOR_HOOK_REGISTRY: dict[str, Any] = {}


def _apply_vendor_hooks(
    normalized: dict[str, Any],
    raw_text: str,
    vendor_name: str | None,
) -> dict[str, Any]:
    if not vendor_name:
        return normalized
    vn_lower = vendor_name.lower()
    for pattern, hook_fn in _VENDOR_HOOK_REGISTRY.items():
        if pattern in vn_lower:
            try:
                result = hook_fn(normalized, raw_text)
                if result:
                    normalized = result
            except Exception:
                logger.warning("Vendor hook '%s' raised an error", pattern, exc_info=True)
    return normalized


# ---------------------------------------------------------------------------
# Confidence scoring
# ---------------------------------------------------------------------------

_CONFIDENCE_REQUIRED = ["vendor_invoice_number", "invoice_date", "total_amount", "currency"]
_CONFIDENCE_OPTIONAL = ["due_date", "po_number", "subtotal_amount", "tax_amount", "description"]


def _score_confidence(normalized: dict[str, Any]) -> float:
    """
    Weighted confidence: required fields 75%, optional fields 25%.
    Returns a float in [0.0, 1.0].
    """
    req = sum(1 for f in _CONFIDENCE_REQUIRED if normalized.get(f)) / len(_CONFIDENCE_REQUIRED)
    opt = sum(1 for f in _CONFIDENCE_OPTIONAL if normalized.get(f)) / len(_CONFIDENCE_OPTIONAL)
    return round(0.75 * req + 0.25 * opt, 4)


def _looks_like_invoice_number(value: str) -> bool:
    text = value.strip()
    if not text:
        return False
    return bool(re.fullmatch(r"[A-Z0-9][A-Z0-9\-\/]{3,}", text, re.IGNORECASE))


def _looks_like_currency(value: str) -> bool:
    return bool(re.fullmatch(r"[A-Z]{3}", value.strip().upper()))


def _extract_pdf_template_values(raw_text: str) -> dict[str, Any]:
    """
    Extract values from the generated fillable invoice PDF template.

    PyPDF text extraction for this template appends the filled values near the
    end of the document in field order, after the instructional text. We detect
    that value block and map it back to invoice fields.
    """
    lines = [line.strip() for line in raw_text.splitlines() if line.strip()]
    if not lines:
        return {}

    start_idx = None
    for idx in range(len(lines) - 4):
        if (
            _looks_like_invoice_number(lines[idx])
            and _try_parse_date(lines[idx + 1]) is not None
            and _try_parse_date(lines[idx + 2]) is not None
            and _looks_like_currency(lines[idx + 3])
        ):
            start_idx = idx

    if start_idx is None:
        return {}

    value_lines = [
        line for line in lines[start_idx:]
        if not _is_pdf_placeholder_line(line)
    ]
    if len(value_lines) < 8:
        return {}

    normalized: dict[str, Any] = {}
    normalized["vendor_invoice_number"] = value_lines[0]
    normalized["invoice_date"] = _try_parse_date(value_lines[1])
    normalized["due_date"] = _try_parse_date(value_lines[2])
    normalized["currency"] = value_lines[3].upper()

    # Description may span multiple lines until the first numeric amount.
    cursor = 4
    description_lines: list[str] = []
    while cursor < len(value_lines) and _try_parse_decimal(value_lines[cursor]) is None:
        description_lines.append(value_lines[cursor])
        cursor += 1
    if description_lines:
        normalized["description"] = " ".join(description_lines).strip()

    numeric_fields = ("subtotal_amount", "tax_amount", "total_amount")
    for field in numeric_fields:
        if cursor >= len(value_lines):
            return normalized
        parsed = _try_parse_decimal(value_lines[cursor])
        if parsed is None:
            return normalized
        normalized[field] = parsed
        cursor += 1

    if cursor < len(value_lines):
        po_candidate = value_lines[cursor].strip()
        if po_candidate and not _is_pdf_placeholder_line(po_candidate):
            normalized["po_number"] = po_candidate

    return {k: v for k, v in normalized.items() if v not in (None, "")}


def extract_excel(file_obj) -> ExtractionResult:
    """
    Extract invoice fields from an Excel workbook (.xlsx/.xls).

    Scans all sheets for labelled rows.  Two layout patterns are supported:
      - key-value:  column-A = label, column-B = value
      - tabular:    first row = headers, subsequent rows = values
    """
    raw_cells: dict[str, Any] = {}
    normalized: dict[str, Any] = {}
    warnings: list[str] = []
    errors: list[str] = []
    confidence = 0.0
    found_fields = 0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as exc:
        errors.append(f"Could not open workbook: {exc}")
        return ExtractionResult(raw_cells={}, normalized={}, confidence=0.0, warnings=warnings, errors=errors)

    # Scan every sheet
    for sheet in wb.worksheets:
        if sheet.max_row < 1 or sheet.max_column < 1:
            continue

        # ── Pattern 1: key-value layout (label in col A) ───────────────────────
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 200), values_only=True):
            if not row:
                continue
            label_cell, value_cell = row[0], (row[1] if len(row) > 1 else None)
            if not label_cell:
                continue
            label = str(label_cell).strip()
            field_name = _guess_field(label)
            if field_name and value_cell is not None:
                # Skip if value_cell also looks like a field label
                # (prevents header-row mis-match: "Invoice Number" col B = "Invoice Date")
                if _guess_field(str(value_cell).strip()):
                    continue
                raw_cells[label] = value_cell
                # Parse value
                parsed = None
                if field_name in ("invoice_date", "due_date"):
                    parsed = _try_parse_date(value_cell)
                elif field_name in ("subtotal_amount", "tax_amount", "total_amount"):
                    parsed = _try_parse_decimal(value_cell)
                else:
                    parsed = str(value_cell).strip()
                if parsed:
                    normalized[field_name] = parsed
                    found_fields += 1

        # ── Pattern 2: tabular header row ──────────────────────────────────────
        header_row = None
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 50), values_only=True), start=1):
            if not row or not row[0]:
                continue
            # Detect header: first cell looks like a field label
            if _guess_field(str(row[0])):
                header_row = [str(c).strip() if c else "" for c in row]
                break

        if header_row:
            # Scan values under header
            for row in sheet.iter_rows(min_row=row_idx + 1, max_row=min(sheet.max_row, row_idx + 100), values_only=True):
                if not row or not row[0]:
                    continue
                for col_idx, cell in enumerate(row):
                    if col_idx >= len(header_row):
                        break
                    header = header_row[col_idx]
                    field_name = _guess_field(header)
                    if field_name and cell is not None:
                        raw_cells[f"{header}:{row_idx}" if field_name == "total_amount" else header] = cell
                        parsed = None
                        if field_name in ("invoice_date", "due_date"):
                            parsed = _try_parse_date(cell)
                        elif field_name in ("subtotal_amount", "tax_amount", "total_amount"):
                            parsed = _try_parse_decimal(cell)
                        else:
                            parsed = str(cell).strip()
                        if parsed and field_name not in normalized:
                            normalized[field_name] = parsed
                            found_fields += 1

    # Compute confidence
    required_fields = ["vendor_invoice_number", "invoice_date", "total_amount", "currency"]
    matched = sum(1 for f in required_fields if f in normalized)
    confidence = matched / len(required_fields)

    if found_fields == 0:
        warnings.append("No invoice fields recognised in the uploaded file.")

    return ExtractionResult(
        raw_cells=raw_cells,
        normalized=normalized,
        confidence=min(1.0, confidence),
        warnings=warnings,
        errors=errors,
    )


# ---------------------------------------------------------------------------
# Submission services
# ---------------------------------------------------------------------------

class SubmissionStateError(ValueError):
    """Submission is in the wrong state for the requested operation."""


class DuplicateInvoiceError(ValueError):
    """A submission for the same vendor + invoice number already exists."""


class VendorRouteError(ValueError):
    """
    The chosen VendorSubmissionRoute is invalid or misconfigured.

    Raised (and transaction rolled back) when:
    - The route is inactive
    - The mapped WorkflowTemplate has no published version
    - The mapped WorkflowTemplate is not active
    - Workflow activation fails due to unresolved assignees
    """


def _file_hash(file_obj) -> str:
    """SHA-256 hash of a file object."""
    hasher = hashlib.sha256()
    for chunk in file_obj.chunks():
        hasher.update(chunk)
    # Reset file pointer for subsequent use
    file_obj.seek(0)
    return hasher.hexdigest()


def create_vendor_invoice_submission(
    *, user, vendor, scope_node, file_obj,
    normalized_data: dict | None = None,
) -> VendorInvoiceSubmission:
    """
    Create a new vendor invoice submission record from an uploaded file.

    Validates:
    - user has an active UserVendorAssignment to this vendor
    - vendor is active
    - scope_node is within vendor's scope_node hierarchy
    """
    from apps.vendors.models import OperationalStatus, UserVendorAssignment

    # Resolve vendor user assignment
    assignment = (
        UserVendorAssignment.objects
        .filter(user=user, is_active=True, vendor=vendor)
        .select_related("vendor__scope_node")
        .first()
    )
    if not assignment:
        raise SubmissionStateError("You are not linked to this vendor.")

    if vendor.operational_status != OperationalStatus.ACTIVE:
        raise SubmissionStateError("Your vendor account is not active.")

    # Scope node bounds check
    vendor_sn = vendor.scope_node
    if not (scope_node.path == vendor_sn.path or scope_node.path.startswith(vendor_sn.path + "/")):
        raise SubmissionStateError("This bill-to entity is outside your vendor scope.")

    # File info
    file_name = file_obj.name
    ext = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""
    file_type = "pdf" if ext == "pdf" else ("xlsx" if ext in ("xlsx", "xls") else ext)

    # Hash for duplicate detection
    file_obj.seek(0)
    file_hash = _file_hash(file_obj)
    file_obj.seek(0)

    submission = VendorInvoiceSubmission.objects.create(
        vendor=vendor,
        submitted_by=user,
        scope_node=scope_node,
        source_file=file_obj,
        source_file_name=file_name,
        source_file_type=file_type,
        source_file_hash=file_hash,
        normalized_data=normalized_data or {},
        status=VendorInvoiceSubmissionStatus.NEEDS_CORRECTION
        if normalized_data
        else VendorInvoiceSubmissionStatus.UPLOADED,
    )
    return submission


def extract_invoice_submission(submission: VendorInvoiceSubmission) -> ExtractionResult:
    """
    Run extraction on the submission's source file and update the submission record.

    Returns an ExtractionResult; the submission's normalized_data and status are
    also updated atomically.
    """
    if submission.status not in (VendorInvoiceSubmissionStatus.UPLOADED, VendorInvoiceSubmissionStatus.NEEDS_CORRECTION):
        raise SubmissionStateError(
            f"Cannot extract — submission is in status '{submission.status}'."
        )

    submission.status = VendorInvoiceSubmissionStatus.EXTRACTING
    submission.save(update_fields=["status"])

    try:
        if submission.source_file_type == "pdf":
            result = extract_pdf(submission)
        else:
            file_obj = submission.source_file.open("rb")
            result = extract_excel(file_obj)
            file_obj.close()

        # Persist extraction results — convert Decimals to floats for JSON serialization
        submission.raw_extracted_data = _json_safe_dict(result.raw_cells)
        submission.original_normalized_data = _json_safe_dict(result.normalized)
        submission.normalized_data = _json_safe_dict(result.normalized)
        submission.confidence_score = Decimal(str(result.confidence)) if result.confidence else None
        if result.errors:
            submission.validation_errors = result.errors
            submission.status = VendorInvoiceSubmissionStatus.NEEDS_CORRECTION
        elif result.confidence < 0.5:
            submission.status = VendorInvoiceSubmissionStatus.NEEDS_CORRECTION
        else:
            submission.status = VendorInvoiceSubmissionStatus.READY
        submission.save(update_fields=[
            "raw_extracted_data", "original_normalized_data", "normalized_data", "confidence_score",
            "validation_errors", "status", "updated_at",
        ])
        return result

    except Exception as exc:
        logger.exception("Extraction failed for submission %s", submission.pk)
        submission.status = VendorInvoiceSubmissionStatus.NEEDS_CORRECTION
        submission.validation_errors = [{"extraction": str(exc)}]
        submission.save(update_fields=["status", "validation_errors"])
        return ExtractionResult(
            raw_cells={}, normalized={}, confidence=0.0,
            errors=[f"Extraction failed: {exc}"],
        )


def _extract_pdf_locally(submission: VendorInvoiceSubmission) -> ExtractionResult:
    warnings: list[str] = []
    errors: list[str] = []
    raw_text = ""
    ocr_used = False

    # Layer 1: native text extraction
    try:
        import PyPDF2
        file_obj = submission.source_file.open("rb")
        reader = PyPDF2.PdfReader(file_obj)
        raw_text = "\n".join(page.extract_text() or "" for page in reader.pages)
        file_obj.close()
    except ImportError:
        warnings.append("PDF text extraction library (PyPDF2) is not installed.")
    except Exception as exc:
        warnings.append(f"PDF text extraction failed: {exc}")

    # Layer 2: classification
    pdf_class = _classify_pdf_text(raw_text)

    # Layer 3: OCR when text is sparse
    if pdf_class == "scanned":
        file_obj = submission.source_file.open("rb")
        ocr_text = _ocr_pdf_pages(file_obj)
        file_obj.close()
        if ocr_text.strip():
            raw_text = ocr_text
            ocr_used = True
            pdf_class = _classify_pdf_text(raw_text)
        else:
            warnings.append(
                "This appears to be a scanned PDF. OCR is unavailable or produced no text. "
                "Please fill in the invoice details manually."
            )

    # Layer 4: field extraction
    normalized: dict[str, Any] = {}
    if pdf_class == "template":
        normalized = _extract_pdf_template_values(raw_text)
    if not normalized:
        normalized = _extract_from_text_with_regex(raw_text)

    # Layer 5: vendor-specific rules
    vendor_name = getattr(submission.vendor, "vendor_name", None)
    normalized = _apply_vendor_hooks(normalized, raw_text, vendor_name)

    # Layer 6: confidence
    confidence = _score_confidence(normalized)

    raw_cells: dict[str, Any] = {"raw_text": raw_text}
    if ocr_used:
        raw_cells["extraction_method"] = "ocr"
        warnings.append("OCR was used to read this PDF. Please review all extracted fields carefully.")
    elif pdf_class == "template":
        raw_cells["extraction_method"] = "template"
    else:
        raw_cells["extraction_method"] = "regex"

    if not normalized:
        warnings.append("No invoice fields could be extracted. Please enter details manually.")
        errors.append("No fields could be extracted from the PDF.")
    elif confidence < 0.5:
        warnings.append("Extraction found only partial invoice data. Please review and complete all fields.")

    return ExtractionResult(
        raw_cells=raw_cells,
        normalized={k: v for k, v in normalized.items() if v not in (None, "")},
        confidence=confidence,
        warnings=warnings,
        errors=errors,
    )


def extract_pdf(submission: VendorInvoiceSubmission) -> ExtractionResult:
    """
    Hybrid PDF extraction pipeline:

    Layer 1 — Azure Document Intelligence (primary) when configured.
    Layer 2 — Local parser fallback / backfill.
    Layer 3 — Review-before-submit remains mandatory.
    """
    local_result: ExtractionResult | None = None

    if _azure_docintel_configured():
        try:
            azure_result = _extract_pdf_with_azure_document_intelligence(submission)
            if azure_result and azure_result.normalized:
                local_result = _extract_pdf_locally(submission)
                return _merge_extraction_results(azure_result, local_result, submission=submission)
        except Exception as exc:
            logger.warning(
                "Azure Document Intelligence extraction failed for submission %s; falling back to local parser: %s",
                submission.pk,
                exc,
            )

    result = _extract_pdf_locally(submission)
    result.normalized = _normalize_extracted_invoice_fields(result.normalized, submission=submission)
    result.confidence = _score_confidence(result.normalized)
    return result


def validate_invoice_submission(submission: VendorInvoiceSubmission) -> list[dict]:
    """
    Run validation rules on normalized_data and return a list of error dicts.
    An empty list means validation passed.
    """
    errors = []
    nd = submission.normalized_data or {}
    vendor = submission.vendor

    # Invoice number
    if not nd.get("vendor_invoice_number"):
        errors.append({"field": "vendor_invoice_number", "message": "Invoice number is required."})
    else:
        # Duplicate check
        existing = VendorInvoiceSubmission.objects.filter(
            vendor=vendor,
            normalized_data__vendor_invoice_number=nd["vendor_invoice_number"],
        ).exclude(pk=submission.pk).exists()
        if existing:
            errors.append({"field": "vendor_invoice_number", "message": "An invoice with this number already exists."})

    # Invoice date
    if not nd.get("invoice_date"):
        errors.append({"field": "invoice_date", "message": "Invoice date is required."})

    # Total amount
    total = nd.get("total_amount")
    if not total:
        errors.append({"field": "total_amount", "message": "Total amount is required."})
    elif isinstance(total, (int, float, Decimal)) and total <= 0:
        errors.append({"field": "total_amount", "message": "Total amount must be greater than zero."})

    # Currency
    if not nd.get("currency"):
        errors.append({"field": "currency", "message": "Currency is required."})

    return errors


def update_invoice_submission_fields(submission: VendorInvoiceSubmission, normalized_data: dict) -> VendorInvoiceSubmission:
    """Update normalized_data and re-validate. Returns updated submission."""
    submission.normalized_data = normalized_data
    submission.validation_errors = validate_invoice_submission(submission)
    submission.status = (
        VendorInvoiceSubmissionStatus.NEEDS_CORRECTION
        if submission.validation_errors
        else VendorInvoiceSubmissionStatus.READY
    )
    submission.save(update_fields=["normalized_data", "validation_errors", "status", "updated_at"])
    return submission


@transaction.atomic
def submit_vendor_invoice_submission(submission: VendorInvoiceSubmission, user) -> Invoice:
    """
    Finalise a vendor invoice submission:
      1. Validate all required fields
      2. Create the final Invoice
      3. Attach documents
      4. Start workflow
      5. Update submission status
    """

    if submission.status not in (VendorInvoiceSubmissionStatus.UPLOADED, VendorInvoiceSubmissionStatus.NEEDS_CORRECTION, VendorInvoiceSubmissionStatus.READY):
        raise SubmissionStateError(
            f"Cannot submit — submission is in status '{submission.status}'."
        )

    # Run full validation
    validation_errors = validate_invoice_submission(submission)
    if validation_errors:
        submission.validation_errors = validation_errors
        submission.status = VendorInvoiceSubmissionStatus.NEEDS_CORRECTION
        submission.save(update_fields=["validation_errors", "status"])
        raise SubmissionStateError("Validation failed. Please correct the highlighted fields.")

    nd = submission.normalized_data

    # Build amount (prefer total_amount; compute from subtotal + tax if missing)
    total = nd.get("total_amount")
    if not total:
        subtotal = float(nd.get("subtotal_amount") or 0)
        tax = float(nd.get("tax_amount") or 0)
        total = Decimal(str(subtotal + tax))
    else:
        total = Decimal(str(total))

    currency = nd.get("currency", "INR")
    invoice_date_str = nd.get("invoice_date")
    due_date_str = nd.get("due_date")

    from django.utils.dateparse import parse_date
    invoice_date = parse_date(invoice_date_str) if invoice_date_str else None
    due_date = parse_date(due_date_str) if due_date_str else None

    # Create final Invoice — stops at pending_workflow; no auto workflow start.
    invoice = Invoice.objects.create(
        scope_node=submission.scope_node,
        title=nd.get("vendor_invoice_number", "Untitled Invoice"),
        amount=total,
        currency=currency,
        vendor=submission.vendor,
        created_by=user,
        po_number=nd.get("po_number", ""),
        vendor_invoice_number=nd.get("vendor_invoice_number", ""),
        invoice_date=invoice_date,
        due_date=due_date,
        subtotal_amount=Decimal(str(nd.get("subtotal_amount", 0))),
        tax_amount=Decimal(str(nd.get("tax_amount", 0))),
        description=nd.get("description", ""),
        status=InvoiceStatus.PENDING_WORKFLOW,
    )

    # Link documents to invoice
    for doc in submission.documents.all():
        doc.invoice = invoice
        doc.save(update_fields=["invoice"])

    # Update submission
    submission.final_invoice = invoice
    submission.status = VendorInvoiceSubmissionStatus.SUBMITTED
    submission.submitted_at = timezone.now()
    submission.correction_note = ""
    submission.correction_requested_by = None
    submission.correction_requested_at = None
    submission.validation_errors = []
    submission.save(update_fields=[
        "final_invoice", "status", "submitted_at",
        "correction_note", "correction_requested_by", "correction_requested_at",
        "validation_errors",
    ])

    return invoice


def submit_vendor_invoice_with_route(
    submission: VendorInvoiceSubmission,
    user,
    send_to_route,
) -> tuple[Invoice, list[dict]]:
    """
    Primary vendor invoice submission path that auto-routes to a workflow.

    Returns (invoice, warnings_list). Validation failures persist submission
    errors/status. Invoice/workflow creation stays atomic so no orphan records
    are left behind.
    """
    if submission.status not in (
        VendorInvoiceSubmissionStatus.UPLOADED,
        VendorInvoiceSubmissionStatus.NEEDS_CORRECTION,
        VendorInvoiceSubmissionStatus.READY,
    ):
        raise SubmissionStateError(
            f"Cannot submit ??? submission is in status '{submission.status}'."
        )

    if submission.vendor_id:
        from apps.vendors.services import assert_vendor_profile_not_on_hold
        from apps.vendors.models import Vendor as _Vendor
        try:
            _vendor = _Vendor.objects.only(
                "profile_change_pending", "profile_hold_reason"
            ).get(pk=submission.vendor_id)
            assert_vendor_profile_not_on_hold(_vendor)
        except _Vendor.DoesNotExist:
            pass

    validation_result = validate_vendor_submission_for_submit(submission, send_to_route)

    if not validation_result.is_valid:
        submission.validation_errors = [
            {"field": k, "message": v[0]} for k, v in validation_result.field_errors.items()
        ]
        submission.status = VendorInvoiceSubmissionStatus.NEEDS_CORRECTION
        submission.save(update_fields=["validation_errors", "status"])
        raise SubmissionValidationError(validation_result)

    with transaction.atomic():
        nd = submission.normalized_data

        total = nd.get("total_amount")
        if not total:
            subtotal = float(nd.get("subtotal_amount") or 0)
            tax = float(nd.get("tax_amount") or 0)
            total = Decimal(str(subtotal + tax))
        else:
            total = Decimal(str(total))

        from django.utils.dateparse import parse_date
        invoice_date = parse_date(nd["invoice_date"]) if nd.get("invoice_date") else None
        due_date = parse_date(nd["due_date"]) if nd.get("due_date") else None

        from apps.workflow.models import WorkflowTemplateVersion, VersionStatus
        published_version = (
            WorkflowTemplateVersion.objects
            .select_related("template")
            .filter(
                template=send_to_route.workflow_template,
                status=VersionStatus.PUBLISHED,
            )
            .order_by("-version_number")
            .first()
        )

        invoice = Invoice.objects.create(
            scope_node=submission.scope_node,
            title=nd.get("vendor_invoice_number", "Untitled Invoice"),
            amount=total,
            currency=nd.get("currency", "INR"),
            vendor=submission.vendor,
            created_by=user,
            po_number=nd.get("po_number", ""),
            vendor_invoice_number=nd.get("vendor_invoice_number", ""),
            invoice_date=invoice_date,
            due_date=due_date,
            subtotal_amount=Decimal(str(nd.get("subtotal_amount", 0))),
            tax_amount=Decimal(str(nd.get("tax_amount", 0))),
            description=nd.get("description", ""),
            status=InvoiceStatus.PENDING_WORKFLOW,
            selected_workflow_template=send_to_route.workflow_template,
            selected_workflow_version=published_version,
            workflow_selected_by=user,
            workflow_selected_at=timezone.now(),
        )

        for doc in submission.documents.all():
            doc.invoice = invoice
            doc.save(update_fields=["invoice"])

        from apps.workflow.services import create_workflow_instance_draft, activate_workflow_instance
        instance = create_workflow_instance_draft(
            template_version=published_version,
            subject_type="invoice",
            subject_id=invoice.pk,
            subject_scope_node=invoice.scope_node,
            started_by=user,
        )

        try:
            activate_workflow_instance(instance, activated_by=user)
        except ValueError as exc:
            raise VendorRouteError(
                f"Route '{send_to_route.label}' cannot start: {exc} "
                "Assign all required approvers before this route can be used."
            )

        submission.send_to_route = send_to_route
        submission.final_invoice = invoice
        submission.status = VendorInvoiceSubmissionStatus.SUBMITTED
        submission.submitted_at = timezone.now()
        submission.correction_note = ""
        submission.correction_requested_by = None
        submission.correction_requested_at = None
        submission.validation_errors = []
        submission.save(update_fields=[
            "send_to_route", "final_invoice", "status", "submitted_at",
            "correction_note", "correction_requested_by", "correction_requested_at",
            "validation_errors",
        ])

        invoice.refresh_from_db()
        return invoice, validation_result.warnings


def cancel_vendor_invoice_submission(submission: VendorInvoiceSubmission) -> None:
    """
    Cancel a vendor invoice submission.
    Allowed from: uploaded, extracting (blocked), needs_correction, ready.
    Not allowed from: submitted, cancelled.
    Not allowed if a final invoice already exists.
    """
    if submission.final_invoice_id is not None:
        raise SubmissionStateError(
            "Cannot cancel — a final invoice has already been created from this submission."
        )

    # extracting is a transitional state — block cancellation while actively extracting
    if submission.status == VendorInvoiceSubmissionStatus.EXTRACTING:
        raise SubmissionStateError(
            "Cannot cancel — extraction is in progress. Please wait for it to complete."
        )

    cancellable = (
        VendorInvoiceSubmissionStatus.UPLOADED,
        VendorInvoiceSubmissionStatus.NEEDS_CORRECTION,
        VendorInvoiceSubmissionStatus.READY,
    )
    if submission.status == VendorInvoiceSubmissionStatus.SUBMITTED:
        raise SubmissionStateError("Cannot cancel — submission has already been submitted.")
    if submission.status == VendorInvoiceSubmissionStatus.CANCELLED:
        raise SubmissionStateError("Submission is already cancelled.")
    if submission.status not in cancellable:
        raise SubmissionStateError(
            f"Cannot cancel — submission is in status '{submission.status}'."
        )

    submission.status = VendorInvoiceSubmissionStatus.CANCELLED
    submission.save(update_fields=["status", "updated_at"])


def _delete_file_field_safely(file_field) -> None:
    """Delete a FileField from storage, swallowing storage-layer failures."""
    if not file_field:
        return
    name = getattr(file_field, "name", "")
    if not name:
        return
    try:
        file_field.storage.delete(name)
    except Exception:
        pass


def discard_vendor_invoice_submission(submission: VendorInvoiceSubmission) -> None:
    """
    Permanently delete an unfinished vendor invoice submission.

    Allowed only for submission-side records that have not created a final Invoice:
      - uploaded
      - needs_correction
      - ready
      - cancelled

    Not allowed for:
      - extracting
      - submitted
      - rejected
      - any submission already linked to a final invoice
    """
    if submission.final_invoice_id is not None:
        raise SubmissionStateError(
            "Cannot discard this record because a final invoice has already been created from it."
        )

    if submission.status == VendorInvoiceSubmissionStatus.EXTRACTING:
        raise SubmissionStateError(
            "Cannot discard while extraction is in progress. Please wait for it to complete."
        )

    discardable = {
        VendorInvoiceSubmissionStatus.UPLOADED,
        VendorInvoiceSubmissionStatus.NEEDS_CORRECTION,
        VendorInvoiceSubmissionStatus.READY,
        VendorInvoiceSubmissionStatus.CANCELLED,
    }
    if submission.status == VendorInvoiceSubmissionStatus.SUBMITTED:
        raise SubmissionStateError("Cannot discard a submission that has already been submitted.")
    if submission.status == VendorInvoiceSubmissionStatus.REJECTED:
        raise SubmissionStateError("Cannot discard a submission that is already part of review history.")
    if submission.status not in discardable:
        raise SubmissionStateError(
            f"Cannot discard submission in status '{submission.status}'."
        )

    source_file = submission.source_file
    document_files = [doc.file for doc in submission.documents.all()]

    submission.delete()

    _delete_file_field_safely(source_file)
    for document_file in document_files:
        _delete_file_field_safely(document_file)


# ---------------------------------------------------------------------------
# Invoice template downloads
# ---------------------------------------------------------------------------

def generate_excel_template() -> HttpResponse:
    """
    Generate a Vendor Invoice Upload Template .xlsx with labelled header row.
    Labels match extract_excel() field detection so uploaded files work directly.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse("openpyxl not installed", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Data"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "Invoice Number", "Invoice Date", "Due Date", "PO Number",
        "Currency", "Subtotal Amount", "Tax Amount", "Total Amount",
        "Description",
    ]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    col_widths = [22, 16, 16, 18, 10, 18, 14, 16, 40]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    ws.append([])  # blank row hint
    ws.append(["INV-001", "2026-04-18", "2026-05-18", "PO-001", "INR", "10000", "1800", "11800", "Services rendered"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=Vendor_Invoice_Template.xlsx"
    return response


def generate_pdf_template() -> HttpResponse:
    """
    Generate a styled Vendor Invoice Upload Template PDF.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.colors import HexColor, white, black
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
        )
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
    except ImportError:
        return _pdf_template_fallback()

    # ── Colours ────────────────────────────────────────────────────────────────
    NAVY      = HexColor("#1E3A5F")
    LIGHT_BLUE = HexColor("#E8F0F8")
    MID_BLUE  = HexColor("#4A7BAD")
    GREEN     = HexColor("#2E7D32")
    RED       = HexColor("#C62828")
    GREY_BG   = HexColor("#F5F5F5")
    BORDER    = HexColor("#B0BEC5")

    # ── Document ───────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm,
        topMargin=18*mm, bottomMargin=18*mm,
    )

    # ── Styles ─────────────────────────────────────────────────────────────────
    def ps(name, **kw):
        return ParagraphStyle(name, **kw)

    title_s = ps("title", fontSize=18, leading=22, textColor=white, alignment=TA_CENTER, fontName="Helvetica-Bold")
    subtitle_s = ps("subtitle", fontSize=9, leading=12, textColor=LIGHT_BLUE, alignment=TA_CENTER, fontName="Helvetica")
    section_s = ps("section", fontSize=8, leading=10, textColor=MID_BLUE, fontName="Helvetica-Bold")
    label_s = ps("label", fontSize=8.5, leading=11, textColor=NAVY, fontName="Helvetica-Bold")
    hint_s  = ps("hint",  fontSize=7.5, leading=10, textColor=HexColor("#607D8B"), fontName="Helvetica-Oblique")
    body_s  = ps("body",  fontSize=8,   leading=11, textColor=black,  fontName="Helvetica")
    note_s  = ps("note",  fontSize=7.5, leading=10, textColor=HexColor("#455A64"), fontName="Helvetica")
    footer_s = ps("footer", fontSize=6.5, leading=9, textColor=HexColor("#90A4AE"), alignment=TA_CENTER)

    # ── Build elements ─────────────────────────────────────────────────────────
    els = []

    # ── Header banner ──────────────────────────────────────────────────────────
    header_data = [[
        Paragraph("Horizon — Vendor Invoice Submission Template (Recommended)", title_s),
    ]]
    header_tbl = Table(header_data, colWidths=[174*mm])
    header_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), NAVY),
        ("TOPPADDING",   (0, 0), (-1, -1), 8*mm),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 8*mm),
        ("LEFTPADDING",  (0, 0), (-1, -1), 4*mm),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4*mm),
        ("ROUNDEDCORNERS", [4]),
    ]))
    els.append(header_tbl)
    els.append(Spacer(1, 3*mm))

    # ── Recommended note ───────────────────────────────────────────────────────
    note_data = [[
        Paragraph(
            "&#10004; <b>Template Usage:</b> You can upload invoices in any format. "
            "Using this template improves automatic field extraction accuracy, reducing manual corrections. "
            "All extracted fields will be reviewed and confirmed before final submission.",
            note_s,
        )
    ]]
    note_tbl = Table(note_data, colWidths=[174*mm])
    note_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), HexColor("#E8F5E9")),
        ("TOPPADDING",   (0, 0), (-1, -1), 3*mm),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3*mm),
        ("LEFTPADDING",  (0, 0), (-1, -1), 4*mm),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4*mm),
        ("BOX",          (0, 0), (-1, -1), 0.5, GREEN),
    ]))
    els.append(note_tbl)
    els.append(Spacer(1, 5*mm))

    # ── Section header helper ──────────────────────────────────────────────────
    def section_row(label):
        return [Paragraph(label.upper(), section_s)]

    # ── Section: Invoice Details ───────────────────────────────────────────────
    els.append(Paragraph("Invoice Details", section_s))
    els.append(Spacer(1, 2*mm))

    field_label = lambda t: Paragraph(t, label_s)
    field_hint  = lambda t: Paragraph(t, hint_s)

    # We'll build a 2-col table: [Field Name | Fillable space]
    def row(label, hint="", required=False):
        req_mark = " <font color='#C62828'>*</font>" if required else ""
        return [field_label(label + req_mark), field_hint(hint) if hint else Paragraph("", hint_s)]

    DETAIL_FIELDS = [
        ("Invoice Number", "e.g. INV-2026-001", True),
        ("Invoice Date", "YYYY-MM-DD  (e.g. 2026-04-18)", True),
        ("Due Date", "YYYY-MM-DD", False),
        ("Currency", "e.g. INR, USD, EUR — use 3-letter ISO code", False),
        ("Description", "Brief description of goods or services", False),
    ]

    detail_rows = [row(l, h, r) for l, h, r in DETAIL_FIELDS]

    detail_tbl = Table(detail_rows, colWidths=[45*mm, 129*mm])
    detail_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (0, -1), LIGHT_BLUE),
        ("BACKGROUND",   (1, 0), (1, -1), white),
        ("BOX",          (0, 0), (-1, -1), 0.5, BORDER),
        ("INNERGRID",    (0, 0), (-1, -1), 0.25, BORDER),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",   (0, 0), (-1, -1), 3*mm),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3*mm),
        ("LEFTPADDING",  (0, 0), (-1, -1), 3*mm),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3*mm),
    ]))
    els.append(detail_tbl)
    els.append(Spacer(1, 5*mm))

    # ── Section: Amounts ───────────────────────────────────────────────────────
    els.append(Paragraph("Amounts", section_s))
    els.append(Spacer(1, 2*mm))

    AMOUNT_FIELDS = [
        ("Subtotal Amount", "Numeric value without currency symbol (e.g. 10000)", True),
        ("Tax Amount", "Tax portion (e.g. 1800). Total = Subtotal + Tax.", False),
        ("Total Amount", "Grand total = Subtotal + Tax. Auto-calculated if using Excel.", False),
        ("PO Number", "Optional purchase order reference, if available.", False),
    ]

    amount_rows = [row(l, h, r) for l, h, r in AMOUNT_FIELDS]

    amount_tbl = Table(amount_rows, colWidths=[45*mm, 129*mm])
    amount_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (0, -1), LIGHT_BLUE),
        ("BACKGROUND",   (1, 0), (1, -1), white),
        ("BOX",          (0, 0), (-1, -1), 0.5, BORDER),
        ("INNERGRID",    (0, 0), (-1, -1), 0.25, BORDER),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",   (0, 0), (-1, -1), 3*mm),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3*mm),
        ("LEFTPADDING",  (0, 0), (-1, -1), 3*mm),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3*mm),
    ]))
    els.append(amount_tbl)
    els.append(Spacer(1, 5*mm))

    # ── PO reference notice ─────────────────────────────────────────────────────
    po_notice_data = [[
        Paragraph(
            "<b>Note:</b> The <b>PO Number</b> field is optional. Add it only if your invoice has one.",
            note_s,
        )
    ]]
    po_notice_tbl = Table(po_notice_data, colWidths=[174*mm])
    po_notice_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), HexColor("#FFF8E1")),
        ("TOPPADDING",   (0, 0), (-1, -1), 3*mm),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3*mm),
        ("LEFTPADDING",  (0, 0), (-1, -1), 4*mm),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4*mm),
        ("BOX",          (0, 0), (-1, -1), 0.5, HexColor("#F9A825")),
    ]))
    els.append(po_notice_tbl)
    els.append(Spacer(1, 5*mm))

    # ── Submission instructions ────────────────────────────────────────────────
    els.append(Paragraph("How to Submit", section_s))
    els.append(Spacer(1, 2*mm))

    steps_data = [
        [Paragraph("1", label_s), Paragraph("(Optional) Fill in fields to help with extraction accuracy.", body_s)],
        [Paragraph("2", label_s), Paragraph("Log in to the Vendor Portal and choose <b>Upload Invoice PDF</b>.", body_s)],
        [Paragraph("3", label_s), Paragraph("Upload your invoice PDF in any format. System will extract fields best-effort.", body_s)],
        [Paragraph("4", label_s), Paragraph("Review and correct extracted fields, then submit for approval.", body_s)],
    ]
    steps_tbl = Table(steps_data, colWidths=[10*mm, 164*mm])
    steps_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (0, -1), MID_BLUE),
        ("BACKGROUND",   (1, 0), (1, -1), GREY_BG),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",   (0, 0), (-1, -1), 2.5*mm),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 2.5*mm),
        ("LEFTPADDING",  (0, 0), (-1, -1), 3*mm),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3*mm),
        ("INNERGRID",    (0, 0), (-1, -1), 0.25, BORDER),
        ("BOX",          (0, 0), (-1, -1), 0.5, BORDER),
        ("ALIGN",        (0, 0), (0, -1), "CENTER"),
        ("TEXTCOLOR",    (0, 0), (0, -1), white),
    ]))
    els.append(steps_tbl)
    els.append(Spacer(1, 6*mm))

    # ── Divider + footer ───────────────────────────────────────────────────────
    els.append(HRFlowable(width="100%", thickness=0.5, color=BORDER))
    els.append(Spacer(1, 2*mm))
    els.append(Paragraph(
        "Horizon Vendor Portal — Template provided for optional use."
        "Invoices in any format are accepted with review-before-submit workflow.",
        footer_s,
    ))

    # ── Build ───────────────────────────────────────────────────────────────────
    doc.build(els)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=Vendor_Invoice_Template.pdf"
    return response


def _pdf_template_fallback() -> HttpResponse:
    """Plain-text fallback when reportlab is not installed."""
    content = (
        "VENDOR INVOICE SUBMISSION TEMPLATE\n"
        "=" * 50 + "\n\n"
        "Fields to fill in:\n"
        "  Invoice Number  : _______________\n"
        "  Invoice Date    : _______________  (YYYY-MM-DD)\n"
        "  Due Date        : _______________  (YYYY-MM-DD)\n"
        "  PO Number        : _______________\n"
        "  Currency        : _______________  (e.g. INR)\n"
        "  Subtotal Amount  : _______________\n"
        "  Tax Amount       : _______________\n"
        "  Total Amount     : _______________\n"
        "  Description       : _______________\n\n"
        "Note: PO Number is optional. Add it only if your invoice has one.\n"
        "Save this file and re-upload it to the Vendor Portal."
    )
    response = HttpResponse(content, content_type="text/plain")
    response["Content-Disposition"] = "attachment; filename=Vendor_Invoice_Template.txt"
    return response


# ---------------------------------------------------------------------------
# Invoice Payment — post-finance payment recording
# ---------------------------------------------------------------------------

class PaymentPermissionError(PermissionError):
    """Raised when user lacks permission to record a payment."""
    pass


class PaymentValidationError(ValueError):
    """Raised when payment data fails validation."""
    pass


def _now():
    from django.utils import timezone
    return timezone.now()


def _get_invoice_eligible_for_payment(invoice) -> bool:
    """
    Return True if invoice is in a payment-eligible state.

    Eligible: FINANCE_APPROVED only (finance has cleared it, payment can be recorded).
    Not eligible: all other statuses including PAID.
    Once PAID, no new payment recording should be initiated.
    """
    from apps.invoices.models import InvoiceStatus
    return invoice.status == InvoiceStatus.FINANCE_APPROVED


def _get_workflow_participants(invoice) -> list:
    """
    Return all users who acted as participants in the invoice's workflow chain.

    Includes:
    - started_by on the WorkflowInstance
    - assigned_user on every WorkflowInstanceStep for every group in the instance
    """
    from apps.workflow.models import WorkflowInstance, WorkflowInstanceStep

    participants = set()

    instances = WorkflowInstance.objects.filter(
        subject_type="invoice",
        subject_id=invoice.pk,
    ).select_related("started_by").prefetch_related("instance_groups__instance_steps__assigned_user")

    for instance in instances:
        if instance.started_by:
            participants.add(instance.started_by.pk)
        for group in instance.instance_groups.all():
            for step in group.instance_steps.all():
                if step.assigned_user_id:
                    participants.add(step.assigned_user_id)

    return list(participants)


def can_user_record_invoice_payment(user, invoice) -> bool:
    """
    Authorization rule for payment recording on an invoice.

    User can record payment if ALL of:
      1. Invoice is in FINANCE_APPROVED status (payment-eligible)
      2. User is a resolved finance recipient for the invoice handoff,
         has a configured finance role, OR is a superuser/org_admin/tenant_admin

    Permission model:
    - Internal workflow participation is NOT sufficient.
    - Finance review and payment recording must use the same finance-recipient
      authority so a finance user who approved the handoff can record payment.
    """
    if not _get_invoice_eligible_for_payment(invoice):
        return False

    if user.is_superuser:
        return True

    # Resolve assigned role codes directly from DB so this check is not dependent
    # on serializer-populated convenience attributes.
    from apps.access.models import UserRoleAssignment

    user_roles = set(
        UserRoleAssignment.objects.filter(user=user, role__is_active=True)
        .values_list("role__code", flat=True)
    )

    # Admin-level roles (same pattern as elsewhere in the codebase)
    if any(r in ("org_admin", "tenant_admin") for r in user_roles):
        return True

    try:
        from apps.finance.services import _get_finance_role_codes
        finance_role_codes = _get_finance_role_codes()
    except Exception:
        finance_role_codes = {"finance_team"}

    if user_roles.intersection(finance_role_codes):
        return True

    email = (getattr(user, "email", "") or "").strip().lower()
    if not email:
        return False

    try:
        from apps.finance.models import FinanceHandoff, FinanceHandoffStatus
        from apps.finance.services import NoFinanceRecipientsError, resolve_finance_recipients_for_handoff

        handoffs = FinanceHandoff.objects.filter(
            module="invoice",
            subject_type="invoice",
            subject_id=invoice.pk,
            status__in=[
                FinanceHandoffStatus.FINANCE_APPROVED,
                FinanceHandoffStatus.SENT,
                FinanceHandoffStatus.PENDING,
            ],
        ).order_by("-updated_at", "-id")

        for handoff in handoffs:
            try:
                recipients = resolve_finance_recipients_for_handoff(handoff)
            except NoFinanceRecipientsError:
                continue
            if email in {r.strip().lower() for r in recipients if r}:
                return True
    except Exception:
        pass

    return False


def get_or_create_invoice_payment(invoice) -> "InvoicePayment":
    """Return existing payment record or create a new pending one."""
    from apps.invoices.models import InvoicePayment, InvoicePaymentStatus

    payment, created = InvoicePayment.objects.get_or_create(
        invoice=invoice,
        defaults={"payment_status": InvoicePaymentStatus.PENDING},
    )
    return payment


def record_invoice_payment(
    invoice,
    actor,
    data: dict,
) -> "InvoicePayment":
    """
    Create or update the payment record for an invoice.

    If no payment record exists yet, creates one in PENDING.
    Subsequent calls update the existing record.

    Validation (raises PaymentValidationError):
      - Invoice must be in FINANCE_APPROVED status
      - When marking PAID: payment_date, paid_amount > 0, and utr_number

    Permission (raises PaymentPermissionError):
      - Actor must pass can_user_record_invoice_payment()

    Side effects: NONE — no workflow mutation, no budget changes.
    """
    from apps.invoices.models import InvoicePayment, InvoicePaymentStatus, PaymentMethod

    # Eligibility check FIRST — before permission
    # This tells the caller WHY payment cannot be recorded
    if not _get_invoice_eligible_for_payment(invoice):
        raise PaymentValidationError(
            "Payment can only be recorded after finance approval. "
            f"Invoice is in '{invoice.status}' status."
        )

    # Permission check
    if not can_user_record_invoice_payment(actor, invoice):
        raise PaymentPermissionError(
            "You do not have permission to record payment for this invoice."
        )

    # Get or create payment record
    payment = get_or_create_invoice_payment(invoice)

    # Validate PAID status requirements at service layer
    if data.get("payment_status") == InvoicePaymentStatus.PAID:
        errors = {}
        if not data.get("payment_date"):
            errors.setdefault("payment_date", []).append("Payment date is required when marking as paid.")
        amount = data.get("paid_amount")
        if not amount or amount <= 0:
            errors.setdefault("paid_amount", []).append("Paid amount must be greater than zero when marking as paid.")
        utr = (data.get("utr_number") or "").strip()
        if not utr:
            errors.setdefault("utr_number", []).append(
                "UTR number is required when marking as paid."
            )
        if errors:
            raise PaymentValidationError(errors)

    # Apply field updates
    allowed_fields = [
        "payment_status", "payment_method",
        "payment_reference_number", "utr_number",
        "transaction_id", "bank_reference_number",
        "payer_bank_name", "beneficiary_name", "beneficiary_bank_name",
        "paid_amount", "currency", "payment_date", "remarks",
    ]
    for field in allowed_fields:
        if field in data:
            setattr(payment, field, data[field])

    # On first record (recorded_by is null), set audit fields
    if payment.recorded_by_id is None:
        payment.recorded_by = actor
        payment.recorded_at = _now()

    # Always update updated_by
    payment.updated_by = actor

    payment.save()

    # If marked PAID, update invoice status to PAID
    if data.get("payment_status") == InvoicePaymentStatus.PAID:
        invoice.status = "paid"
        invoice.save(update_fields=["status", "updated_at"])

    return payment
