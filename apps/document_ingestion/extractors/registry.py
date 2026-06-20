import csv
import io
import json
import re
from pathlib import Path

from django.conf import settings
from openpyxl import load_workbook
from PyPDF2 import PdfReader

from apps.document_ingestion.extractors.base import ExtractedRecord, ExtractionError, ExtractionResult
from apps.document_ingestion.extractors.normalization import (
    confidence_for,
    infer_document_type,
    normalize_amount,
    normalize_record,
)


TEXT_PATTERNS = {
    "vendor_code": [r"(?:SAP\s*)?Vendor\s*(?:Code|ID)\s*[:#-]?\s*([A-Z0-9_-]+)"],
    "vendor_name": [r"Vendor\s*Name\s*[:#-]?\s*([^\r\n]+)"],
    "vendor_email": [r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}"],
    "invoice_number": [r"Invoice\s*(?:Number|No\.?|#)\s*[:#-]?\s*([A-Z0-9/_-]+)"],
    "sap_document_number": [r"SAP\s*Document\s*(?:Number|No\.?)\s*[:#-]?\s*([A-Z0-9/_-]+)"],
    "utr_number": [r"UTR\s*(?:Number|No\.?|#)?\s*[:#-]?\s*([A-Z0-9/_-]+)"],
    "payment_reference_number": [r"Payment\s*Reference\s*(?:Number|No\.?)?\s*[:#-]?\s*([A-Z0-9/_-]+)"],
    "payment_date": [r"Payment\s*Date\s*[:#-]?\s*([0-9]{1,4}[-/.][0-9]{1,2}[-/.][0-9]{1,4})"],
    "invoice_date": [r"Invoice\s*Date\s*[:#-]?\s*([0-9]{1,4}[-/.][0-9]{1,2}[-/.][0-9]{1,4})"],
    "gstin": [r"GSTIN\s*[:#-]?\s*([0-9A-Z]{15})"],
    "pan": [r"PAN\s*[:#-]?\s*([A-Z]{5}[0-9]{4}[A-Z])"],
    "amount": [r"(?:Paid\s*Amount|Invoice\s*Amount|Total\s*Amount|Grand\s*Total)\s*[:#-]?\s*(?:INR|Rs\.?|₹)?\s*([0-9,]+(?:\.\d{1,2})?)"],
}


def _record(raw: dict, filename: str) -> ExtractedRecord:
    normalized = normalize_record(raw)
    document_type = infer_document_type(normalized, filename)
    confidence, errors = confidence_for(normalized, document_type)
    return ExtractedRecord(raw, normalized, document_type, confidence, errors)


def _extract_text_fields(text: str) -> dict:
    result = {"full_text": text}
    for field, patterns in TEXT_PATTERNS.items():
        for pattern in patterns:
            match = re.search(pattern, text, flags=re.IGNORECASE)
            if match:
                result[field] = match.group(1) if match.lastindex else match.group(0)
                break
    if result.get("amount"):
        result["amount"] = normalize_amount(result["amount"])
    return result


def _pdf_text(content: bytes) -> tuple[str, str]:
    try:
        reader = PdfReader(io.BytesIO(content))
        max_pages = getattr(settings, "DOCUMENT_INGESTION_MAX_PDF_PAGES", 50)
        if len(reader.pages) > max_pages:
            raise ExtractionError(f"PDF exceeds the {max_pages}-page processing limit.")
        text = "\n".join(page.extract_text() or "" for page in reader.pages).strip()
    except Exception as exc:
        raise ExtractionError(f"PDF could not be read: {exc}") from exc
    if text:
        return text, "pdf_text"
    try:
        import fitz
        import pytesseract
        from PIL import Image

        doc = fitz.open(stream=content, filetype="pdf")
        max_pages = getattr(settings, "DOCUMENT_INGESTION_MAX_PDF_PAGES", 50)
        if doc.page_count > max_pages:
            raise ExtractionError(f"PDF exceeds the {max_pages}-page processing limit.")
        pages = []
        for page in doc:
            pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            image = Image.open(io.BytesIO(pixmap.tobytes("png")))
            pages.append(pytesseract.image_to_string(image))
        return "\n".join(pages).strip(), "pdf_ocr"
    except Exception as exc:
        raise ExtractionError(f"Scanned PDF OCR failed: {exc}") from exc


def extract_document(content: bytes, filename: str) -> ExtractionResult:
    suffix = Path(filename).suffix.lower()
    if suffix == ".pdf":
        text, extractor = _pdf_text(content)
        raw = _extract_text_fields(text)
        return ExtractionResult({"text": text}, [_record(raw, filename)], extractor)

    if suffix in {".xlsx", ".xlsm"}:
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        except Exception as exc:
            raise ExtractionError(f"Excel workbook could not be read: {exc}") from exc
        if not rows:
            raise ExtractionError("Excel workbook is empty.")
        headers = [str(value or "").strip() for value in rows[0]]
        raw_records = [dict(zip(headers, row)) for row in rows[1:] if any(value not in (None, "") for value in row)]
        return ExtractionResult({"sheet": sheet.title, "row_count": len(raw_records)}, [_record(raw, filename) for raw in raw_records], "excel")

    if suffix == ".csv":
        try:
            decoded = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            decoded = content.decode("latin-1")
        raw_records = list(csv.DictReader(io.StringIO(decoded)))
        return ExtractionResult({"row_count": len(raw_records)}, [_record(raw, filename) for raw in raw_records], "csv")

    if suffix == ".json":
        try:
            payload = json.loads(content.decode("utf-8-sig"))
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise ExtractionError(f"JSON could not be read: {exc}") from exc
        raw_records = payload if isinstance(payload, list) else [payload]
        if not all(isinstance(item, dict) for item in raw_records):
            raise ExtractionError("JSON must contain an object or list of objects.")
        return ExtractionResult({"record_count": len(raw_records)}, [_record(raw, filename) for raw in raw_records], "json")

    if suffix in {".txt", ".text"}:
        text = content.decode("utf-8-sig", errors="replace")
        raw = _extract_text_fields(text)
        return ExtractionResult({"text": text}, [_record(raw, filename)], "text")

    raise ExtractionError(f"Unsupported file type '{suffix or 'unknown'}'.")
