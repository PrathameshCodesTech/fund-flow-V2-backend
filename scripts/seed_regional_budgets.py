from decimal import Decimal
import re

from django.db import transaction
from openpyxl import load_workbook

from apps.budgets.models import Budget, BudgetCategory, BudgetLine, BudgetStatus
from apps.core.models import Organization, ScopeNode


WORKBOOK_PATH = r"C:\Users\Prathmesh Marathe\Downloads\Marketing Budget_FY27_V2_10Mar26.xlsx"

SHEET_CONFIG = {
    "North": {
        "budget_name": "FY27 Marketing - North",
        "budget_code": "FY27-MKT-NORTH",
        "scope_code": "north",
        "header_row": 1,
        "total_row": 7,
        "first_category_col": 7,
        "last_category_col": 17,
    },
    "South": {
        "budget_name": "FY27 Marketing - South",
        "budget_code": "FY27-MKT-SOUTH",
        "scope_code": "south",
        "header_row": 2,
        "total_row": 17,
        "first_category_col": 7,
        "last_category_col": 17,
    },
    "West": {
        "budget_name": "FY27 Marketing - West",
        "budget_code": "FY27-MKT-WEST",
        "scope_code": "west",
        "header_row": 2,
        "total_row": 12,
        "first_category_col": 7,
        "last_category_col": 16,
    },
    "Incity": {
        "budget_name": "FY27 Marketing - Incity",
        "budget_code": "FY27-MKT-INCITY",
        "scope_code": "incity",
        "header_row": 2,
        "total_row": 20,
        "first_category_col": 8,
        "last_category_col": 16,
    },
}


def to_decimal(value) -> Decimal:
    if value in (None, "", "-"):
        return Decimal("0")
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    cleaned = str(value).replace(",", "").strip()
    if not cleaned or cleaned == "-":
        return Decimal("0")
    return Decimal(cleaned)


def make_code(value: str, fallback: str) -> str:
    code = re.sub(r"[^A-Z0-9]+", "-", value.upper()).strip("-")
    return code[:50] or fallback


def get_or_create_category(org, name: str, fallback_code: str):
    category = BudgetCategory.objects.filter(org=org, name=name).first()
    if category:
        return category
    base_code = make_code(name, fallback_code)
    code = base_code
    suffix = 2
    while BudgetCategory.objects.filter(org=org, code=code).exists():
        code = f"{base_code[:45]}-{suffix}"
        suffix += 1
    return BudgetCategory.objects.create(
        org=org,
        name=name,
        code=code,
        is_active=True,
    )


def parse_regional_totals():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    parsed = {}
    for sheet_name, cfg in SHEET_CONFIG.items():
        sheet = workbook[sheet_name]
        categories = []
        total_allocated = Decimal("0")
        for col in range(cfg["first_category_col"], cfg["last_category_col"] + 1):
            raw_header = sheet.cell(cfg["header_row"], col).value
            header = str(raw_header).strip() if raw_header is not None else ""
            if not header or header.lower() == "total":
                continue
            allocated = to_decimal(sheet.cell(cfg["total_row"], col).value)
            if allocated == 0:
                continue
            categories.append({"name": header, "allocated": allocated})
            total_allocated += allocated
        parsed[sheet_name] = {"categories": categories, "total": total_allocated}
    return parsed


@transaction.atomic
def seed():
    parsed = parse_regional_totals()
    org = Organization.objects.get(code="horizon")

    for sheet_name, cfg in SHEET_CONFIG.items():
        scope = ScopeNode.objects.get(org=org, code=cfg["scope_code"])
        BudgetLine.objects.filter(budget__code=cfg["budget_code"]).delete()
        Budget.objects.filter(code=cfg["budget_code"]).delete()

        budget = Budget.objects.create(
            org=org,
            scope_node=scope,
            name=cfg["budget_name"],
            code=cfg["budget_code"],
            financial_year="2026-27",
            status=BudgetStatus.ACTIVE,
            allocated_amount=parsed[sheet_name]["total"],
            reserved_amount=Decimal("0"),
            consumed_amount=Decimal("0"),
        )

        for index, item in enumerate(parsed[sheet_name]["categories"], start=1):
            category = get_or_create_category(
                org=org,
                name=item["name"],
                fallback_code=f"{cfg['scope_code'].upper()}-{index}",
            )
            BudgetLine.objects.create(
                budget=budget,
                category=category,
                subcategory=None,
                allocated_amount=item["allocated"],
                reserved_amount=Decimal("0"),
                consumed_amount=Decimal("0"),
            )

        print(f"{budget.name}: allocated={budget.allocated_amount} categories={len(parsed[sheet_name]['categories'])}")
        for item in parsed[sheet_name]["categories"]:
            print(f"  - {item['name']}: {item['allocated']}")


seed()
