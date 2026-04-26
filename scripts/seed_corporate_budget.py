from collections import OrderedDict
from decimal import Decimal
import re

from django.db import transaction
from openpyxl import load_workbook

from apps.budgets.models import (
    Budget,
    BudgetCategory,
    BudgetConsumption,
    BudgetLine,
    BudgetRule,
    BudgetSubCategory,
    BudgetVarianceRequest,
    BudgetStatus,
    ConsumptionStatus,
    ConsumptionType,
    SourceType,
)
from apps.core.models import Organization, ScopeNode


WORKBOOK_PATH = r"C:\Users\Prathmesh Marathe\Downloads\Marketing Budget_FY27_V2_10Mar26.xlsx"
SHEET_NAME = "FY26 Master Budget_v1"
ALLOWED_SECTIONS = OrderedDict(
    [
        ("Customer/IPC Events", []),
        ("Retainerships/ annual payments", []),
        ("Content Marketing & Assets", []),
        ("Print/Outdoor", []),
        ("Others", []),
        ("ESG Initiatives", []),
    ]
)
EXCLUDED_SECTION_HEADERS = {
    "Project specific (incl. Core)",
    "TOTAL CORPORATE (A)",
    "TOTAL PROJECTS (B)",
    "TOTAL CORP + PROJECTS + ESG (A+B+C)",
}


def to_decimal(value) -> Decimal:
    if value in (None, "", "-"):
        return Decimal("0")
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    cleaned = str(value).replace(",", "").strip()
    if not cleaned or cleaned == "-":
        return Decimal("0")
    return Decimal(cleaned)


def make_code(value: str, fallback: str) -> str:
    code = re.sub(r"[^A-Z0-9]+", "-", value.upper()).strip("-")
    return code[:50] or fallback


def parse_corporate_sections():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook[SHEET_NAME]

    current_category = None
    for row in sheet.iter_rows(values_only=True):
        raw_name = row[1]
        if not raw_name:
            continue
        name = str(raw_name).strip()

        if name in ALLOWED_SECTIONS:
            current_category = name
            continue
        if name in EXCLUDED_SECTION_HEADERS:
            current_category = None
            continue
        if current_category is None:
            continue
        if name.lower().startswith("total"):
            continue

        allocated = to_decimal(row[4])
        consumed = to_decimal(row[5])
        if allocated == 0 and consumed == 0:
            continue

        ALLOWED_SECTIONS[current_category].append(
            {
                "name": name,
                "allocated": allocated,
                "consumed": consumed,
            }
        )

    merged = OrderedDict()
    for category_name, items in ALLOWED_SECTIONS.items():
        merged_items = OrderedDict()
        for item in items:
            if item["name"] not in merged_items:
                merged_items[item["name"]] = {
                    "allocated": Decimal("0"),
                    "consumed": Decimal("0"),
                }
            merged_items[item["name"]]["allocated"] += item["allocated"]
            merged_items[item["name"]]["consumed"] += item["consumed"]
        merged[category_name] = merged_items
    return merged


def wipe_budget_data():
    BudgetConsumption.objects.all().delete()
    BudgetRule.objects.all().delete()
    BudgetVarianceRequest.objects.all().delete()
    BudgetLine.objects.all().delete()
    BudgetSubCategory.objects.all().delete()
    BudgetCategory.objects.all().delete()
    Budget.objects.all().delete()


@transaction.atomic
def seed():
    corporate_tree = parse_corporate_sections()

    wipe_budget_data()

    org = Organization.objects.get(code="horizon")
    scope = ScopeNode.objects.get(org=org, code="corporate")
    budget = Budget.objects.create(
        org=org,
        scope_node=scope,
        name="FY27 Marketing - Corporate",
        code="FY27-MKT-CORP",
        financial_year="2026-27",
        status=BudgetStatus.ACTIVE,
        allocated_amount=Decimal("0"),
        reserved_amount=Decimal("0"),
        consumed_amount=Decimal("0"),
    )

    total_allocated = Decimal("0")
    total_consumed = Decimal("0")

    for category_index, (category_name, subcategories) in enumerate(corporate_tree.items(), start=1):
        category = BudgetCategory.objects.create(
            org=org,
            name=category_name,
            code=make_code(category_name, f"CAT-{category_index}"),
            is_active=True,
        )

        for subcat_index, (subcategory_name, values) in enumerate(subcategories.items(), start=1):
            subcategory = BudgetSubCategory.objects.create(
                category=category,
                name=subcategory_name,
                code=make_code(subcategory_name, f"SUB-{category_index}-{subcat_index}"),
                is_active=True,
            )
            line = BudgetLine.objects.create(
                budget=budget,
                category=category,
                subcategory=subcategory,
                allocated_amount=values["allocated"],
                reserved_amount=Decimal("0"),
                consumed_amount=values["consumed"],
            )
            if values["consumed"] > 0:
                BudgetConsumption.objects.create(
                    budget=budget,
                    budget_line=line,
                    source_type=SourceType.MANUAL_ADJUSTMENT,
                    source_id=f"seed-corporate-{line.id}",
                    amount=values["consumed"],
                    consumption_type=ConsumptionType.CONSUMED,
                    status=ConsumptionStatus.APPLIED,
                    note="FY26 spent baseline seeded for FY27 UAT",
                )
            total_allocated += values["allocated"]
            total_consumed += values["consumed"]

    budget.allocated_amount = total_allocated
    budget.reserved_amount = Decimal("0")
    budget.consumed_amount = total_consumed
    budget.save(update_fields=["allocated_amount", "reserved_amount", "consumed_amount", "updated_at"])

    print(f"BUDGET {budget.id} {budget.name}")
    print(f"ALLOCATED {budget.allocated_amount}")
    print(f"CONSUMED {budget.consumed_amount}")
    print(f"RESERVED {budget.reserved_amount}")
    print(f"CATEGORIES {BudgetCategory.objects.count()}")
    print(f"SUBCATEGORIES {BudgetSubCategory.objects.count()}")
    print(f"LINES {BudgetLine.objects.count()}")
    print(f"CONSUMPTIONS {BudgetConsumption.objects.count()}")
    print("CATEGORY_BREAKDOWN")
    for category_name, subcategories in corporate_tree.items():
        cat_allocated = sum((item["allocated"] for item in subcategories.values()), Decimal("0"))
        cat_consumed = sum((item["consumed"] for item in subcategories.values()), Decimal("0"))
        print(f"- {category_name}: allocated={cat_allocated} consumed={cat_consumed} subcats={len(subcategories)}")
        for subcategory_name, values in subcategories.items():
            print(
                f"  * {subcategory_name}: allocated={values['allocated']} consumed={values['consumed']}"
            )


seed()
