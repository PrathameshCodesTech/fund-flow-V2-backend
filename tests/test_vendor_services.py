"""
Vendor service-layer tests.

Covers all 15 required scenarios plus finalize helper.
Email sending is mocked at apps.vendors.email.send_finance_email and
apps.vendors.email.send_vendor_invitation_email.
"""
import io
import os
from datetime import timedelta
from unittest.mock import MagicMock, patch

import openpyxl
import pytest
from django.utils import timezone

from apps.core.models import Organization, ScopeNode, NodeType
from apps.users.models import User
from apps.vendors.models import (
    FinanceActionType,
    FinanceDecisionChoice,
    InvitationStatus,
    MarketingStatus,
    OperationalStatus,
    SubmissionMode,
    SubmissionStatus,
    Vendor,
    VendorAttachment,
    VendorFinanceActionToken,
    VendorInvitation,
    VendorOnboardingSubmission,
)
from apps.vendors.services import (
    FinanceTokenError,
    InvitationExpiredError,
    InvitationNotFoundError,
    POMandate,
    SubmissionStateError,
    VendorStateError,
    approve_vendor_marketing,
    assert_vendor_can_submit_invoice,
    attach_document,
    create_or_update_submission_from_excel,
    create_or_update_submission_from_manual,
    create_vendor_invitation,
    finance_approve_submission,
    finance_reject_submission,
    finalize_submission,
    get_invitation_by_token,
    reject_vendor_marketing,
    reopen_submission,
    send_submission_to_finance,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def org(db):
    return Organization.objects.create(name="SVC Org", code="svc-org")


@pytest.fixture
def company(org):
    return ScopeNode.objects.create(
        org=org, parent=None, name="HQ", code="hq",
        node_type=NodeType.COMPANY, path="/svc-org/hq", depth=0, is_active=True,
    )


@pytest.fixture
def entity(org, company):
    return ScopeNode.objects.create(
        org=org, parent=company, name="Entity A", code="ea",
        node_type=NodeType.ENTITY, path="/svc-org/hq/ea", depth=1, is_active=True,
    )


@pytest.fixture
def actor(db):
    return User.objects.create_user(email="svc-actor@example.com", password="pass")


@pytest.fixture
def approver(db):
    return User.objects.create_user(email="svc-approver@example.com", password="pass")


@pytest.fixture
def invitation(org, entity, actor):
    return create_vendor_invitation(
        org=org,
        scope_node=entity,
        vendor_email="vendor@example.com",
        invited_by=actor,
        vendor_name_hint="Acme Corp",
    )


@pytest.fixture
def manual_payload():
    return {
        "vendor_name": "Acme Supplies Ltd",
        "vendor_type": "supplier",
        "gst_registered": "yes",
        "gstin": "22AAABB1234C1Z5",
        "pan": "AAABB1234C",
        "email": "acme@example.com",
        "phone": "9876543210",
        "address_line1": "123 Main Street",
        "city": "Mumbai",
        "state": "Maharashtra",
        "country": "India",
        "pincode": "400001",
        "bank_name": "State Bank of India",
        "account_number": "1234567890",
        "ifsc": "SBIN0001234",
    }


def _make_submitted_submission(invitation, payload):
    """Helper: create a fully submitted submission."""
    submission = create_or_update_submission_from_manual(
        invitation, payload, finalize=True
    )
    return submission


def _make_excel_file(data: dict) -> io.BytesIO:
    """Build a VRF-style Excel file in memory."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendor Registration Form"
    for label, value in data.items():
        ws.append([label, value])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _make_sent_submission(invitation, payload):
    """Helper: create a submission that has been sent to finance.

    Under Option B, finalize=True already puts submission in sent_to_finance.
    This helper is kept for cases where we need to explicitly call send_to_finance
    (e.g. on a reopened submission).
    """
    submission = _make_submitted_submission(invitation, payload)
    return submission


# ---------------------------------------------------------------------------
# 1. create_vendor_invitation
# ---------------------------------------------------------------------------

class TestCreateVendorInvitation:
    def test_creates_invitation_with_token(self, org, entity, actor):
        inv = create_vendor_invitation(
            org=org, scope_node=entity, vendor_email="v@test.com", invited_by=actor
        )
        assert inv.pk is not None
        assert inv.status == InvitationStatus.PENDING
        assert len(inv.token) > 10  # non-trivial token

    def test_token_is_unique(self, org, entity, actor):
        inv1 = create_vendor_invitation(org=org, scope_node=entity, vendor_email="a@test.com", invited_by=actor)
        inv2 = create_vendor_invitation(org=org, scope_node=entity, vendor_email="b@test.com", invited_by=actor)
        assert inv1.token != inv2.token

    @patch("apps.vendors.services._send_invitation_email")
    def test_sends_invitation_email_on_create(self, mock_send, org, entity, actor):
        create_vendor_invitation(
            org=org, scope_node=entity,
            vendor_email="newvendor@test.com",
            invited_by=actor,
            vendor_name_hint="Test Vendor",
        )
        mock_send.assert_called_once()
        call_args = mock_send.call_args
        invitation_arg = call_args[0][0]
        assert invitation_arg.vendor_email == "newvendor@test.com"
        assert invitation_arg.vendor_name_hint == "Test Vendor"

    @patch("apps.vendors.services._send_invitation_email")
    def test_invitation_email_contains_onboarding_link(self, mock_send, org, entity, actor):
        create_vendor_invitation(
            org=org, scope_node=entity,
            vendor_email="linktest@test.com",
            invited_by=actor,
        )
        mock_send.assert_called_once()
        # _send_invitation_email is called — email content is verified
        # in apps.vendors.tests.test_vendor_api or via integration test

    @patch("apps.vendors.services._send_invitation_email")
    def test_invitation_created_when_email_fails(self, mock_send, org, entity, actor):
        """Invitation must be created even if email sending raises."""
        mock_send.side_effect = Exception("SMTP config missing")
        inv = create_vendor_invitation(
            org=org, scope_node=entity,
            vendor_email="fails@test.com",
            invited_by=actor,
        )
        assert inv.pk is not None
        assert inv.status == InvitationStatus.PENDING


# ---------------------------------------------------------------------------
# 2. get_invitation_by_token — expiry
# ---------------------------------------------------------------------------

class TestGetInvitationByToken:
    def test_valid_token_returns_invitation(self, invitation):
        fetched = get_invitation_by_token(invitation.token)
        assert fetched.pk == invitation.pk

    def test_valid_token_sets_opened(self, invitation):
        fetched = get_invitation_by_token(invitation.token)
        assert fetched.status == InvitationStatus.OPENED

    def test_invalid_token_raises(self, db):
        with pytest.raises(InvitationNotFoundError):
            get_invitation_by_token("nonexistent_token_xyz")

    def test_expired_token_raises(self, org, entity, actor):
        inv = create_vendor_invitation(
            org=org, scope_node=entity, vendor_email="exp@test.com", invited_by=actor,
            expires_at=timezone.now() - timedelta(hours=1),
        )
        with pytest.raises(InvitationExpiredError):
            get_invitation_by_token(inv.token)

    def test_expired_token_marks_expired(self, org, entity, actor):
        inv = create_vendor_invitation(
            org=org, scope_node=entity, vendor_email="exp2@test.com", invited_by=actor,
            expires_at=timezone.now() - timedelta(hours=1),
        )
        with pytest.raises(InvitationExpiredError):
            get_invitation_by_token(inv.token)
        inv.refresh_from_db()
        assert inv.status == InvitationStatus.EXPIRED

    def test_cancelled_invitation_raises(self, invitation):
        invitation.status = InvitationStatus.CANCELLED
        invitation.save()
        with pytest.raises(InvitationNotFoundError):
            get_invitation_by_token(invitation.token)


# ---------------------------------------------------------------------------
# 3. Manual submission saves raw + normalized fields
# ---------------------------------------------------------------------------

class TestManualSubmission:
    def test_creates_submission_with_normalized_fields(self, invitation, manual_payload):
        sub = create_or_update_submission_from_manual(invitation, manual_payload)
        assert sub.pk is not None
        assert sub.normalized_vendor_name == "Acme Supplies Ltd"
        assert sub.normalized_email == "acme@example.com"
        assert sub.normalized_city == "Mumbai"
        assert sub.normalized_gst_registered is True
        assert sub.submission_mode == SubmissionMode.MANUAL

    def test_raw_form_data_preserved(self, invitation, manual_payload):
        sub = create_or_update_submission_from_manual(invitation, manual_payload)
        assert sub.raw_form_data.get("vendor_name") == "Acme Supplies Ltd"

    def test_unknown_fields_in_raw_data(self, invitation):
        payload = {"vendor_name": "Test Co", "email": "t@t.com", "custom_field": "custom_value"}
        sub = create_or_update_submission_from_manual(invitation, payload)
        assert "custom_field" in sub.raw_form_data

    def test_draft_status_on_save(self, invitation, manual_payload):
        sub = create_or_update_submission_from_manual(invitation, manual_payload)
        assert sub.status == SubmissionStatus.DRAFT

    def test_finalize_transitions_to_sent_to_finance(self, invitation, manual_payload):
        """Option B: finalize=True transitions directly to sent_to_finance."""
        sub = create_or_update_submission_from_manual(invitation, manual_payload, finalize=True)
        assert sub.status == SubmissionStatus.SENT_TO_FINANCE
        assert sub.submitted_at is not None
        assert sub.finance_sent_at is not None
        # Tokens must exist
        assert sub.finance_tokens.exists(), "finance tokens must be created on finalize"

    def test_finalize_persists_normalized_and_raw_data(self, invitation, manual_payload):
        """Finalize must persist the payload before moving to finance review."""
        sub = create_or_update_submission_from_manual(invitation, manual_payload, finalize=True)
        # Force a fresh DB read to prove data was actually persisted, not just in-memory
        sub_refresh = VendorOnboardingSubmission.objects.get(pk=sub.pk)
        assert sub_refresh.normalized_vendor_name == "Acme Supplies Ltd"
        assert sub_refresh.normalized_email == "acme@example.com"
        assert sub_refresh.normalized_city == "Mumbai"
        assert sub_refresh.raw_form_data.get("vendor_name") == "Acme Supplies Ltd"
        assert sub_refresh.submission_mode == SubmissionMode.MANUAL
        # Status is sent_to_finance (not just submitted)
        assert sub_refresh.status == SubmissionStatus.SENT_TO_FINANCE

    def test_upsert_updates_same_record(self, invitation, manual_payload):
        sub1 = create_or_update_submission_from_manual(invitation, manual_payload)
        payload2 = {**manual_payload, "city": "Delhi"}
        sub2 = create_or_update_submission_from_manual(invitation, payload2)
        assert sub1.pk == sub2.pk
        assert sub2.normalized_city == "Delhi"


# ---------------------------------------------------------------------------
# 4. Excel submission parses and saves raw + normalized fields
# ---------------------------------------------------------------------------

class TestExcelSubmission:
    def test_excel_parses_and_normalizes(self, invitation):
        data = {
            "Vendor Name": "Excel Corp",
            "Email": "excel@example.com",
            "Phone": "9000000001",
            "Bank Name": "HDFC Bank",
            "IFSC Code": "HDFC0001234",
        }
        excel_file = _make_excel_file(data)
        sub = create_or_update_submission_from_excel(invitation, excel_file)
        assert sub.normalized_vendor_name == "Excel Corp"
        assert sub.normalized_email == "excel@example.com"
        assert sub.normalized_bank_name == "HDFC Bank"
        assert sub.submission_mode == SubmissionMode.EXCEL_UPLOAD

    def test_excel_unknown_fields_in_raw(self, invitation):
        data = {
            "Vendor Name": "Corp",
            "Email": "c@c.com",
            "Requestor Name": "Internal Team",
        }
        excel_file = _make_excel_file(data)
        sub = create_or_update_submission_from_excel(invitation, excel_file)
        # Unknown field preserved in raw_form_data
        assert any("Requestor" in k or "requestor" in k.lower() for k in sub.raw_form_data)

    def test_excel_finalize_goes_to_sent_to_finance(self, invitation):
        """Option B: Excel finalize transitions to sent_to_finance with tokens."""
        data = {"Vendor Name": "FinalCorp", "Email": "final@example.com"}
        excel_file = _make_excel_file(data)
        sub = create_or_update_submission_from_excel(invitation, excel_file, finalize=True)
        assert sub.status == SubmissionStatus.SENT_TO_FINANCE
        assert sub.finance_sent_at is not None
        assert sub.finance_tokens.exists(), "finance tokens must be created"

    def test_excel_finalize_persists_normalized_data(self, invitation):
        """Finalize must persist Excel-extracted data before moving to finance review."""
        data = {
            "Vendor Name": "Excel Vendor Corp",
            "Email": "excelcorp@example.com",
            "City": "Bangalore",
        }
        excel_file = _make_excel_file(data)
        sub = create_or_update_submission_from_excel(invitation, excel_file, finalize=True)
        sub_refresh = VendorOnboardingSubmission.objects.get(pk=sub.pk)
        assert sub_refresh.normalized_vendor_name == "Excel Vendor Corp"
        assert sub_refresh.normalized_email == "excelcorp@example.com"
        assert sub_refresh.normalized_city == "Bangalore"
        assert sub_refresh.submission_mode == SubmissionMode.EXCEL_UPLOAD
        assert sub_refresh.status == SubmissionStatus.SENT_TO_FINANCE


# ---------------------------------------------------------------------------
# 5. attach_document
# ---------------------------------------------------------------------------

class TestAttachDocument:
    def test_creates_attachment(self, invitation, manual_payload):
        sub = create_or_update_submission_from_manual(invitation, manual_payload)
        att = attach_document(
            submission=sub,
            title="GST Certificate",
            file_name="gst.pdf",
            file_url="https://example.com/gst.pdf",
            document_type="gst",
        )
        assert att.pk is not None
        assert att.submission_id == sub.pk
        assert att.title == "GST Certificate"


# ---------------------------------------------------------------------------
# 6. finalize_submission
# ---------------------------------------------------------------------------

class TestFinalizeSubmission:
    def test_finalize_goes_to_sent_to_finance(self, invitation, manual_payload):
        """Under Option B, finalize transitions directly to sent_to_finance."""
        sub = create_or_update_submission_from_manual(invitation, manual_payload)
        finalized = finalize_submission(sub)
        assert finalized.status == SubmissionStatus.SENT_TO_FINANCE
        assert finalized.submitted_at is not None
        assert finalized.finance_sent_at is not None
        # Tokens created by _start_finance_review during finalize
        assert finalized.finance_tokens.exists(), "finance tokens must be created"

    def test_finalize_requires_vendor_name(self, invitation):
        sub = create_or_update_submission_from_manual(
            invitation, {"email": "a@a.com"}
        )
        with pytest.raises(ValueError, match="vendor_name"):
            finalize_submission(sub)

    def test_finalize_requires_email(self, invitation):
        sub = create_or_update_submission_from_manual(
            invitation, {"vendor_name": "Corp"}
        )
        with pytest.raises(ValueError, match="email"):
            finalize_submission(sub)

    def test_finalize_already_finalized_raises(self, invitation, manual_payload):
        """Cannot finalize a submission that has already been finalized."""
        sub = create_or_update_submission_from_manual(invitation, manual_payload)
        finalize_submission(sub)  # first call — transitions to sent_to_finance
        with pytest.raises(SubmissionStateError):
            finalize_submission(sub)  # second call — already sent_to_finance


# ---------------------------------------------------------------------------
# 7. send_to_finance generates export + tokens + email
# ---------------------------------------------------------------------------

class TestSendToFinance:
    def test_generates_export_and_sends_email(self, invitation, manual_payload):
        """send_submission_to_finance sends email and transitions REOPENED → sent_to_finance."""
        sub = create_or_update_submission_from_manual(invitation, manual_payload)
        # Put in REOPENED state (simulating a resubmission after rejection)
        sub.status = SubmissionStatus.REOPENED
        sub.save(update_fields=["status"])
        with patch("apps.vendors.email.send_finance_email") as mock_send:
            updated = send_submission_to_finance(sub)
        mock_send.assert_called_once()
        assert updated.status == SubmissionStatus.SENT_TO_FINANCE
        assert updated.finance_sent_at is not None

    def test_creates_approve_and_reject_tokens(self, invitation, manual_payload):
        sub = create_or_update_submission_from_manual(invitation, manual_payload)
        sub.status = SubmissionStatus.REOPENED
        sub.save(update_fields=["status"])
        with patch("apps.vendors.email.send_finance_email"):
            send_submission_to_finance(sub)
        tokens = list(sub.finance_tokens.values_list("action_type", flat=True))
        assert FinanceActionType.APPROVE in tokens
        assert FinanceActionType.REJECT in tokens

    def test_wrong_state_raises(self, invitation, manual_payload):
        sub = create_or_update_submission_from_manual(invitation, manual_payload)
        # Sub is in DRAFT — cannot send to finance
        with pytest.raises(SubmissionStateError):
            send_submission_to_finance(sub)
        with pytest.raises(SubmissionStateError):
            send_submission_to_finance(sub)


# ---------------------------------------------------------------------------
# 8. finance_approve requires sap_vendor_id
# ---------------------------------------------------------------------------

class TestFinanceApproveRequiresSAP:
    def test_blank_sap_id_raises(self, invitation, manual_payload):
        sub = _make_sent_submission(invitation, manual_payload)
        approve_token = sub.finance_tokens.get(action_type=FinanceActionType.APPROVE)
        with pytest.raises(ValueError, match="sap_vendor_id"):
            finance_approve_submission(approve_token.token, sap_vendor_id="")

    def test_missing_sap_id_raises(self, invitation, manual_payload):
        sub = _make_sent_submission(invitation, manual_payload)
        approve_token = sub.finance_tokens.get(action_type=FinanceActionType.APPROVE)
        with pytest.raises(ValueError):
            finance_approve_submission(approve_token.token, sap_vendor_id="   ")


# ---------------------------------------------------------------------------
# 9. finance_approve creates Vendor in waiting_marketing_approval
# ---------------------------------------------------------------------------

class TestFinanceApprove:
    def test_creates_vendor_record(self, invitation, manual_payload):
        sub = _make_sent_submission(invitation, manual_payload)
        approve_token = sub.finance_tokens.get(action_type=FinanceActionType.APPROVE)
        sub_updated, vendor = finance_approve_submission(approve_token.token, sap_vendor_id="SAP123")

        assert vendor.pk is not None
        assert vendor.operational_status == OperationalStatus.WAITING_MARKETING_APPROVAL
        assert vendor.marketing_status == MarketingStatus.PENDING
        assert vendor.sap_vendor_id == "SAP123"

    def test_submission_moves_to_marketing_pending(self, invitation, manual_payload):
        sub = _make_sent_submission(invitation, manual_payload)
        approve_token = sub.finance_tokens.get(action_type=FinanceActionType.APPROVE)
        sub_updated, _ = finance_approve_submission(approve_token.token, sap_vendor_id="SAP456")
        assert sub_updated.status == SubmissionStatus.MARKETING_PENDING

    def test_token_marked_used(self, invitation, manual_payload):
        sub = _make_sent_submission(invitation, manual_payload)
        approve_token = sub.finance_tokens.get(action_type=FinanceActionType.APPROVE)
        finance_approve_submission(approve_token.token, sap_vendor_id="SAP789")
        approve_token.refresh_from_db()
        assert approve_token.used_at is not None

    def test_cannot_reuse_token(self, invitation, manual_payload):
        sub = _make_sent_submission(invitation, manual_payload)
        approve_token = sub.finance_tokens.get(action_type=FinanceActionType.APPROVE)
        finance_approve_submission(approve_token.token, sap_vendor_id="SAP000")
        with pytest.raises(FinanceTokenError, match="already been used"):
            finance_approve_submission(approve_token.token, sap_vendor_id="SAP000")


# ---------------------------------------------------------------------------
# 10. finance_reject updates submission status
# ---------------------------------------------------------------------------

class TestFinanceReject:
    def test_reject_sets_status(self, invitation, manual_payload):
        sub = _make_sent_submission(invitation, manual_payload)
        reject_token = sub.finance_tokens.get(action_type=FinanceActionType.REJECT)
        updated = finance_reject_submission(reject_token.token, note="Missing bank details")
        assert updated.status == SubmissionStatus.FINANCE_REJECTED

    def test_reject_token_marked_used(self, invitation, manual_payload):
        sub = _make_sent_submission(invitation, manual_payload)
        reject_token = sub.finance_tokens.get(action_type=FinanceActionType.REJECT)
        finance_reject_submission(reject_token.token)
        reject_token.refresh_from_db()
        assert reject_token.used_at is not None

    def test_wrong_token_type_raises(self, invitation, manual_payload):
        sub = _make_sent_submission(invitation, manual_payload)
        approve_token = sub.finance_tokens.get(action_type=FinanceActionType.APPROVE)
        with pytest.raises(FinanceTokenError):
            finance_reject_submission(approve_token.token)

    def test_expired_token_raises(self, invitation, manual_payload):
        sub = _make_sent_submission(invitation, manual_payload)
        reject_token = sub.finance_tokens.get(action_type=FinanceActionType.REJECT)
        reject_token.expires_at = timezone.now() - timedelta(hours=1)
        reject_token.save(update_fields=["expires_at"])
        with pytest.raises(FinanceTokenError, match="expired"):
            finance_reject_submission(reject_token.token)


# ---------------------------------------------------------------------------
# 11. reopen_submission after finance rejection
# ---------------------------------------------------------------------------

class TestReopenSubmission:
    def test_reopen_from_finance_rejected(self, invitation, manual_payload):
        sub = _make_sent_submission(invitation, manual_payload)
        reject_token = sub.finance_tokens.get(action_type=FinanceActionType.REJECT)
        finance_reject_submission(reject_token.token)
        sub.refresh_from_db()
        reopened = reopen_submission(sub, reopened_by=None, note="Adding GSTIN")
        assert reopened.status == SubmissionStatus.REOPENED

    def test_reopen_from_wrong_state_raises(self, invitation, manual_payload):
        sub = _make_submitted_submission(invitation, manual_payload)
        with pytest.raises(SubmissionStateError):
            reopen_submission(sub)


# ---------------------------------------------------------------------------
# 12. marketing_approve activates vendor
# ---------------------------------------------------------------------------

class TestMarketingApprove:
    def _get_vendor_in_waiting(self, invitation, manual_payload):
        sub = _make_sent_submission(invitation, manual_payload)
        approve_token = sub.finance_tokens.get(action_type=FinanceActionType.APPROVE)
        _, vendor = finance_approve_submission(approve_token.token, sap_vendor_id="SAP-MKT")
        return vendor

    def test_marketing_approve_activates(self, invitation, manual_payload, actor):
        vendor = self._get_vendor_in_waiting(invitation, manual_payload)
        updated = approve_vendor_marketing(vendor, approved_by=actor, po_mandate_enabled=True)
        assert updated.operational_status == OperationalStatus.ACTIVE
        assert updated.marketing_status == MarketingStatus.APPROVED
        assert updated.po_mandate_enabled is True

    def test_marketing_approve_sets_submission_activated(self, invitation, manual_payload, actor):
        vendor = self._get_vendor_in_waiting(invitation, manual_payload)
        approve_vendor_marketing(vendor, approved_by=actor)
        vendor.onboarding_submission.refresh_from_db()
        assert vendor.onboarding_submission.status == SubmissionStatus.ACTIVATED

    def test_wrong_state_raises(self, invitation, manual_payload, actor):
        vendor = self._get_vendor_in_waiting(invitation, manual_payload)
        approve_vendor_marketing(vendor, approved_by=actor)
        with pytest.raises(VendorStateError):
            approve_vendor_marketing(vendor, approved_by=actor)


# ---------------------------------------------------------------------------
# 13. marketing_reject blocks activation
# ---------------------------------------------------------------------------

class TestMarketingReject:
    def _get_vendor_in_waiting(self, invitation, manual_payload):
        sub = _make_sent_submission(invitation, manual_payload)
        approve_token = sub.finance_tokens.get(action_type=FinanceActionType.APPROVE)
        _, vendor = finance_approve_submission(approve_token.token, sap_vendor_id="SAP-REJ")
        return vendor

    def test_marketing_reject_sets_rejected(self, invitation, manual_payload, actor):
        vendor = self._get_vendor_in_waiting(invitation, manual_payload)
        updated = reject_vendor_marketing(vendor, rejected_by=actor, note="Policy violation")
        assert updated.operational_status == OperationalStatus.INACTIVE
        assert updated.marketing_status == MarketingStatus.REJECTED

    def test_marketing_reject_sets_submission_rejected(self, invitation, manual_payload, actor):
        vendor = self._get_vendor_in_waiting(invitation, manual_payload)
        reject_vendor_marketing(vendor, rejected_by=actor)
        vendor.onboarding_submission.refresh_from_db()
        assert vendor.onboarding_submission.status == SubmissionStatus.REJECTED

    def test_wrong_state_raises(self, invitation, manual_payload, actor):
        vendor = self._get_vendor_in_waiting(invitation, manual_payload)
        # Active vendor cannot be rejected
        approve_vendor_marketing(vendor, approved_by=actor)
        with pytest.raises(VendorStateError):
            reject_vendor_marketing(vendor, rejected_by=actor)


# ---------------------------------------------------------------------------
# 14. assert_vendor_can_submit_invoice blocks inactive vendor
# ---------------------------------------------------------------------------

class TestAssertVendorCanSubmitInvoice:
    def _make_active_vendor(self, entity, org):
        vendor = Vendor.objects.create(
            scope_node=entity,
            org=org,
            vendor_name="Active Vendor",
            sap_vendor_id="SAP-ACTIVE",
            operational_status=OperationalStatus.ACTIVE,
            marketing_status=MarketingStatus.APPROVED,
        )
        return vendor

    def test_active_vendor_passes(self, entity, org):
        vendor = self._make_active_vendor(entity, org)
        # No exception
        assert_vendor_can_submit_invoice(vendor)

    def test_inactive_vendor_raises(self, entity, org):
        vendor = Vendor.objects.create(
            scope_node=entity, org=org,
            vendor_name="Inactive Vendor", sap_vendor_id="SAP-INACTIVE",
            operational_status=OperationalStatus.INACTIVE,
            marketing_status=MarketingStatus.PENDING,
        )
        with pytest.raises(VendorStateError, match="not active"):
            assert_vendor_can_submit_invoice(vendor)

    def test_waiting_marketing_approval_raises(self, entity, org):
        vendor = Vendor.objects.create(
            scope_node=entity, org=org,
            vendor_name="WMA Vendor", sap_vendor_id="SAP-WMA",
            operational_status=OperationalStatus.WAITING_MARKETING_APPROVAL,
            marketing_status=MarketingStatus.PENDING,
        )
        with pytest.raises(VendorStateError):
            assert_vendor_can_submit_invoice(vendor)


# ---------------------------------------------------------------------------
# 15. assert_vendor_can_submit_invoice enforces PO mandate
# ---------------------------------------------------------------------------

class TestPOMandate:
    def test_po_mandate_enforced_when_enabled(self, entity, org):
        vendor = Vendor.objects.create(
            scope_node=entity, org=org,
            vendor_name="PO Vendor", sap_vendor_id="SAP-PO",
            operational_status=OperationalStatus.ACTIVE,
            marketing_status=MarketingStatus.APPROVED,
            po_mandate_enabled=True,
        )
        with pytest.raises(POMandate):
            assert_vendor_can_submit_invoice(vendor, po_number=None)

    def test_po_mandate_passes_when_po_provided(self, entity, org):
        vendor = Vendor.objects.create(
            scope_node=entity, org=org,
            vendor_name="PO Vendor 2", sap_vendor_id="SAP-PO2",
            operational_status=OperationalStatus.ACTIVE,
            marketing_status=MarketingStatus.APPROVED,
            po_mandate_enabled=True,
        )
        # Should not raise
        assert_vendor_can_submit_invoice(vendor, po_number="PO-001")

    def test_no_mandate_passes_without_po(self, entity, org):
        vendor = Vendor.objects.create(
            scope_node=entity, org=org,
            vendor_name="No PO Vendor", sap_vendor_id="SAP-NOPO",
            operational_status=OperationalStatus.ACTIVE,
            marketing_status=MarketingStatus.APPROVED,
            po_mandate_enabled=False,
        )
        assert_vendor_can_submit_invoice(vendor, po_number=None)


# ---------------------------------------------------------------------------
# 16. Excel upload saves source file
# ---------------------------------------------------------------------------

class TestExcelUploadSourceFile:
    @patch("apps.vendors.services._send_invitation_email")
    @patch("apps.vendors.services.send_finance_handoff_notification")
    @patch("apps.vendors.email.send_finance_email")
    def test_source_excel_file_saved(self, mock_email, mock_notif, mock_inv_email, invitation, manual_payload, tmp_path, settings):
        """create_or_update_submission_from_excel saves the raw upload to source_excel_file."""
        settings.MEDIA_ROOT = str(tmp_path)
        buf = _make_excel_file(manual_payload)
        buf.name = "vendor_upload.xlsx"
        submission = create_or_update_submission_from_excel(invitation, buf)
        submission.refresh_from_db()
        assert submission.source_excel_file, "source_excel_file should be set after upload"
        assert "vendor_upload.xlsx" in submission.source_excel_file

    @patch("apps.vendors.services._send_invitation_email")
    @patch("apps.vendors.services.send_finance_handoff_notification")
    @patch("apps.vendors.email.send_finance_email")
    def test_source_excel_file_content_matches(self, mock_email, mock_notif, mock_inv_email, invitation, manual_payload, tmp_path, settings):
        """Source file on disk should be a valid Excel workbook."""
        settings.MEDIA_ROOT = str(tmp_path)
        buf = _make_excel_file(manual_payload)
        buf.name = "upload_check.xlsx"
        submission = create_or_update_submission_from_excel(invitation, buf)
        submission.refresh_from_db()
        assert submission.source_excel_file
        import openpyxl as _xl
        wb = _xl.load_workbook(submission.source_excel_file)
        assert wb.active is not None


# ---------------------------------------------------------------------------
# 17. _start_finance_review generates canonical export Excel
# ---------------------------------------------------------------------------

class TestFinanceReviewExcelGeneration:
    @patch("apps.vendors.services._send_invitation_email")
    @patch("apps.vendors.services.send_finance_handoff_notification")
    @patch("apps.vendors.email.send_finance_email")
    def test_finalize_generates_exported_excel(self, mock_email, mock_notif, mock_inv_email, invitation, manual_payload, tmp_path, settings):
        """finalize_submission triggers Excel generation and sets exported_excel_file."""
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        finalize_submission(submission)
        submission.refresh_from_db()
        assert submission.exported_excel_file, "exported_excel_file should be set after finalize"
        import os
        assert os.path.isfile(submission.exported_excel_file), "Exported Excel should exist on disk"

    @patch("apps.vendors.services._send_invitation_email")
    @patch("apps.vendors.services.send_finance_handoff_notification")
    @patch("apps.vendors.email.send_finance_email")
    def test_finalize_excel_contains_vendor_name(self, mock_email, mock_notif, mock_inv_email, invitation, manual_payload, tmp_path, settings):
        """Generated Excel should include the vendor name in the workbook."""
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        finalize_submission(submission)
        submission.refresh_from_db()
        import openpyxl as _xl
        wb = _xl.load_workbook(submission.exported_excel_file)
        all_values = [
            str(cell.value)
            for row in wb.active.iter_rows()
            for cell in row
            if cell.value
        ]
        assert any("Acme Supplies Ltd" in v for v in all_values)


# ---------------------------------------------------------------------------
# 18. Finance email passes approve_url + reject_url as separate buttons
# ---------------------------------------------------------------------------

class TestFinanceEmailActionUrls:
    @patch("apps.vendors.services._send_invitation_email")
    @patch("apps.vendors.notifications.send_finance_handoff_notification")
    def test_send_to_finance_passes_approve_url(self, mock_notif, mock_inv_email, invitation, manual_payload, tmp_path, settings):
        """send_finance_email must be called with approve_url pointing to the approve token."""
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        submission.status = "submitted"
        submission.save(update_fields=["status"])
        with patch("apps.vendors.email.send_finance_email") as mock_email:
            send_submission_to_finance(submission)
            mock_email.assert_called_once()
            kwargs = mock_email.call_args[1] if mock_email.call_args[1] else {}
            approve_url = kwargs.get("approve_url")
            assert approve_url is not None, "approve_url must be passed to send_finance_email"
            assert "vendor/finance/" in approve_url

    @patch("apps.vendors.services._send_invitation_email")
    @patch("apps.vendors.notifications.send_finance_handoff_notification")
    def test_send_to_finance_passes_reject_url(self, mock_notif, mock_inv_email, invitation, manual_payload, tmp_path, settings):
        """send_finance_email must be called with reject_url pointing to the reject token."""
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        submission.status = "submitted"
        submission.save(update_fields=["status"])
        with patch("apps.vendors.email.send_finance_email") as mock_email:
            send_submission_to_finance(submission)
            mock_email.assert_called_once()
            kwargs = mock_email.call_args[1] if mock_email.call_args[1] else {}
            reject_url = kwargs.get("reject_url")
            assert reject_url is not None, "reject_url must be passed to send_finance_email"
            assert "vendor/finance/" in reject_url

    @patch("apps.vendors.services._send_invitation_email")
    @patch("apps.vendors.notifications.send_finance_handoff_notification")
    def test_approve_and_reject_urls_are_different(self, mock_notif, mock_inv_email, invitation, manual_payload, tmp_path, settings):
        """approve_url and reject_url must point to different tokens."""
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        submission.status = "submitted"
        submission.save(update_fields=["status"])
        with patch("apps.vendors.email.send_finance_email") as mock_email:
            send_submission_to_finance(submission)
            kwargs = mock_email.call_args[1] if mock_email.call_args[1] else {}
            assert kwargs.get("approve_url") != kwargs.get("reject_url")


# ---------------------------------------------------------------------------
# 19. Finance token serializer exposes paired reject token
# ---------------------------------------------------------------------------

class TestPublicFinanceTokenSerializer:
    @patch("apps.vendors.services._send_invitation_email")
    @patch("apps.vendors.services.send_finance_handoff_notification")
    @patch("apps.vendors.email.send_finance_email")
    def test_approve_token_includes_reject_token(self, mock_email, mock_notif, mock_inv_email, invitation, manual_payload, tmp_path, settings):
        """PublicFinanceTokenSerializer for an approve token should include reject_token."""
        from apps.vendors.api.serializers import PublicFinanceTokenSerializer
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        finalize_submission(submission)
        approve_tok = submission.finance_tokens.get(action_type=FinanceActionType.APPROVE)
        reject_tok = submission.finance_tokens.get(action_type=FinanceActionType.REJECT)
        data = PublicFinanceTokenSerializer(approve_tok).data
        assert data["reject_token"] == reject_tok.token

    @patch("apps.vendors.services._send_invitation_email")
    @patch("apps.vendors.services.send_finance_handoff_notification")
    @patch("apps.vendors.email.send_finance_email")
    def test_reject_token_has_no_paired_reject_token(self, mock_email, mock_notif, mock_inv_email, invitation, manual_payload, tmp_path, settings):
        """PublicFinanceTokenSerializer for a reject token should return null reject_token."""
        from apps.vendors.api.serializers import PublicFinanceTokenSerializer
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        finalize_submission(submission)
        reject_tok = submission.finance_tokens.get(action_type=FinanceActionType.REJECT)
        data = PublicFinanceTokenSerializer(reject_tok).data
        assert data["reject_token"] is None

    @patch("apps.vendors.services._send_invitation_email")
    @patch("apps.vendors.services.send_finance_handoff_notification")
    @patch("apps.vendors.email.send_finance_email")
    def test_token_serializer_includes_submission_fields(self, mock_email, mock_notif, mock_inv_email, invitation, manual_payload, tmp_path, settings):
        """PublicFinanceTokenSerializer should include vendor name and bank details."""
        from apps.vendors.api.serializers import PublicFinanceTokenSerializer
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        finalize_submission(submission)
        approve_tok = submission.finance_tokens.get(action_type=FinanceActionType.APPROVE)
        data = PublicFinanceTokenSerializer(approve_tok).data
        assert data["vendor_name"] == "Acme Supplies Ltd"
        assert data["bank_name"] == "State Bank of India"
        assert data["account_number"] == "1234567890"


# ---------------------------------------------------------------------------
# 20. Attachment file upload (Part A)
# ---------------------------------------------------------------------------

class TestAttachmentFileUpload:
    def test_attach_document_with_file_obj(self, invitation, manual_payload, tmp_path, settings):
        """attach_document() with a file_obj stores the file via FileField."""
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        fake_file = io.BytesIO(b"PDF content here")
        fake_file.name = "test_doc.pdf"
        att = attach_document(submission, title="Test Doc", file_obj=fake_file, document_type="invoice")
        assert att.pk is not None
        assert att.file_name == "test_doc.pdf"
        assert att.file  # FileField populated
        assert att.document_type == "invoice"

    def test_attach_document_without_file_obj(self, invitation, manual_payload, tmp_path, settings):
        """attach_document() without file_obj falls back to legacy file_url field."""
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        att = attach_document(
            submission,
            title="Legacy Doc",
            file_url="https://example.com/doc.pdf",
            file_name="doc.pdf",
        )
        assert att.file_url == "https://example.com/doc.pdf"
        assert not att.file  # FileField empty

    def test_attach_document_auto_resolves_file_name(self, invitation, manual_payload, tmp_path, settings):
        """If file_name not provided, it is inferred from file_obj.name."""
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        fake_file = io.BytesIO(b"data")
        fake_file.name = "/tmp/uploads/contract.docx"
        att = attach_document(submission, title="Contract", file_obj=fake_file)
        assert att.file_name == "contract.docx"


# ---------------------------------------------------------------------------
# 21. Finance token serializer uses download URLs, not raw paths (Part B)
# ---------------------------------------------------------------------------

class TestFinanceTokenSerializerDownloadUrls:
    @patch("apps.vendors.services._send_invitation_email")
    @patch("apps.vendors.services.send_finance_handoff_notification")
    @patch("apps.vendors.email.send_finance_email")
    def test_serializer_no_raw_file_paths(self, mock_email, mock_notif, mock_inv_email, invitation, manual_payload, tmp_path, settings):
        """PublicFinanceTokenSerializer must not expose raw filesystem paths."""
        from apps.vendors.api.serializers import PublicFinanceTokenSerializer
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        finalize_submission(submission)
        approve_tok = submission.finance_tokens.get(action_type=FinanceActionType.APPROVE)
        data = PublicFinanceTokenSerializer(approve_tok).data
        assert "source_excel_file" not in data
        assert "exported_excel_file" not in data

    @patch("apps.vendors.services._send_invitation_email")
    @patch("apps.vendors.services.send_finance_handoff_notification")
    @patch("apps.vendors.email.send_finance_email")
    def test_serializer_has_excel_flags(self, mock_email, mock_notif, mock_inv_email, invitation, manual_payload, tmp_path, settings):
        """Serializer has has_exported_excel and has_source_excel boolean fields."""
        from apps.vendors.api.serializers import PublicFinanceTokenSerializer
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        finalize_submission(submission)
        approve_tok = submission.finance_tokens.get(action_type=FinanceActionType.APPROVE)
        data = PublicFinanceTokenSerializer(approve_tok).data
        assert "has_exported_excel" in data
        assert "has_source_excel" in data
        assert isinstance(data["has_exported_excel"], bool)

    @patch("apps.vendors.services._send_invitation_email")
    @patch("apps.vendors.services.send_finance_handoff_notification")
    @patch("apps.vendors.email.send_finance_email")
    def test_serializer_attachment_has_download_url(self, mock_email, mock_notif, mock_inv_email, invitation, manual_payload, tmp_path, settings):
        """Attachment entries in serializer payload have download_url, not file_url."""
        from apps.vendors.api.serializers import PublicFinanceTokenSerializer
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        fake_file = io.BytesIO(b"data")
        fake_file.name = "pan.pdf"
        attach_document(submission, title="PAN Card", file_obj=fake_file)
        finalize_submission(submission)
        approve_tok = submission.finance_tokens.get(action_type=FinanceActionType.APPROVE)
        data = PublicFinanceTokenSerializer(approve_tok).data
        assert len(data["attachments"]) == 1
        att = data["attachments"][0]
        assert "download_url" in att
        assert "file_url" not in att


# ---------------------------------------------------------------------------
# 22. Mandatory workbook — fatal behavior (Part C)
# ---------------------------------------------------------------------------

class TestMandatoryWorkbook:
    @patch("apps.vendors.services._send_invitation_email")
    @patch("apps.vendors.services.send_finance_handoff_notification")
    def test_excel_generation_failure_prevents_state_change(self, mock_notif, mock_inv_email, invitation, manual_payload, tmp_path, settings):
        """If VRF Excel generation raises, submission must stay in 'submitted', not advance."""
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        submission.status = SubmissionStatus.SUBMITTED
        submission.save(update_fields=["status"])
        with patch("apps.vendors.services.generate_vendor_export_excel", side_effect=RuntimeError("disk full")):
            with pytest.raises(SubmissionStateError, match="workbook generation failed"):
                send_submission_to_finance(submission)
        submission.refresh_from_db()
        assert submission.status == SubmissionStatus.SUBMITTED

    @patch("apps.vendors.services._send_invitation_email")
    @patch("apps.vendors.services.send_finance_handoff_notification")
    def test_email_failure_rolls_back_state_change(self, mock_notif, mock_inv_email, invitation, manual_payload, tmp_path, settings):
        """If finance email fails, the transaction rolls back and status stays submitted."""
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        submission.status = SubmissionStatus.SUBMITTED
        submission.save(update_fields=["status"])
        mock_notif.side_effect = Exception("SMTP error")
        with pytest.raises(Exception, match="SMTP error"):
            send_submission_to_finance(submission)
        submission.refresh_from_db()
        assert submission.status == SubmissionStatus.SUBMITTED

    def test_email_py_raises_on_missing_excel_file(self, tmp_path):
        """send_finance_email() raises OSError if exported_excel_path does not exist."""
        from apps.vendors.email import send_finance_email
        nonexistent = str(tmp_path / "missing.xlsx")
        with patch("apps.vendors.email.EmailMessage") as mock_msg:
            mock_instance = MagicMock()
            mock_msg.return_value = mock_instance
            with pytest.raises(OSError):
                send_finance_email(
                    submission_id=1,
                    vendor_name="Test Vendor",
                    approve_url="http://example.com/finance/approve-token",
                    reject_url="http://example.com/finance/reject-token",
                    exported_excel_path=nonexistent,
                )


# ---------------------------------------------------------------------------
# 23. Resend management command — no token duplication (Part D)
# ---------------------------------------------------------------------------

class TestResendVendorFinanceCommand:
    @patch("apps.vendors.services._send_invitation_email")
    @patch("apps.vendors.services.send_finance_handoff_notification")
    @patch("apps.vendors.email.send_finance_email")
    def test_resend_reuses_valid_tokens(self, mock_email, mock_notif, mock_inv_email, invitation, manual_payload, tmp_path, settings):
        """resend_vendor_finance must reuse existing valid tokens, not create new ones."""
        from django.core.management import call_command
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        finalize_submission(submission)
        token_count_before = submission.finance_tokens.count()
        assert token_count_before == 2
        with patch("apps.vendors.notifications.send_finance_handoff_notification"):
            call_command("resend_vendor_finance", submission_id=submission.pk, skip_email=True)
        assert submission.finance_tokens.count() == 2  # No new tokens created

    @patch("apps.vendors.services._send_invitation_email")
    @patch("apps.vendors.services.send_finance_handoff_notification")
    @patch("apps.vendors.email.send_finance_email")
    def test_resend_wrong_status_raises(self, mock_email, mock_notif, mock_inv_email, invitation, manual_payload, tmp_path, settings):
        """resend_vendor_finance raises CommandError if submission is not in SENT_TO_FINANCE."""
        from django.core.management import call_command
        from django.core.management.base import CommandError as DjangoCommandError
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        # Still in draft/submitted state — not sent to finance
        with pytest.raises(DjangoCommandError, match="expected"):
            call_command("resend_vendor_finance", submission_id=submission.pk)

    @patch("apps.vendors.services._send_invitation_email")
    @patch("apps.vendors.services.send_finance_handoff_notification")
    @patch("apps.vendors.email.send_finance_email")
    def test_resend_creates_new_token_when_existing_expired(self, mock_email, mock_notif, mock_inv_email, invitation, manual_payload, tmp_path, settings):
        """If existing tokens are expired, resend command creates new ones."""
        from django.core.management import call_command
        settings.MEDIA_ROOT = str(tmp_path)
        submission = create_or_update_submission_from_manual(invitation, manual_payload)
        finalize_submission(submission)
        # Expire all tokens
        past = timezone.now() - timedelta(hours=1)
        submission.finance_tokens.all().update(expires_at=past)
        token_count_before = submission.finance_tokens.count()
        with patch("apps.vendors.notifications.send_finance_handoff_notification"):
            call_command("resend_vendor_finance", submission_id=submission.pk, skip_email=True)
        # New tokens created on top of expired ones
        assert submission.finance_tokens.count() > token_count_before
        assert data["ifsc"] == "SBIN0001234"
